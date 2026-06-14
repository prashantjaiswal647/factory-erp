import os
import subprocess
import sys
from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from db import Base
from models import (
    BillPayment,
    Customer,
    Factory,
    FinalProductStock,
    InvoiceDocument,
    OutstandingBill,
    PaymentCollection,
)
from services.master_backup import (
    RestoreFailure,
    SHEETS,
    build_master_backup,
    build_validation_report,
    restore_staged_backup,
    stage_backup,
    staged_backup_metadata_path,
    validate_backup,
)


def test_master_backup_import_uses_override_without_creating_directory(tmp_path):
    backup_root = tmp_path / "docker-storage" / "backups"
    env = os.environ.copy()
    env["BACKUP_ROOT"] = str(backup_root)
    api_root = Path(__file__).resolve().parents[1]

    result = subprocess.run(
        [
            sys.executable,
            "-c",
            "import services.master_backup as module; print(module.BACKUP_ROOT)",
        ],
        cwd=api_root,
        env=env,
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    assert result.stdout.strip() == str(backup_root)
    assert not backup_root.exists()


@pytest.fixture()
def backup_db(tmp_path, monkeypatch):
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine, autocommit=False, autoflush=False)()
    factory = Factory(id=1, name="Backup Factory", factory_name="Backup Factory", subscription_status="active")
    customer = Customer(id=1, factory_id=1, name="Raju Traders", phone_number="9999999999", previous_due=0, total_due=100)
    invoice = InvoiceDocument(
        id=1, factory_id=1, customer_id=1, invoice_number="INV-1", invoice_date=date(2026, 6, 12),
        customer_name="Raju Traders", payment_method="Cash", bill_total=100, amount_paid=0,
        customer_total_due=100, payload_json={"items": [{"product_size_ml": 210, "quantity": 10}]},
    )
    outstanding = OutstandingBill(
        id=1, factory_id=1, customer_id=1, invoice_document_id=1, source_type="invoice",
        tracking_number="INV-1", bill_date=date(2026, 6, 12), bill_amount=100, amount_paid=0,
        balance_amount=100, status="active",
    )
    db.add_all([factory, customer, invoice, outstanding])
    db.commit()
    monkeypatch.setattr("services.master_backup.BACKUP_ROOT", tmp_path / "backups")
    monkeypatch.setattr("services.master_backup.STAGING_ROOT", tmp_path / "staging")
    try:
        yield db
    finally:
        db.close()
        engine.dispose()


def test_export_contains_all_master_sheets(backup_db):
    workbook = load_workbook(build_master_backup(backup_db, 1), read_only=True)
    assert set(SHEETS).issubset(workbook.sheetnames)
    assert {"Backup Metadata", "Invoice Items", "Payment Allocations"}.issubset(workbook.sheetnames)


def test_corrupt_and_cross_factory_backup_rejected(backup_db):
    corrupt = validate_backup(b"not an excel file", 1)
    assert corrupt["fatal"] is True
    backup = build_master_backup(backup_db, 1).getvalue()
    cross_factory = validate_backup(backup, 2)
    assert cross_factory["fatal"] is True
    assert "Cross-factory" in cross_factory["errors"][0]["error"]
    report_workbook = load_workbook(build_validation_report(backup, 2), read_only=True)
    assert {"Validation Summary", "Validation Errors", "Sheet Counts"}.issubset(report_workbook.sheetnames)


def test_restore_same_backup_twice_is_idempotent(backup_db):
    backup = build_master_backup(backup_db, 1).getvalue()
    backup_db.query(OutstandingBill).delete()
    backup_db.query(InvoiceDocument).delete()
    backup_db.query(Customer).delete()
    backup_db.commit()

    first_id, first_report = stage_backup(backup, 1)
    assert first_report["fatal"] is False
    first = restore_staged_backup(backup_db, 1, first_id)
    assert first["inserted"] >= 3
    assert backup_db.query(Customer).count() == 1
    assert backup_db.query(InvoiceDocument).count() == 1
    assert backup_db.query(OutstandingBill).count() == 1

    second_id, second_report = stage_backup(backup, 1)
    assert second_report["fatal"] is False
    second = restore_staged_backup(backup_db, 1, second_id)
    assert second["updated"] >= 3
    assert backup_db.query(Customer).count() == 1
    assert backup_db.query(InvoiceDocument).count() == 1
    assert backup_db.query(OutstandingBill).count() == 1


def test_validated_session_confirm_restore_succeeds_and_keeps_filename(backup_db):
    backup = build_master_backup(backup_db, 1).getvalue()
    restore_id, report = stage_backup(backup, 1, "factory-master-backup.xlsx")

    assert report["fatal"] is False
    assert staged_backup_metadata_path(1, restore_id).read_text(encoding="utf-8").find(
        "factory-master-backup.xlsx"
    ) >= 0
    result = restore_staged_backup(backup_db, 1, restore_id)

    assert result["updated"] >= 3
    assert not staged_backup_metadata_path(1, restore_id).exists()


def test_restore_ignores_read_me_and_restore_mapping_sheets(backup_db):
    backup = build_master_backup(backup_db, 1)
    workbook = load_workbook(backup)
    workbook.create_sheet("READ_ME").append(["Documentation only"])
    workbook.create_sheet("RESTORE_MAPPING").append(["source", "target"])
    output = BytesIO()
    workbook.save(output)
    restore_id, report = stage_backup(output.getvalue(), 1, "formatted-master-backup.xlsx")

    assert report["fatal"] is False
    result = restore_staged_backup(backup_db, 1, restore_id)
    assert result["updated"] >= 3


def test_restore_replaces_snapshot_and_stock_is_not_additive(backup_db):
    stock = FinalProductStock(
        factory_id=1,
        product_size_ml=210,
        variety="Plain",
        packaging_size_name="210 ML Plain",
        current_quantity=500,
        total_boxes=10,
        loose_packets=5,
        packets_per_box_limit=50,
    )
    backup_db.add(stock)
    backup_db.commit()
    backup = build_master_backup(backup_db, 1).getvalue()

    stock.current_quantity = 900
    stock.total_boxes = 18
    backup_db.add(Customer(factory_id=1, name="Created Later", phone_number="8888888888"))
    backup_db.commit()

    restore_id, report = stage_backup(backup, 1)
    assert report["fatal"] is False
    result = restore_staged_backup(backup_db, 1, restore_id)

    restored_stock = backup_db.query(FinalProductStock).one()
    assert restored_stock.current_quantity == 500
    assert restored_stock.total_boxes == 10
    assert backup_db.query(Customer).filter(Customer.name == "Created Later").count() == 0
    assert result["deleted"] >= 1


def test_invoice_payment_history_restore_does_not_change_finished_stock(backup_db):
    stock = FinalProductStock(
        factory_id=1,
        product_size_ml=210,
        variety="Plain",
        packaging_size_name="210 ML Plain",
        current_quantity=500,
        total_boxes=10,
        loose_packets=0,
        packets_per_box_limit=50,
    )
    payment = PaymentCollection(
        id=1,
        factory_id=1,
        customer_id=1,
        outstanding_bill_id=1,
        amount_collected=25,
        payment_mode="Cash",
        collection_date=date(2026, 6, 13),
    )
    allocation = BillPayment(
        id=1,
        factory_id=1,
        bill_id=1,
        amount_allocated=25,
        payment_date=date(2026, 6, 13),
    )
    backup_db.add_all([stock, payment, allocation])
    backup_db.commit()
    backup = build_master_backup(backup_db, 1).getvalue()

    stock.current_quantity = 700
    backup_db.commit()
    restore_id, report = stage_backup(backup, 1)
    assert report["fatal"] is False
    restore_staged_backup(backup_db, 1, restore_id)

    assert backup_db.query(FinalProductStock).one().current_quantity == 500
    assert backup_db.query(InvoiceDocument).count() == 1
    assert backup_db.query(PaymentCollection).count() == 1
    assert backup_db.query(BillPayment).count() == 1
    assert backup_db.query(OutstandingBill).count() == 1


def test_restore_rolls_back_all_changes_on_fatal_error(backup_db, monkeypatch):
    backup = build_master_backup(backup_db, 1).getvalue()
    customer = backup_db.query(Customer).one()
    customer.name = "Current Database Value"
    backup_db.commit()
    restore_id, report = stage_backup(backup, 1)
    assert report["fatal"] is False

    original_coerce = __import__("services.master_backup", fromlist=["_coerce"])._coerce
    calls = {"count": 0}

    def fail_after_mutation(column, value):
        calls["count"] += 1
        if calls["count"] > 4:
            raise RuntimeError("forced restore failure")
        return original_coerce(column, value)

    monkeypatch.setattr("services.master_backup._coerce", fail_after_mutation)
    with pytest.raises(RestoreFailure, match="could not be applied") as caught:
        restore_staged_backup(backup_db, 1, restore_id)

    assert "Customers" in caught.value.detail
    backup_db.expire_all()
    assert backup_db.query(Customer).one().name == "Current Database Value"


def test_api_runtime_installs_pg_dump_client():
    dockerfile = Path(__file__).resolve().parents[1] / "Dockerfile"
    content = dockerfile.read_text(encoding="utf-8")
    assert "postgresql-client" in content
