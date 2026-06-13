from io import BytesIO
from decimal import Decimal
from types import SimpleNamespace

from openpyxl import Workbook
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from db import Base
from auth import get_current_user
from main import app
from models import BlankStock, Customer, Factory, FinalProductStock, Machine, Worker
from routers.onboarding import (
    apply_bulk_rows,
    build_owner_onboarding_workbook,
    dedupe_valid_bulk_rows,
    read_master_bulk_excel,
    validate_bulk_cross_sheet,
)
from services.bulk_validation import enrich_failed_rows


def owner_workbook() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {
        "Factory_Profile": (
            ["Factory Name", "GST Number", "Factory Address"],
            ["Owner Factory", "07ABCDE1234F1Z5", "Delhi"],
        ),
        "Customers": (
            ["Customer Name", "Phone", "Firm Name", "Place"],
            ["Raj", "9876543210", "Raj Traders", "Delhi"],
        ),
        "Workers": (
            ["Worker Name", "Mobile", "Daily Wages", "Duty Hours"],
            ["Raju", "9876500001", 500, 8],
        ),
        "Machines": (
            ["Machine Number", "Machine Name", "Machine Size ML", "Bottom Size MM"],
            ["M-1", "Machine 210", 210, 68],
        ),
        "Cup_Blank": (
            ["Material Name", "Cup Size ML", "Design", "Linked Bottom Size MM", "Weight Per Bora KG", "Total Boras Sacks"],
            ["Cup Blank", 210, "White", 68, 40, 10],
        ),
        "Bottom_Reel": (
            ["Bottom Size MM", "Design", "Total Individual Rolls", "Total Weight KG"],
            [68, "White", 20, 100],
        ),
        "Box_Stock": (
            ["Carton Type", "Carton Quantity", "Price Per Box Rs"],
            ["210 White Carton", 50, 20],
        ),
        "Plastic_Stock": (
            ["Plastic Size Type", "Cup Size ML", "Total Boras Sacks", "Weight Per Bora KG", "Price Per KG Rs"],
            ["210 Sleeve", 210, 5, 20, 120],
        ),
        "Finished_Goods": (
            ["Cup Size ML", "Design", "Carton Type", "Pieces Per Packet", "Packets Per Carton", "Opening Boxes"],
            [210, "White", "210 White Carton", 48, 10, 5],
        ),
        "Costing_Optional": (["Paper Price Per KG"], [100]),
    }
    for name, (headers, values) in sheets.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        sheet.append(values)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_owner_friendly_excel_generates_keys_maps_and_reuploads_idempotently():
    valid, failed = read_master_bulk_excel(owner_workbook())
    assert not [row for row in failed if "error" in row and "defaulted" not in row["error"]]
    assert valid["customer"][0]["customer_restore_key"].startswith("CUS-")
    assert valid["worker"][0]["worker_restore_key"].startswith("WRK-")
    assert valid["machine"][0]["machine_restore_key"].startswith("MAC-")
    assert valid["blank_stock"][0]["material_restore_key"].startswith("MAT-BL-")
    assert valid["bottom_reel"][0]["material_restore_key"].startswith("MAT-BT-")
    assert valid["finished_goods"][0]["product_restore_key"].startswith("SKU-")
    assert validate_bulk_cross_sheet(valid, strict_validation=True) == []

    valid, duplicate_warnings = dedupe_valid_bulk_rows(valid)
    assert duplicate_warnings == []
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autocommit=False, autoflush=False)
    db = session_factory()
    db.add(Factory(id=1, name="Owner Factory"))
    db.commit()
    user = SimpleNamespace(id=1, factory_id=1)
    import_order = (
        "customer", "worker", "machine", "blank_stock", "bottom_reel",
        "box_stock", "plastic_stock", "finished_goods",
    )
    for _ in range(2):
        for sub_tab_type in import_order:
            apply_bulk_rows(db, user, sub_tab_type, valid[sub_tab_type], {})
        db.commit()

    assert db.query(Customer).count() == 1
    assert db.query(Worker).count() == 1
    assert db.query(Machine).count() == 1
    assert db.query(BlankStock).count() == 1
    assert db.query(FinalProductStock).count() == 1
    db.close()
    Base.metadata.drop_all(engine)
    engine.dispose()


def test_downloaded_owner_template_hides_technical_restore_keys():
    from openpyxl import load_workbook

    workbook = load_workbook(build_owner_onboarding_workbook(), read_only=True)
    assert set(workbook.sheetnames) == {
        "Factory_Profile", "Customers", "Workers", "Machines", "Cup_Blank",
        "Bottom_Reel", "Box_Stock", "Plastic_Stock", "Finished_Goods", "Costing_Optional",
    }
    headers = {
        str(value).strip().lower()
        for sheet in workbook.worksheets
        for value in next(sheet.iter_rows(values_only=True))
        if value
    }
    assert not any("restore_key" in header for header in headers)


def test_bulk_upload_decimal_validation_error_returns_422_not_500():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Factory_Profile"
    sheet.append(["Factory Name"])
    sheet.append(["Owner Factory"])
    for name in (
        "Customers", "Workers", "Machines", "Cup_Blank", "Bottom_Reel",
        "Box_Stock", "Plastic_Stock", "Finished_Goods", "Costing_Optional",
    ):
        workbook.create_sheet(name)
    cup_blank = workbook["Cup_Blank"]
    cup_blank.append([
        "Material Name", "Cup Size ML", "Design", "Linked Bottom Size MM",
        "Weight Per Bora KG", "Total Boras Sacks",
    ])
    cup_blank.append(["Blank", 210, "White", 0, Decimal("40.50"), Decimal("2.25")])
    output = BytesIO()
    workbook.save(output)

    app.dependency_overrides[get_current_user] = lambda: SimpleNamespace(
        id=1, factory_id=1, role="Owner", username="owner", full_name="Owner"
    )
    try:
        response = TestClient(app, raise_server_exceptions=False).post(
            "/api/v1/onboarding/bulk-upload/master",
            files={
                "file": (
                    "Munshi_AI_Factory_Owner_Onboarding_Template.xlsx",
                    output.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    finally:
        app.dependency_overrides.pop(get_current_user, None)

    assert response.status_code == 422
    assert response.json()["detail"]["overall_status"] == "failed"
    assert "40.5" in response.text or "2.25" in response.text


def test_owner_validation_auto_normalizes_mapping_and_plastic_units():
    workbook = Workbook()
    workbook.remove(workbook.active)
    data = {
        "Factory_Profile": (["Factory Name"], [["Factory"]]),
        "Customers": (["Customer Name", "Phone"], []),
        "Workers": (["Worker Name"], []),
        "Machines": (["Machine Number", "Machine Name", "Machine Size ML", "Bottom Size MM"], [["1", "M1", 210, 68]]),
        "Cup_Blank": (
            ["Material Name", "Cup Size ML", "Design", "Linked Bottom Size MM", "Weight Per Bora KG", "Total Boras Sacks"],
            [["Blank", 210, "", 68, 40, 2]],
        ),
        "Bottom_Reel": (
            ["Bottom Size MM", "Design", "Total Individual Rolls", "Total Weight KG"],
            [[68, "White", 10, 20]],
        ),
        "Box_Stock": (["Carton Type", "Carton Quantity"], [["210 - White", 5]]),
        "Plastic_Stock": (
            ["Plastic Size Type", "Cup Size ML", "Total Boras Sacks", "Weight Per Bora KG", "Price Per KG Rs"],
            [["Sleeve", "210 ml", 2, 20, 100], [None, None, None, None, None]],
        ),
        "Finished_Goods": (
            ["Cup Size ML", "Design", "Carton Type", "Pieces Per Packet", "Packets Per Carton", "Opening Boxes"],
            [[210, "", " 210-White ", 48, 10, 1]],
        ),
        "Costing_Optional": (["Paper Price Per KG"], []),
    }
    for name, (headers, rows) in data.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)

    valid, failed = read_master_bulk_excel(output.getvalue())
    issues = enrich_failed_rows(failed)
    assert valid["plastic_stock"][0]["used_for_cup_size_ml"] == 210
    assert len(valid["plastic_stock"]) == 1
    assert valid["blank_stock"][0]["variety_design"] == "White"
    assert valid["finished_goods"][0]["variety_design"] == "White"
    assert not [issue for issue in validate_bulk_cross_sheet(valid, strict_validation=True) if issue.severity.value == "fatal"]
    assert all("BulkRow" not in issue.error for issue in issues)


def test_partial_plastic_row_has_simple_error_and_duplicates_are_grouped():
    workbook_bytes = owner_workbook()
    valid, _ = read_master_bulk_excel(workbook_bytes)
    valid["customer"].append({**valid["customer"][0], "_row_number": 9})
    _, duplicate_warnings = dedupe_valid_bulk_rows(valid)
    customer_warnings = [issue for issue in duplicate_warnings if issue.sheet == "Customers"]
    assert len(customer_warnings) == 1
    assert "duplicate rows were combined" in customer_warnings[0].error

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plastic_Stock"
    sheet.append(["Plastic Size Type", "Cup Size ML", "Total Boras Sacks"])
    sheet.append(["Sleeve", "not a size", 2])
    from routers.onboarding import read_owner_friendly_sheet

    warnings: list[dict] = []
    _, errors = read_owner_friendly_sheet(
        __import__("pandas").read_excel(BytesIO(_single_sheet_bytes(workbook)), header=None),
        "plastic_stock",
        "Plastic_Stock",
        warnings,
    )
    issues = enrich_failed_rows(errors)
    assert issues
    assert all("PlasticStockBulkRow" not in issue.error for issue in issues)
    assert any(
        issue.error == "Cup size can be written as 210 or 210,250,300."
        for issue in issues
    )


def test_owner_template_plastic_bottom_and_missing_blank_validation():
    workbook = Workbook()
    workbook.remove(workbook.active)
    data = {
        "Factory_Profile": (["Factory Name"], [["Factory"]]),
        "Customers": (["Customer Name", "Phone"], []),
        "Workers": (["Worker Name"], []),
        "Machines": (
            ["Machine Number", "Machine Name", "Machine Size ML", "Bottom Size MM"],
            [
                ["1", "M1", 65, 65],
                ["2", "M2", 52, 52],
                ["3", "M3", 48, 48],
                ["4", "M4", 75, 75],
                ["5", "M5", 57, 57],
            ],
        ),
        "Cup_Blank": (
            ["Material Name", "Cup Size ML", "Design", "Linked Bottom Size MM", "Weight Per Bora KG"],
            [
                ["Blank 65", 65, "White", 65, 40],
                ["Blank 52", 52, "White", 52, 40],
                ["Blank 48", 48, "White", 48, 40],
                ["Blank 75", 75, "White", 75, 40],
                ["Blank 57", 57, "White", 57, 40],
            ],
        ),
        "Bottom_Reel": (
            ["Bottom Size MM", "Opening Rolls", "Total Weight KG"],
            [
                [65, 10, 20],
                [52, 10, 20],
                [48, 10, 20],
                [75, 10, 20],
                [57, 10, 20],
            ],
        ),
        "Box_Stock": (["Carton Type", "Carton Quantity"], [["65 White", 5]]),
        "Plastic_Stock": (
            ["Plastic Type", "Used For Cup Sizes ML", "Total Boras Sacks"],
            [
                ["Sleeve A", 210, 1],
                ["Sleeve B", "210 ml", 1],
                ["Sleeve C", "210,250,300", 1],
                ["Sleeve D", "55,65", 1],
            ],
        ),
        "Finished_Goods": (
            ["Cup Size ML", "Design", "Carton Type", "Pieces Per Packet", "Packets Per Carton"],
            [
                [65, "White", "65 White", 50, 20],
                [65, "black", "65 Black", 50, 20],
                [65, "hot", "65 Hot", 50, 20],
            ],
        ),
        "Costing_Optional": (["Paper Price Per KG"], []),
    }
    for name, (headers, rows) in data.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)

    valid, failed = read_master_bulk_excel(output.getvalue())
    assert [row["used_for_cup_size_ml"] for row in valid["plastic_stock"]] == [
        210, 210, 210, 250, 300, 55, 65,
    ]
    assert valid["bottom_reel"][0]["variety_design"] == "Plain White"
    assert not [row for row in failed if row.get("entity_type") in {"plastic_stock", "bottom_reel"}]
    assert len([
        row
        for row in failed
        if row.get("sheet") == "Bottom_Reel"
        and "auto-defaulted" in row.get("error", "")
    ]) == 5

    fatal = [
        issue
        for issue in validate_bulk_cross_sheet(valid, strict_validation=True)
        if issue.severity.value == "fatal"
    ]
    assert len(fatal) == 2
    assert {issue.suggested_correction for issue in fatal} == {
        "Add Cup Blank row for 65 ml black",
        "Add Cup Blank row for 65 ml hot",
    }


def test_partial_plastic_size_uses_business_validation_message():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plastic_Stock"
    sheet.append(["Plastic Type", "Used For Cup Sizes ML"])
    sheet.append(["Sleeve", "not a cup size"])
    from routers.onboarding import read_owner_friendly_sheet

    rows, errors = read_owner_friendly_sheet(
        __import__("pandas").read_excel(BytesIO(_single_sheet_bytes(workbook)), header=None),
        "plastic_stock",
        "Plastic_Stock",
        [],
    )

    assert rows == []
    assert errors[0]["error"] == "Cup size can be written as 210 or 210,250,300."


def _single_sheet_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
