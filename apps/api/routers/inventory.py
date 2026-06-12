from decimal import Decimal
from datetime import date as date_cls, datetime
from io import BytesIO
from typing import List, Optional, Union

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Response, UploadFile, File, status
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import String, func, or_
from sqlalchemy import func as sql_func
from sqlalchemy.orm import Session
from PIL import Image
from io import BytesIO
import os
import logging

from dependencies import INVENTORY_ROLES, check_permissions
from db import get_db
from services.activity_logger import log_activity
from models import (
    BlankStock,
    BottomStock,
    BoxStock,
    DailyProduction,
    DailySale,
    Factory,
    FinalProductStock,
    FinishedGoodsStock,
    Inventory,
    Machine,
    PackagingProfile,
    PackagingMetrics,
    PlasticStock,
    PolybagStock,
    User,
)


router = APIRouter()
log = logging.getLogger(__name__)


def decimal_or_zero(value) -> Decimal:
    return Decimal(value or 0)


def int_or_zero(value) -> int:
    return int(value or 0)


def calculate_live_sku_stock(
    db: Session,
    factory_id: str,
    product_size_ml: int,
    variety: str,
    packaging_size_name: str,
    onboarding_boxes: int,
    onboarding_loose: int,
    packets_per_box_limit: int,
) -> tuple[int, int]:
    # Sum production made
    prod_boxes = (
        db.query(func.coalesce(func.sum(DailyProduction.total_boxes_made), 0))
        .filter(DailyProduction.factory_id == str(factory_id))
        .filter(DailyProduction.product_size_ml == product_size_ml)
        .filter(func.lower(DailyProduction.variety) == variety.strip().lower())
        .filter(func.lower(DailyProduction.packaging_size_name) == packaging_size_name.strip().lower())
        .filter(DailyProduction.status == "ACTIVE")
        .scalar()
    ) or 0

    prod_loose = (
        db.query(func.coalesce(func.sum(DailyProduction.loose_packets_made), 0))
        .filter(DailyProduction.factory_id == str(factory_id))
        .filter(DailyProduction.product_size_ml == product_size_ml)
        .filter(func.lower(DailyProduction.variety) == variety.strip().lower())
        .filter(func.lower(DailyProduction.packaging_size_name) == packaging_size_name.strip().lower())
        .filter(DailyProduction.status == "ACTIVE")
        .scalar()
    ) or 0

    # Sum sales sold
    sales_boxes = (
        db.query(func.coalesce(func.sum(DailySale.boxes_sold), 0))
        .filter(DailySale.factory_id == str(factory_id))
        .filter(DailySale.product_size_ml == product_size_ml)
        .filter(func.lower(DailySale.variety) == variety.strip().lower())
        .filter(func.lower(DailySale.packaging_size_name) == packaging_size_name.strip().lower())
        .scalar()
    ) or 0

    sales_loose = (
        db.query(func.coalesce(func.sum(DailySale.loose_packets_sold), 0))
        .filter(DailySale.factory_id == str(factory_id))
        .filter(DailySale.product_size_ml == product_size_ml)
        .filter(func.lower(DailySale.variety) == variety.strip().lower())
        .filter(func.lower(DailySale.packaging_size_name) == packaging_size_name.strip().lower())
        .scalar()
    ) or 0

    limit = packets_per_box_limit if packets_per_box_limit > 0 else 1000
    total_packets = (
        (onboarding_boxes * limit + onboarding_loose)
        + (prod_boxes * limit + prod_loose)
        - (sales_boxes * limit + sales_loose)
    )

    live_boxes = total_packets // limit
    live_loose = total_packets % limit
    return int(live_boxes), int(live_loose)


def recalculate_and_sync_sku_stock(
    db: Session,
    factory_id: str,
    product_size_ml: int,
    variety: str,
    packaging_size_name: str,
) -> tuple[int, int]:
    stock = (
        db.query(FinalProductStock)
        .filter(
            FinalProductStock.factory_id == str(factory_id),
            FinalProductStock.product_size_ml == product_size_ml,
            func.lower(FinalProductStock.variety) == variety.strip().lower(),
            func.lower(FinalProductStock.packaging_size_name) == packaging_size_name.strip().lower(),
        )
        .first()
    )

    if not stock:
        return 0, 0

    onboarding_boxes = stock.total_boxes or 0
    onboarding_loose = stock.loose_packets or 0

    live_boxes, live_loose = calculate_live_sku_stock(
        db=db,
        factory_id=str(factory_id),
        product_size_ml=product_size_ml,
        variety=variety,
        packaging_size_name=packaging_size_name,
        onboarding_boxes=onboarding_boxes,
        onboarding_loose=onboarding_loose,
        packets_per_box_limit=stock.packets_per_box_limit or 1000,
    )

    stock.current_quantity = max(live_boxes, 0)
    db.add(stock)

    profile = (
        db.query(PackagingProfile)
        .filter(
            PackagingProfile.factory_id == str(factory_id),
            PackagingProfile.cup_size_ml == product_size_ml,
            func.lower(PackagingProfile.profile_name) == packaging_size_name.strip().lower(),
        )
        .first()
    )

    if profile:
        fg_stock = (
            db.query(FinishedGoodsStock)
            .filter(
                FinishedGoodsStock.factory_id == str(factory_id),
                FinishedGoodsStock.packaging_profile_id == profile.id,
            )
            .first()
        )
        if fg_stock:
            fg_stock.boxes_available = max(live_boxes, 0)
            db.add(fg_stock)

    db.flush()
    return live_boxes, live_loose


class LiveStockRow(BaseModel):
    id: Union[int, str]
    stock_type: str
    item_name: str
    category: Optional[str] = None
    packaging_size: Optional[str] = None
    packaging_size_name: Optional[str] = None
    product_size_ml: Optional[int] = None
    variety: Optional[str] = None
    pieces_per_packet: Optional[int] = None
    packets_per_box: Optional[int] = None
    packets_per_box_limit: Optional[int] = None
    current_quantity: Optional[Decimal] = None
    quantity: Decimal
    unit: str
    price_per_unit: Optional[Decimal] = None
    price_per_box: Optional[Decimal] = None
    price_per_kg: Optional[Decimal] = None
    box_type: Optional[str] = None
    size_ml: Optional[int] = None
    size_mm: Optional[int] = None
    kg_per_sack: Optional[Decimal] = None
    total_weight_kg: Optional[Decimal] = None
    total_rolls: Optional[int] = None
    total_boras: Optional[int] = None
    weight_per_bora_kg: Optional[Decimal] = None
    cup_size_ml: Optional[int] = None
    image_url: Optional[str] = None
    variant_name: Optional[str] = None
    bucket: Optional[str] = None
    quantity_source: Optional[str] = None


class FinalStockRow(BaseModel):
    id: int
    product_size_ml: Optional[int] = None
    variety: Optional[str] = None
    packaging_size: Optional[str] = None
    packaging_size_name: Optional[str] = None
    pieces_per_packet: Optional[int] = None
    packets_per_box: Optional[int] = None
    packets_per_box_limit: Optional[int] = None
    current_quantity: Optional[int] = None
    total_boxes: Optional[int] = None
    loose_packets: Optional[int] = None
    image_url: Optional[str] = None
    variant_name: Optional[str] = None


class FinalStockCreate(BaseModel):
    product_id: Optional[int] = Field(default=None, gt=0)
    product_size_ml: Optional[int] = Field(default=None, gt=0)
    variety: str = "Standard/White"
    packaging_size: Optional[str] = Field(default=None, max_length=100)
    packaging_size_name: Optional[str] = Field(default=None, max_length=100)
    initial_quantity: int = Field(default=0, ge=0)
    current_quantity: Optional[int] = Field(default=None, ge=0)
    total_boxes: Optional[int] = Field(default=None, ge=0)
    loose_packets: int = Field(default=0, ge=0)
    pieces_per_packet: int = Field(default=1, gt=0)
    packets_per_box: Optional[int] = Field(default=None, gt=0)
    packets_per_box_limit: Optional[int] = Field(default=None, gt=0)
    category: Optional[str] = None


class FinishedGoodsOnboardRequest(BaseModel):
    product_size_ml: float
    variety_design: str = Field(..., min_length=1)
    packaging_size_name: Optional[str] = None
    pcs_per_packet: int = Field(..., gt=0)
    packets_per_box: int = Field(..., gt=0)
    initial_quantity_boxes: int = Field(..., ge=0)


@router.get("/")
def list_live_stock_endpoint(
    response: Response,
    current_user: User = Depends(check_permissions(INVENTORY_ROLES)),
    db: Session = Depends(get_db),
):
    return list_live_stock(current_user=current_user, db=db, response=response)


def list_live_stock(
    current_user: User,
    db: Session,
    response: Optional[Response] = None,
):
    try:
        processed_inventory = []
        loop_errors = {
            "inventory": 0,
            "blank_stock": 0,
            "bottom_stock": 0,
            "box_stock": 0,
            "plastic_stock": 0,
            "polybag_stock": 0,
            "final_product_stock": 0,
        }

        def record_row_error(loop_name: str, item, exc: Exception) -> None:
            loop_errors[loop_name] += 1
            if loop_errors[loop_name] <= 3:
                log.warning(
                    "live_stock %s row %s: %s",
                    loop_name,
                    getattr(item, "id", "?"),
                    exc,
                )

        def set_warning_header() -> None:
            if response is not None:
                response.headers["X-Inventory-Warnings"] = ",".join(
                    f"{name}={count}" for name, count in loop_errors.items()
                )
        try:
            factory_id = str(current_user.factory_id)
        except Exception:
            return []

        # 1. Fetch standard Inventory items
        try:
            inventory_items = (
                db.query(Inventory)
                .filter(Inventory.factory_id == factory_id)
                .filter(or_(Inventory.item_name.is_(None), ~Inventory.item_name.like("[DELETED]%")))
                .order_by(Inventory.item_name.asc().nullslast(), Inventory.id.asc())
                .all()
            )
            for item in inventory_items:
                try:
                    quantity = item.quantity if item.quantity is not None else 0
                    processed_inventory.append(
                        {
                            "id": item.id,
                            "factory_id": item.factory_id,
                            "product_id": getattr(item, "product_id", None),
                            "item_name": item.item_name if item.item_name is not None else "Unknown Item",
                            "stock_type": "Inventory",
                            "current_quantity": getattr(item, "current_quantity", None) if getattr(item, "current_quantity", None) is not None else float(quantity),
                            "quantity": float(quantity),
                            "unit": item.unit if item.unit is not None else "pieces",
                            "price_per_unit": float(item.price_per_unit if item.price_per_unit is not None else 0),
                            "packaging_size": item.packaging_size if item.packaging_size is not None else "Standard",
                            "pieces_per_packet": item.pieces_per_packet if item.pieces_per_packet is not None else 0,
                            "packets_per_box": item.packets_per_box if item.packets_per_box is not None else 0,
                            "category": item.category,
                            "bucket": (
                                "polybags_packing"
                                if item.category == "Packaging"
                                else "raw_other"
                                if item.category == "Raw"
                                else "needs_mapping_review"
                            ),
                            "quantity_source": None,
                        }
                    )
                except Exception as exc:
                    record_row_error("inventory", item, exc)
                    continue
        except Exception:
            pass

        # 2. Fetch BlankStock items (Raw Material Tab)
        try:
            blank_items = (
                db.query(BlankStock)
                .filter(BlankStock.factory_id == factory_id)
                .order_by(BlankStock.blank_size_ml.asc())
                .all()
            )
            for item in blank_items:
                try:
                    qty = float(item.total_qty_kg or 0)
                    quantity_source = (
                        "not_recorded"
                        if (
                            qty == 0
                            and float(getattr(item, "weight_per_bora_kg", 0) or 0) > 0
                            and float(getattr(item, "total_boras", 0) or 0) == 0
                        )
                        else "excel_upload"
                    )
                    processed_inventory.append(
                        {
                            "id": f"blank-{item.id}",
                            "factory_id": item.factory_id,
                            "product_id": None,
                            "item_name": f"{getattr(item, 'blank_size_ml', 0)}ml Blank - {getattr(item, 'variety', None) or 'Plain White'}",
                            "stock_type": "Blank",
                            "current_quantity": qty,
                            "quantity": qty,
                            "unit": "kg",
                            "price_per_unit": 0.0,
                            "packaging_size": "Standard",
                            "category": "Blank",
                            "size_ml": getattr(item, "blank_size_ml", None),
                            "kg_per_sack": float(getattr(item, "weight_per_bora_kg", 0) or 0),
                            "bucket": "cup_blanks",
                            "quantity_source": quantity_source,
                        }
                    )
                except Exception as exc:
                    record_row_error("blank_stock", item, exc)
                    continue
        except Exception:
            pass

        # 3. Fetch BottomStock items (Raw Material Tab)
        try:
            bottom_items = (
                db.query(BottomStock)
                .filter(BottomStock.factory_id == factory_id)
                .order_by(BottomStock.bottom_size_mm.asc())
                .all()
            )
            for item in bottom_items:
                try:
                    qty = float(item.total_qty_kg or 0)
                    processed_inventory.append(
                        {
                            "id": f"bottom-{item.id}",
                            "factory_id": item.factory_id,
                            "product_id": None,
                            "item_name": f"{getattr(item, 'bottom_size_mm', 0)}mm Bottom Roll - {getattr(item, 'variety', None) or 'Plain White'}",
                            "stock_type": "Bottom",
                            "current_quantity": qty,
                            "quantity": qty,
                            "unit": "kg",
                            "price_per_unit": 0.0,
                            "packaging_size": "Standard",
                            "category": "Bottom",
                            "size_mm": getattr(item, 'bottom_size_mm', 0),
                            "total_weight_kg": qty,
                            "total_rolls": getattr(item, 'total_rolls', 0) or 0,
                            "bucket": "bottom_reels",
                            "quantity_source": "excel_upload",
                        }
                    )
                except Exception as exc:
                    record_row_error("bottom_stock", item, exc)
                    continue
        except Exception:
            pass

        # 4. Fetch BoxStock items (Packaging Tab)
        try:
            box_items = (
                db.query(BoxStock)
                .filter(BoxStock.factory_id == factory_id)
                .order_by(BoxStock.packaging_size_name.asc())
                .all()
            )
            for item in box_items:
                try:
                    qty = float(getattr(item, "quantity", 0) or 0)
                    box_name = getattr(item, "packaging_size_name", None) or getattr(item, "box_type", None) or "Standard"
                    processed_inventory.append(
                        {
                            "id": f"box-{item.id}",
                            "factory_id": item.factory_id,
                            "product_id": None,
                            "item_name": f"{box_name} Carton Box",
                            "stock_type": "Carton Box",
                            "current_quantity": qty,
                            "quantity": qty,
                            "unit": "pcs",
                            "price_per_unit": float(getattr(item, "price_per_box", 0) or 0),
                            "price_per_box": float(getattr(item, "price_per_box", 0) or 0),
                            "packaging_size": box_name,
                            "packaging_size_name": box_name,
                            "category": "Carton Box",
                            "box_type": getattr(item, "box_type", None) or box_name,
                            "bucket": "boxes",
                            "quantity_source": "excel_upload",
                        }
                    )
                except Exception as exc:
                    record_row_error("box_stock", item, exc)
                    continue
        except Exception:
            pass

        # 5. Fetch PlasticStock items (Packaging Tab)
        try:
            plastic_items = (
                db.query(PlasticStock)
                .filter(PlasticStock.factory_id == factory_id)
                .order_by(PlasticStock.plastic_size_name.asc(), PlasticStock.cup_size_ml.asc())
                .all()
            )
            for item in plastic_items:
                try:
                    qty_kg = float((getattr(item, "total_boras", 0) or 0) * (getattr(item, "weight_per_bora_kg", 0.0) or 0.0))
                    total_boras = getattr(item, "total_boras", 0) or 0
                    weight_per_bora_kg = getattr(item, "weight_per_bora_kg", 0.0) or 0.0
                    price_per_kg = getattr(item, "price_per_kg", 0) or 0
                    processed_inventory.append(
                        {
                            "id": f"plastic-{item.id}",
                            "factory_id": item.factory_id,
                            "product_id": None,
                            "item_name": f"{getattr(item, 'plastic_size_name', 'Standard')} ({getattr(item, 'cup_size_ml', 0)}ml) Plastic",
                            "stock_type": "Polybag",
                            "current_quantity": qty_kg,
                            "quantity": qty_kg,
                            "unit": "kg",
                            "price_per_unit": float(price_per_kg),
                            "price_per_kg": float(price_per_kg),
                            "packaging_size": getattr(item, "plastic_size_name", "Standard"),
                            "category": "Polybag",
                            "cup_size_ml": getattr(item, "cup_size_ml", None),
                            "total_boras": int(total_boras),
                            "weight_per_bora_kg": float(weight_per_bora_kg),
                            "bucket": "polybags_packing",
                            "quantity_source": "excel_upload",
                        }
                    )
                except Exception as exc:
                    record_row_error("plastic_stock", item, exc)
                    continue
        except Exception:
            pass

        # 6. Fetch PolybagStock items
        try:
            polybag_items = (
                db.query(PolybagStock)
                .filter(PolybagStock.factory_id == factory_id)
                .order_by(PolybagStock.packaging_size_name.asc())
                .all()
            )
            for item in polybag_items:
                try:
                    qty = float(getattr(item, "total_packets", 0) or 0)
                    processed_inventory.append(
                        {
                            "id": f"polybag-{item.id}",
                            "factory_id": item.factory_id,
                            "product_id": None,
                            "item_name": f"{getattr(item, 'packaging_size_name', 'Standard')} Polybag",
                            "stock_type": "Polybag",
                            "current_quantity": qty,
                            "quantity": qty,
                            "unit": "packets",
                            "price_per_unit": 0.0,
                            "packaging_size": getattr(item, "packaging_size_name", "Standard"),
                            "category": "Polybag",
                            "bucket": "polybags_packing",
                            "quantity_source": "excel_upload",
                        }
                    )
                except Exception as exc:
                    record_row_error("polybag_stock", item, exc)
                    continue
        except Exception:
            pass

        # 7. Fetch Final Product Stocks
        try:
            final_stock_items = (
                db.query(FinalProductStock)
                .filter(FinalProductStock.factory_id == factory_id)
                .order_by(FinalProductStock.product_size_ml.asc(), FinalProductStock.packaging_size_name.asc())
                .all()
            )
            for item in final_stock_items:
                try:
                    live_boxes, live_loose = calculate_live_sku_stock(
                        db=db,
                        factory_id=factory_id,
                        product_size_ml=getattr(item, "product_size_ml", 0),
                        variety=getattr(item, "variety", "Standard"),
                        packaging_size_name=getattr(item, "packaging_size_name", "Standard"),
                        onboarding_boxes=getattr(item, "total_boxes", 0) or 0,
                        onboarding_loose=getattr(item, "loose_packets", 0) or 0,
                        packets_per_box_limit=getattr(item, "packets_per_box_limit", 1000) or 1000,
                    )
                    item.current_quantity = live_boxes
                    db.add(item)
                    
                    # Join PackagingProfile to get FinishedGoodsStock metadata
                    fg_stock = (
                        db.query(FinishedGoodsStock)
                        .join(PackagingProfile, FinishedGoodsStock.packaging_profile_id == PackagingProfile.id)
                        .filter(
                            FinishedGoodsStock.factory_id == factory_id,
                            PackagingProfile.cup_size_ml == getattr(item, "product_size_ml", 0),
                            func.lower(PackagingProfile.profile_name) == getattr(item, "packaging_size_name", "Standard").strip().lower()
                        )
                        .first()
                    )
                    
                    item_image_url = fg_stock.image_url if fg_stock else None
                    item_category = "CUP_FINISHED"
                    item_variant_name = fg_stock.variant_name if fg_stock else f"{getattr(item, 'product_size_ml', 0)}ml_{getattr(item, 'variety', 'Standard')}"
                    
                    packaging_size = getattr(item, "packaging_size_name", None) or "Standard"
                    processed_inventory.append(
                        {
                            "id": f"final-{item.id}",
                            "factory_id": item.factory_id,
                            "product_id": item.id,
                            "product_size_ml": getattr(item, "product_size_ml", 0) or 0,
                            "variety": getattr(item, "variety", None) or "Standard/White",
                            "item_name": f"{getattr(item, 'product_size_ml', 0) or 0}ml {getattr(item, 'variety', None) or 'Standard/White'} - {packaging_size}",
                            "stock_type": "Final Product",
                            "current_quantity": int(live_boxes),
                            "quantity": int(live_boxes),
                            "unit": "boxes",
                            "price_per_unit": 0,
                            "packaging_size": packaging_size,
                            "packaging_size_name": packaging_size,
                            "pieces_per_packet": getattr(item, "pieces_per_packet", 0) if getattr(item, "pieces_per_packet", None) is not None else 0,
                            "packets_per_box": getattr(item, "packets_per_box_limit", 0) if getattr(item, "packets_per_box_limit", None) is not None else 0,
                            "packets_per_box_limit": getattr(item, "packets_per_box_limit", 0) if getattr(item, "packets_per_box_limit", None) is not None else 0,
                            "category": item_category,
                            "image_url": item_image_url,
                            "variant_name": item_variant_name,
                            "bucket": "finished_goods",
                            "quantity_source": "live_calc",
                        }
                    )
                except Exception as exc:
                    record_row_error("final_product_stock", item, exc)
                    continue
            db.commit()
        except Exception:
            db.rollback()

        set_warning_header()
        return processed_inventory
    except Exception:
        return []


class FinishedGoodVariantCreate(BaseModel):
    """Schema for creating a new finished-good variant from the production form.
    Distinct from FinalStockCreate (which is used by the inventory page for opening
    stock + image upload) — this one is dedicated to variant creation and enforces
    duplicate prevention at the API layer.
    """
    product_size_ml: int = Field(..., gt=0)
    variety: str = Field(default="Standard/White", min_length=1, max_length=100)
    packaging_size_name: str = Field(..., min_length=1, max_length=100)
    pieces_per_packet: int = Field(..., gt=0)
    packets_per_box_limit: int = Field(..., gt=0)
    opening_stock_boxes: int = Field(default=0, ge=0)
    opening_stock_loose_packets: int = Field(default=0, ge=0)


class FinishedGoodVariantResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    factory_id: int
    product_size_ml: int
    variety: str
    packaging_size_name: str
    pieces_per_packet: int
    packets_per_box_limit: int
    current_quantity: int
    total_boxes: int
    loose_packets: int
    created_existing: bool = False


def _build_final_stock_row_response(
    db: Session, factory_id: int, stock: FinalProductStock
) -> FinalStockRow:
    """Helper: serialise a FinalProductStock row into FinalStockRow, including
    the FinishedGoodsStock metadata join (image_url, category, variant_name).
    """
    fg_stock = (
        db.query(FinishedGoodsStock)
        .join(PackagingProfile, FinishedGoodsStock.packaging_profile_id == PackagingProfile.id)
        .filter(
            FinishedGoodsStock.factory_id == str(factory_id),
            PackagingProfile.cup_size_ml == stock.product_size_ml,
            func.lower(PackagingProfile.profile_name) == stock.packaging_size_name.strip().lower()
        )
        .first()
    )
    return FinalStockRow(
        id=stock.id,
        product_size_ml=stock.product_size_ml,
        variety=stock.variety or "Standard/White",
        packaging_size=stock.packaging_size_name,
        packaging_size_name=stock.packaging_size_name,
        pieces_per_packet=stock.pieces_per_packet,
        packets_per_box=stock.packets_per_box_limit,
        current_quantity=stock.current_quantity or 0,
        total_boxes=stock.total_boxes or 0,
        loose_packets=stock.loose_packets or 0,
        packets_per_box_limit=stock.packets_per_box_limit,
        image_url=fg_stock.image_url if fg_stock else None,
        category=fg_stock.category if fg_stock else "CUP_FINISHED",
        variant_name=fg_stock.variant_name if fg_stock else f"{stock.product_size_ml}ml_{stock.variety}",
    )


@router.post("/finished-goods/variants", response_model=FinishedGoodVariantResponse)
def create_finished_good_variant(
    payload: FinishedGoodVariantCreate,
    current_user: User = Depends(check_permissions(INVENTORY_ROLES)),
    db: Session = Depends(get_db),
):
    """Create a new finished-good variant (canonical model: FinalProductStock).
    Duplicate prevention: a variant is uniquely identified by
    (factory_id, product_size_ml, variety, packaging_size_name,
     pieces_per_packet, packets_per_box_limit). If a duplicate exists,
    return 409 with the existing product_id so the frontend can offer
    "select existing" instead of creating a duplicate.
    """
    try:
        factory_id = str(current_user.factory_id)
        variety = (payload.variety or "Standard/White").strip() or "Standard/White"
        packaging_size_name = payload.packaging_size_name.strip()

        existing = (
            db.query(FinalProductStock)
            .filter(FinalProductStock.factory_id == factory_id)
            .filter(FinalProductStock.product_size_ml == payload.product_size_ml)
            .filter(sql_func.lower(FinalProductStock.variety) == variety.lower())
            .filter(sql_func.lower(FinalProductStock.packaging_size_name) == packaging_size_name.lower())
            .filter(FinalProductStock.pieces_per_packet == payload.pieces_per_packet)
            .filter(FinalProductStock.packets_per_box_limit == payload.packets_per_box_limit)
            .first()
        )
        if existing is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={
                    "message": "A finished good variant with these specifications already exists.",
                    "existing_product_id": existing.id,
                    "existing": {
                        "id": existing.id,
                        "product_size_ml": existing.product_size_ml,
                        "variety": existing.variety,
                        "packaging_size_name": existing.packaging_size_name,
                        "pieces_per_packet": existing.pieces_per_packet,
                        "packets_per_box_limit": existing.packets_per_box_limit,
                        "current_quantity": existing.current_quantity or 0,
                    },
                },
            )

        stock = FinalProductStock(
            factory_id=factory_id,
            product_size_ml=payload.product_size_ml,
            variety=variety,
            packaging_size_name=packaging_size_name,
            pieces_per_packet=payload.pieces_per_packet,
            packets_per_box_limit=payload.packets_per_box_limit,
            total_boxes=payload.opening_stock_boxes,
            loose_packets=payload.opening_stock_loose_packets,
            current_quantity=payload.opening_stock_boxes,
        )
        db.add(stock)
        db.flush()

        from services.activity_logger import log_activity
        log_activity(
            db=db,
            factory_id=current_user.factory_id,
            user_id=current_user.id,
            user_name=current_user.username,
            user_role=current_user.role,
            action_type="FINISHED_GOODS_VARIANT_CREATED",
            action_summary=(
                f"Finished good variant created from production: "
                f"{stock.product_size_ml}ml {stock.variety} "
                f"{stock.packaging_size_name} "
                f"({stock.pieces_per_packet}pcs/pkt, "
                f"{stock.packets_per_box_limit}pkts/box, "
                f"opening={payload.opening_stock_boxes} boxes)"
            ),
            entity_type="finished_goods_stock",
            entity_id=stock.id,
            metadata=None,
        )
        db.commit()
        db.refresh(stock)
        return FinishedGoodVariantResponse(
            id=stock.id,
            factory_id=int(stock.factory_id),
            product_size_ml=stock.product_size_ml,
            variety=stock.variety or "Standard/White",
            packaging_size_name=stock.packaging_size_name,
            pieces_per_packet=stock.pieces_per_packet,
            packets_per_box_limit=stock.packets_per_box_limit,
            current_quantity=stock.current_quantity or 0,
            total_boxes=stock.total_boxes or 0,
            loose_packets=stock.loose_packets or 0,
            created_existing=False,
        )
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        log.exception("create_finished_good_variant failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/finished-goods/export")
def export_finished_goods_snapshot(
    date: Optional[date_cls] = Query(None, description="Snapshot date (defaults to today, IST)"),
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    current_user: User = Depends(check_permissions(INVENTORY_ROLES)),
    db: Session = Depends(get_db),
):
    """Download a per-factory finished-goods snapshot.
    Returns one row per (factory × FinalProductStock variant), with:
        - opening stock (onboarding + earliest date TotalBoxes/LoosePackets)
        - produced till date (sum DailyProduction.total_boxes_made +
          loose_packets_made up to snapshot_date)
        - sold till date (sum DailySale.boxes_sold + loose_packets_sold
          up to snapshot_date)
        - current boxes = max(0, opening + produced - sold) as full boxes
        - loose packets = current remainder after full-box conversion
        - total pieces = (current_boxes * packets_per_box_limit) + loose_packets
        - last_updated timestamp
    Factory-scoped (P0 gate). Excel preferred; CSV fallback.
    """
    try:
        from services.timezone_utils import get_kolkata_now

        snapshot_date = date or get_kolkata_now().date()
        factory_id = str(current_user.factory_id)

        factory = db.query(Factory).filter(Factory.id == current_user.factory_id).first()
        factory_name = factory.name if factory else f"Factory {current_user.factory_id}"

        stocks = (
            db.query(FinalProductStock)
            .filter(FinalProductStock.factory_id == factory_id)
            .order_by(
                FinalProductStock.product_size_ml.asc(),
                FinalProductStock.variety.asc(),
                FinalProductStock.packaging_size_name.asc(),
            )
            .all()
        )

        # Headers per the spec
        headers = [
            "Snapshot Date",
            "Factory Name",
            "Product Size (ml)",
            "Variety / Design",
            "Packaging Size Name",
            "Pieces per Packet",
            "Packets per Box",
            "Opening Boxes",
            "Opening Loose Packets",
            "Produced (Boxes) Till Date",
            "Produced (Loose Packets) Till Date",
            "Sold (Boxes) Till Date",
            "Sold (Loose Packets) Till Date",
            "Current Boxes",
            "Loose Packets",
            "Total Pieces",
            "Last Updated",
        ]

        rows_data: List[list] = []
        for stock in stocks:
            limit = stock.packets_per_box_limit or 1
            # Opening: from the row itself (this is the onboarding data
            # imported via Excel or saved via the inventory page).
            opening_boxes = int(stock.total_boxes or 0)
            opening_loose = int(stock.loose_packets or 0)

            # Produced till date (exclude future production beyond snapshot_date)
            prod_boxes = (
                db.query(func.coalesce(func.sum(DailyProduction.total_boxes_made), 0))
                .filter(DailyProduction.factory_id == factory_id)
                .filter(DailyProduction.product_size_ml == stock.product_size_ml)
                .filter(sql_func.lower(DailyProduction.variety) == (stock.variety or "").strip().lower())
                .filter(sql_func.lower(DailyProduction.packaging_size_name) == stock.packaging_size_name.strip().lower())
                .filter(DailyProduction.date <= snapshot_date)
                .scalar()
                or 0
            )
            prod_loose = (
                db.query(func.coalesce(func.sum(DailyProduction.loose_packets_made), 0))
                .filter(DailyProduction.factory_id == factory_id)
                .filter(DailyProduction.product_size_ml == stock.product_size_ml)
                .filter(sql_func.lower(DailyProduction.variety) == (stock.variety or "").strip().lower())
                .filter(sql_func.lower(DailyProduction.packaging_size_name) == stock.packaging_size_name.strip().lower())
                .filter(DailyProduction.date <= snapshot_date)
                .scalar()
                or 0
            )

            # Sold till date
            sold_boxes = (
                db.query(func.coalesce(func.sum(DailySale.boxes_sold), 0))
                .filter(DailySale.factory_id == factory_id)
                .filter(DailySale.product_size_ml == stock.product_size_ml)
                .filter(sql_func.lower(DailySale.variety) == (stock.variety or "").strip().lower())
                .filter(sql_func.lower(DailySale.packaging_size_name) == stock.packaging_size_name.strip().lower())
                .filter(DailySale.date <= snapshot_date)
                .scalar()
                or 0
            )
            sold_loose = (
                db.query(func.coalesce(func.sum(DailySale.loose_packets_sold), 0))
                .filter(DailySale.factory_id == factory_id)
                .filter(DailySale.product_size_ml == stock.product_size_ml)
                .filter(sql_func.lower(DailySale.variety) == (stock.variety or "").strip().lower())
                .filter(sql_func.lower(DailySale.packaging_size_name) == stock.packaging_size_name.strip().lower())
                .filter(DailySale.date <= snapshot_date)
                .scalar()
                or 0
            )

            # Compute current = (opening + produced) - sold (in packets)
            opening_packets = opening_boxes * limit + opening_loose
            produced_packets = int(prod_boxes) * limit + int(prod_loose)
            sold_packets = int(sold_boxes) * limit + int(sold_loose)
            current_packets = opening_packets + produced_packets - sold_packets
            if current_packets < 0:
                current_packets = 0
            current_boxes = current_packets // limit
            current_loose = current_packets % limit
            total_pieces = (current_boxes * stock.pieces_per_packet) + (
                current_loose * stock.pieces_per_packet // limit if limit else 0
            )
            # safer formula: total_pieces = current_packets * pieces_per_packet
            total_pieces = current_packets * (stock.pieces_per_packet or 1)

            last_updated = (
                stock.updated_at.isoformat()
                if stock.updated_at is not None
                else ""
            )

            rows_data.append([
                snapshot_date.isoformat(),
                factory_name,
                stock.product_size_ml,
                stock.variety or "Standard/White",
                stock.packaging_size_name,
                stock.pieces_per_packet,
                limit,
                opening_boxes,
                opening_loose,
                int(prod_boxes),
                int(prod_loose),
                int(sold_boxes),
                int(sold_loose),
                int(current_boxes),
                int(current_loose),
                int(total_pieces),
                last_updated,
            ])

        if format == "csv":
            import csv
            output = BytesIO()
            # open with utf-8 BOM for Excel-friendliness
            text_buf = BytesIO()
            text_wrapper = __import__("io").TextIOWrapper(text_buf, encoding="utf-8-sig", newline="")
            writer = csv.writer(text_wrapper)
            writer.writerow(headers)
            for row in rows_data:
                writer.writerow(row)
            text_wrapper.flush()
            text_buf.seek(0)
            output.write(text_buf.getvalue())
            output.seek(0)
            return Response(
                content=output.read(),
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": (
                        f"attachment; filename="
                        f"finished_goods_snapshot_{factory_name}_"
                        f"{snapshot_date.isoformat()}.csv"
                    )
                },
            )

        # Default: xlsx
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Finished Goods"
        ws.append(headers)
        for row in rows_data:
            ws.append(row)
        # Bold header
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Auto column widths
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": (
                    f"attachment; filename="
                    f"finished_goods_snapshot_{factory_name}_"
                    f"{snapshot_date.isoformat()}.xlsx"
                )
            },
        )
    except HTTPException:
        raise
    except Exception as exc:
        log.exception("export_finished_goods_snapshot failed: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.post("/final-stock", response_model=FinalStockRow)
def save_final_stock(
    payload: FinalStockCreate,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(check_permissions(INVENTORY_ROLES)),
    db: Session = Depends(get_db),
):
    try:
        packaging_size_name = (payload.packaging_size_name or payload.packaging_size or "").strip()
        packets_per_box_limit = payload.packets_per_box_limit or payload.packets_per_box
        quantity = payload.current_quantity
        if quantity is None:
            quantity = payload.total_boxes
        if quantity is None:
            quantity = payload.initial_quantity

        if not packaging_size_name:
            raise HTTPException(status_code=422, detail="packaging_size or packaging_size_name is required")
        if packets_per_box_limit is None:
            raise HTTPException(status_code=422, detail="packets_per_box or packets_per_box_limit is required")

        if payload.product_id:
            stock = (
                db.query(FinalProductStock)
                .filter(FinalProductStock.factory_id == str(current_user.factory_id))
                .filter(FinalProductStock.id == payload.product_id)
                .with_for_update()
                .first()
            )
            if stock is None:
                raise HTTPException(status_code=404, detail="Final product stock item not found")
            if payload.product_size_ml is not None:
                stock.product_size_ml = payload.product_size_ml
        else:
            if payload.product_size_ml is None:
                raise HTTPException(status_code=422, detail="product_size_ml is required")
            stock = (
                db.query(FinalProductStock)
                .filter(FinalProductStock.factory_id == str(current_user.factory_id))
                .filter(FinalProductStock.product_size_ml == payload.product_size_ml)
                .filter(FinalProductStock.variety == payload.variety)
                .filter(FinalProductStock.packaging_size_name == packaging_size_name)
                .with_for_update()
                .first()
            )
            if stock is None:
                stock = FinalProductStock(
                    factory_id=str(current_user.factory_id),
                    product_size_ml=payload.product_size_ml,
                    variety=payload.variety,
                    packaging_size_name=packaging_size_name,
                    pieces_per_packet=payload.pieces_per_packet,
                    packets_per_box_limit=packets_per_box_limit,
                )
                db.add(stock)

        stock.variety = payload.variety
        stock.packaging_size_name = packaging_size_name
        stock.pieces_per_packet = payload.pieces_per_packet
        stock.packets_per_box_limit = packets_per_box_limit
        stock.current_quantity = quantity
        stock.total_boxes = quantity
        stock.loose_packets = payload.loose_packets
        db.commit()
        background_tasks.add_task(
            log_activity,
            db=db,
            factory_id=int(current_user.factory_id),
            user_id=current_user.id,
            user_name=current_user.username,
            user_role=current_user.role,
            action_type="FINISHED_GOODS_STOCK_SAVED",
            action_summary=f"Finished goods stock saved - {stock.product_size_ml}ml {stock.variety or 'Standard/White'} ({stock.total_boxes or 0} boxes)",
            entity_type="finished_goods_stock",
            entity_id=stock.id,
            metadata=None,
        )
        db.refresh(stock)
        fg_stock = (
            db.query(FinishedGoodsStock)
            .join(PackagingProfile, FinishedGoodsStock.packaging_profile_id == PackagingProfile.id)
            .filter(
                FinishedGoodsStock.factory_id == str(current_user.factory_id),
                PackagingProfile.cup_size_ml == stock.product_size_ml,
                func.lower(PackagingProfile.profile_name) == stock.packaging_size_name.strip().lower()
            )
            .first()
        )
        return FinalStockRow(
            id=stock.id,
            product_size_ml=stock.product_size_ml,
            variety=stock.variety or "Standard/White",
            packaging_size=stock.packaging_size_name,
            packaging_size_name=stock.packaging_size_name,
            pieces_per_packet=stock.pieces_per_packet,
            packets_per_box=stock.packets_per_box_limit,
            current_quantity=stock.current_quantity if stock.current_quantity is not None else stock.total_boxes or 0,
            total_boxes=stock.total_boxes or 0,
            loose_packets=stock.loose_packets or 0,
            packets_per_box_limit=stock.packets_per_box_limit,
            image_url=fg_stock.image_url if fg_stock else None,
            category=fg_stock.category if fg_stock else "CUP_FINISHED",
            variant_name=fg_stock.variant_name if fg_stock else f"{stock.product_size_ml}ml_{stock.variety}",
        )
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/final-stock", response_model=List[FinalStockRow])
def list_final_stock(
    search: Optional[str] = None,
    current_user: User = Depends(check_permissions(INVENTORY_ROLES)),
    db: Session = Depends(get_db),
):
    """List finished-good variants for the current factory.
    Optional `?search=` substring-matches against:
      - product_size_ml (digits)
      - variety
      - packaging_size_name
      - pieces_per_packet
      - packets_per_box_limit
    Case-insensitive. Server-side factory scoped (P0 gate).
    """
    try:
        query = db.query(FinalProductStock).filter(
            FinalProductStock.factory_id == str(current_user.factory_id)
        )
        if search:
            term = search.strip().lower()
            if term:
                like_term = f"%{term}%"
                query = query.filter(
                    or_(
                        func.lower(FinalProductStock.variety).like(like_term),
                        func.lower(FinalProductStock.packaging_size_name).like(like_term),
                        func.cast(FinalProductStock.product_size_ml, String).like(like_term),
                        func.cast(FinalProductStock.pieces_per_packet, String).like(like_term),
                        func.cast(FinalProductStock.packets_per_box_limit, String).like(like_term),
                    )
                )
        rows = query.order_by(
            FinalProductStock.product_size_ml.asc(),
            FinalProductStock.variety.asc(),
            FinalProductStock.packaging_size_name.asc(),
        ).all()
        processed_rows = []
        for row in rows:
            live_boxes, live_loose = calculate_live_sku_stock(
                db=db,
                factory_id=str(current_user.factory_id),
                product_size_ml=row.product_size_ml,
                variety=row.variety,
                packaging_size_name=row.packaging_size_name,
                onboarding_boxes=row.total_boxes or 0,
                onboarding_loose=row.loose_packets or 0,
                packets_per_box_limit=row.packets_per_box_limit or 1000,
            )
            # Sync cache
            row.current_quantity = live_boxes
            db.add(row)
            
            fg_stock = (
                db.query(FinishedGoodsStock)
                .join(PackagingProfile, FinishedGoodsStock.packaging_profile_id == PackagingProfile.id)
                .filter(
                    FinishedGoodsStock.factory_id == str(current_user.factory_id),
                    PackagingProfile.cup_size_ml == row.product_size_ml,
                    func.lower(PackagingProfile.profile_name) == row.packaging_size_name.strip().lower()
                )
                .first()
            )
            processed_rows.append(
                FinalStockRow(
                    id=row.id,
                    product_size_ml=row.product_size_ml,
                    variety=row.variety or "Standard/White",
                    packaging_size=row.packaging_size_name,
                    packaging_size_name=row.packaging_size_name,
                    pieces_per_packet=row.pieces_per_packet,
                    packets_per_box=row.packets_per_box_limit,
                    current_quantity=int(live_boxes),
                    total_boxes=row.total_boxes or 0,
                    loose_packets=row.loose_packets or 0,
                    packets_per_box_limit=row.packets_per_box_limit,
                    image_url=fg_stock.image_url if fg_stock else None,
                    category=fg_stock.category if fg_stock else "CUP_FINISHED",
                    variant_name=fg_stock.variant_name if fg_stock else f"{row.product_size_ml}ml_{row.variety}",
                )
            )
        db.commit()
        return processed_rows
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/final-stock/{product_id}/image")
async def upload_product_image(
    product_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(check_permissions(INVENTORY_ROLES)),
    db: Session = Depends(get_db)
):
    try:
        if not file.content_type or not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="Uploaded file is not a valid image.")

        product = db.query(FinalProductStock).filter(
            FinalProductStock.factory_id == str(current_user.factory_id),
            FinalProductStock.id == product_id
        ).first()

        if not product:
            raise HTTPException(status_code=404, detail="Product stock not found")

        profile = db.query(PackagingProfile).filter(
            PackagingProfile.factory_id == str(current_user.factory_id),
            PackagingProfile.cup_size_ml == product.product_size_ml,
            func.lower(PackagingProfile.profile_name) == product.packaging_size_name.strip().lower()
        ).first()

        if not profile:
            raise HTTPException(status_code=404, detail="Packaging profile not found for this product specification")

        fg_stock = db.query(FinishedGoodsStock).filter(
            FinishedGoodsStock.factory_id == str(current_user.factory_id),
            FinishedGoodsStock.packaging_profile_id == profile.id
        ).first()

        if not fg_stock:
            fg_stock = FinishedGoodsStock(
                factory_id=str(current_user.factory_id),
                cup_size_ml=product.product_size_ml,
                packaging_profile_id=profile.id,
                boxes_available=product.current_quantity or 0,
                category="CUP_FINISHED",
                variant_name=f"{product.product_size_ml}ml_{product.variety}"
            )
            db.add(fg_stock)
            db.flush()

        contents = await file.read()
        try:
            with Image.open(BytesIO(contents)) as img:
                if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
                    pass
                else:
                    img = img.convert("RGB")

                img.thumbnail((800, 800))

                os.makedirs("./volumes/media", exist_ok=True)
                filename = f"factory_{current_user.factory_id}_stock_{product_id}.webp"
                filepath = os.path.join("./volumes/media", filename)
                
                img.save(filepath, "WEBP", quality=75)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to process and compress image: {str(e)}")

        relative_url = f"/media/{filename}"
        fg_stock.image_url = relative_url
        db.add(fg_stock)
        db.commit()

        return {"status": "success", "image_url": relative_url}

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/finished-goods/onboard", response_model=FinalStockRow)
def onboard_finished_goods(
    payload: FinishedGoodsOnboardRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(check_permissions(INVENTORY_ROLES)),
    db: Session = Depends(get_db),
):
    factory_id = current_user.factory_id
    product_size_ml_int = int(payload.product_size_ml)
    variety_design = payload.variety_design.strip()
    
    pcs_per_packet = payload.pcs_per_packet
    packets_per_box = payload.packets_per_box
    initial_quantity_boxes = payload.initial_quantity_boxes

    # Auto-generate packaging_size_name if not provided
    if not payload.packaging_size_name or not payload.packaging_size_name.strip():
        size_val = int(payload.product_size_ml) if int(payload.product_size_ml) == payload.product_size_ml else payload.product_size_ml
        packaging_size_name = f"{size_val}ML_{variety_design}_Box"
    else:
        packaging_size_name = payload.packaging_size_name.strip()

    # Enforce safe nested transaction block
    try:
        with db.begin(nested=True):
            # Check if identical product configuration already exists for the same factory_id
            stock = (
                db.query(FinalProductStock)
                .filter(FinalProductStock.factory_id == str(factory_id))
                .filter(FinalProductStock.product_size_ml == product_size_ml_int)
                .filter(func.lower(func.trim(FinalProductStock.variety)) == variety_design.lower())
                .filter(func.lower(func.trim(FinalProductStock.packaging_size_name)) == packaging_size_name.lower())
                .with_for_update()
                .first()
            )

            if stock is not None:
                # Dynamically append initial_quantity_boxes to the existing rows metrics
                stock.total_boxes = (stock.total_boxes or 0) + initial_quantity_boxes
                stock.current_quantity = (stock.current_quantity or 0) + initial_quantity_boxes
                db.add(stock)
                db.flush()
            else:
                # Instantiate a new product row entry first
                stock = FinalProductStock(
                    factory_id=str(factory_id),
                    product_size_ml=product_size_ml_int,
                    variety=variety_design,
                    packaging_size_name=packaging_size_name,
                    pieces_per_packet=pcs_per_packet,
                    packets_per_box_limit=packets_per_box,
                    current_quantity=initial_quantity_boxes,
                    total_boxes=initial_quantity_boxes,
                    loose_packets=0,
                )
                db.add(stock)
                db.flush()

            # Automatically log opening balance context inside the asset registry tracking layers
            def get_or_create_local_inventory(item_name: str, category: str, unit: str) -> Inventory:
                item = (
                    db.query(Inventory)
                    .filter(Inventory.factory_id == str(factory_id))
                    .filter(func.lower(Inventory.item_name) == item_name.lower())
                    .first()
                )
                if item is None:
                    item = Inventory(
                        factory_id=str(factory_id),
                        item_name=item_name.strip(),
                        category=category,
                        unit=unit,
                        quantity=Decimal("0.000"),
                        price_per_unit=Decimal("0.00"),
                    )
                    db.add(item)
                    db.flush()
                return item

            poly_name = f"{product_size_ml_int}ml Polybag"
            box_name = packaging_size_name

            poly_inventory = get_or_create_local_inventory(poly_name, "Packaging", "pieces")
            box_inventory = get_or_create_local_inventory(box_name, "Packaging", "pieces")

            # Check if PackagingProfile exists
            profile = (
                db.query(PackagingProfile)
                .filter(PackagingProfile.factory_id == str(factory_id))
                .filter(func.lower(PackagingProfile.profile_name) == packaging_size_name.lower())
                .first()
            )
            if profile is None:
                profile = PackagingProfile(
                    factory_id=str(factory_id),
                    profile_name=packaging_size_name,
                    product_name=f"{product_size_ml_int}ml Paper Cup",
                    product_name_ml=product_size_ml_int,
                    cup_size_ml=product_size_ml_int,
                    polybag_capacity=pcs_per_packet,
                    box_capacity=pcs_per_packet * packets_per_box,
                    box_size_name=packaging_size_name,
                    cups_per_poly=pcs_per_packet,
                    cups_per_polybag=pcs_per_packet,
                    polys_per_box=packets_per_box,
                    polybags_per_box=packets_per_box,
                    box_inventory_id=box_inventory.id,
                    poly_inventory_id=poly_inventory.id,
                )
                db.add(profile)
                db.flush()

            # Cross-sync FinishedGoodsStock
            finished = (
                db.query(FinishedGoodsStock)
                .filter(FinishedGoodsStock.factory_id == str(factory_id))
                .filter(FinishedGoodsStock.packaging_profile_id == profile.id)
                .first()
            )
            if finished is None:
                finished = FinishedGoodsStock(
                    factory_id=str(factory_id),
                    cup_size_ml=product_size_ml_int,
                    packaging_profile_id=profile.id,
                    boxes_available=initial_quantity_boxes,
                    category="CUP_FINISHED",
                    variant_name=f"{product_size_ml_int}ml_{variety_design}"
                )
                db.add(finished)
                db.flush()
            else:
                finished.boxes_available = (finished.boxes_available or 0) + initial_quantity_boxes
                db.add(finished)
                db.flush()

            # Synchronize PackagingMetrics
            metric = (
                db.query(PackagingMetrics)
                .filter(PackagingMetrics.factory_id == str(factory_id))
                .filter(PackagingMetrics.cup_size_ml == product_size_ml_int)
                .filter(func.lower(PackagingMetrics.variant_name) == variety_design.lower())
                .first()
            )
            if metric is None:
                metric = PackagingMetrics(
                    factory_id=str(factory_id),
                    cup_size_ml=product_size_ml_int,
                    variant_name=variety_design,
                    kg_per_box=Decimal("10.000"),
                    cups_per_box=packets_per_box,
                )
                db.add(metric)
                db.flush()
            else:
                metric.cups_per_box = packets_per_box
                db.add(metric)
                db.flush()

            # Recalculate dynamic stock caches
            recalculate_and_sync_sku_stock(
                db=db,
                factory_id=str(factory_id),
                product_size_ml=product_size_ml_int,
                variety=variety_design,
                packaging_size_name=packaging_size_name,
            )

        db.commit()
        background_tasks.add_task(
            log_activity,
            db=db,
            factory_id=int(current_user.factory_id),
            user_id=current_user.id,
            user_name=current_user.username,
            user_role=current_user.role,
            action_type="FINISHED_GOODS_STOCK_SAVED",
            action_summary=f"Finished goods stock saved - {product_size_ml_int}ml {variety_design} ({initial_quantity_boxes} boxes)",
            entity_type="finished_goods_stock",
            entity_id=stock.id,
            metadata=None,
        )
        db.refresh(stock)

        # Retrieve image details and other info for output mapping
        fg_stock = (
            db.query(FinishedGoodsStock)
            .join(PackagingProfile, FinishedGoodsStock.packaging_profile_id == PackagingProfile.id)
            .filter(
                FinishedGoodsStock.factory_id == str(factory_id),
                PackagingProfile.cup_size_ml == stock.product_size_ml,
                func.lower(PackagingProfile.profile_name) == stock.packaging_size_name.strip().lower()
            )
            .first()
        )

        return FinalStockRow(
            id=stock.id,
            product_size_ml=stock.product_size_ml,
            variety=stock.variety or "Standard/White",
            packaging_size=stock.packaging_size_name,
            packaging_size_name=stock.packaging_size_name,
            pieces_per_packet=stock.pieces_per_packet,
            packets_per_box=stock.packets_per_box_limit,
            current_quantity=stock.current_quantity,
            total_boxes=stock.total_boxes or 0,
            loose_packets=stock.loose_packets or 0,
            packets_per_box_limit=stock.packets_per_box_limit,
            image_url=fg_stock.image_url if fg_stock else None,
            variant_name=fg_stock.variant_name if fg_stock else f"{stock.product_size_ml}ml_{stock.variety}",
        )

    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to onboard finished goods opening stock: {str(exc)}"
        )
