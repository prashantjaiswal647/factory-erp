import logging
import hashlib
import re
from pathlib import Path
from uuid import uuid4
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, status, BackgroundTasks, UploadFile
from pydantic import BaseModel, ConfigDict, Field, model_validator
from fastapi.responses import StreamingResponse
from sqlalchemy import func as sql_func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from auth import assert_owner_delete_permission, normalize_phone_number, require_owner, require_owner_delete
from dependencies import FACTORY_VIEW_ROLES, OWNER_ROLES, check_permissions
from db import get_db
from models import (
    BlankStock,
    BottomStock,
    BoxStock,
    CostingMaster,
    Customer,
    Factory,
    FactorySettings,
    FinalProductStock,
    FinishedGoodsStock,
    Inventory,
    Machine,
    MaterialYield,
    PackagingProfile,
    RawMaterial,
    RawMaterialMetrics,
    PackagingMetrics,
    PlasticStock,
    User,
    Worker,
    WorkerOpeningAttendance,
    OutstandingBill,
)
from schemas import (
    CustomerPayload,
    MachinePayload,
    MaterialYieldPayload,
    OnboardingCompleteRequest,
    OnboardingCompleteResponse,
    PackagingProfilePayload,
    RawMaterialPayload,
    Step1Request,
    Step1Response,
    Step2MachineItem,
    Step2Request,
    Step2Response,
    Step3PackagingMetricItem,
    Step3RawMaterialMetricItem,
    Step3Request,
    Step3Response,
    BoxStockCreate,
    PlasticStockCreate,
    WorkerCreate,
    WorkerPayload,
    FactoryProfileUpdate,
    FactoryProfileResponse,
    WorkerResponse,
)
from routers.operations import log_factory_operation
from services.activity_logger import log_activity
from services.carton_mapping import normalize_carton_type, parse_allowed_sizes, serialize_finished_product_sizes
from services.n8n_sync import sync_data_to_n8n_bg
from services.bulk_validation import (
    BulkValidationReport,
    ValidationIssue,
    ValidationSeverity,
    enrich_failed_rows,
    make_report,
)
from services.accounting import create_outstanding_bill
from subscription_limits import check_machine_limit, get_machine_limit_usage

router = APIRouter(prefix="/api/onboarding", tags=["onboarding"])
v1_router = APIRouter(prefix="/api/v1/onboarding", tags=["onboarding"])
logger = logging.getLogger(__name__)

BULK_TEMPLATE_COLUMNS = {
    "company_profile": ["row_type", "factory_name", "gstin", "factory_address", "invoice_prefix", "advance_upi_discount", "bill_of_supply_start_seq", "tax_invoice_start_seq", "bill_of_supply_simple_start_seq"],
    "worker": ["row_type", "worker_restore_key", "name", "mobile_number", "daily_wages", "duty_hours", "shift_timing", "shift_type", "previous_attendance_details"],
    "customer": ["row_type", "customer_restore_key", "name", "firm_name", "contact_number", "phone_number", "email", "place", "address", "gst_number", "previous_due", "opening_outstanding_date", "opening_outstanding_note", "advance_balance", "advance_balance_date", "advance_balance_note"],
    "machine": ["row_type", "machine_restore_key", "machine_number", "machine_name", "machine_type", "default_operating_speed", "target_output_per_shift", "mould_size_ml", "bottom_size_mm"],
    "blank_stock": ["row_type", "material_restore_key", "material_name", "variety_design", "size_ml", "linked_bottom_size_mm", "weight_per_bora_kg", "total_boras_sacks"],
    "bottom_reel": ["row_type", "material_restore_key", "bottom_size_mm", "variety_design", "total_individual_rolls", "total_weight_kg", "bottom_price_per_kg"],
    "box_stock": ["row_type", "box_type", "box_quantity_pieces", "price_per_box_rs", "size_for_finished_product"],
    "plastic_stock": ["row_type", "plastic_size_type", "used_for_cup_size_ml", "total_boras_sacks", "weight_per_bora_kg", "price_per_kg_rs"],
    "finished_goods": ["row_type", "product_restore_key", "product_size_ml", "variety_design", "packaging_size_name", "carton_type", "pcs_per_packet", "packets_per_box", "initial_stock_boxes", "initial_loose_packets"],
}

BULK_MASTER_SHEETS = {
    "Company Profile": "company_profile",
    "Workers": "worker",
    "Customers": "customer",
    "Machines": "machine",
    "Raw Materials": "raw_materials",
    "Finished Goods": "finished_goods",
}

OWNER_FRIENDLY_SHEETS = {
    "Factory_Profile": "company_profile",
    "Customers": "customer",
    "Workers": "worker",
    "Machines": "machine",
    "Cup_Blank": "blank_stock",
    "Bottom_Reel": "bottom_reel",
    "Box_Stock": "box_stock",
    "Plastic_Stock": "plastic_stock",
    "Finished_Goods": "finished_goods",
    "Costing_Optional": "costing_optional",
}

OWNER_HEADER_ALIASES = {
    "customer": {
        "customer_phone": "phone_number",
        "mobile_number": "phone_number",
        "customer_firm_name": "firm_name",
        "opening_outstanding": "previous_due",
        "opening_outstanding_amount": "previous_due",
    },
    "worker": {"worker_phone": "mobile_number", "phone_number": "mobile_number"},
    "machine": {"size_ml": "mould_size_ml", "machine_speed": "default_operating_speed"},
    "blank_stock": {
        "cup_size_ml": "size_ml",
        "product_size_ml": "size_ml",
        "carton_type": "material_name",
        "blank_description": "material_name",
        "material_name": "material_name",
        "design_/_variety_name": "variety_design",
        "cup_design": "variety_design",
        "design": "variety_design",
        "variety": "variety_design",
        "opening_boras": "total_boras_sacks",
        "total_bora": "total_boras_sacks",
        "opening_bora_quantity": "total_boras_sacks",
        "total_boras": "total_boras_sacks",
        "weight_per_bora_kg": "weight_per_bora_kg",
        "linked_bottom_size_mm": "linked_bottom_size_mm",
    },
    "bottom_reel": {
        "linked_bottom_size_mm": "bottom_size_mm",
        "opening_rolls": "total_individual_rolls",
        "total_weight": "total_weight_kg",
        "design_/_variety_name": "variety_design",
        "design": "variety_design",
    },
    "box_stock": {
        "carton_type": "box_type",
        "packaging_size_name": "box_type",
        "quantity": "box_quantity_pieces",
        "opening_box_quantity": "box_quantity_pieces",
        "price_per_box": "price_per_box_rs",
        "size_for_finished_product": "size_for_finished_product",
    },
    "plastic_stock": {
        "cup_size_ml": "used_for_cup_size_ml",
        "product_size_ml": "used_for_cup_size_ml",
        "plastic_type": "plastic_size_type",
        "used_for_cup_sizes_ml": "used_for_cup_size_ml",
        "opening_boras": "total_boras_sacks",
        "price_per_kg": "price_per_kg_rs",
    },
    "finished_goods": {
        "cup_size_ml": "product_size_ml",
        "box_type": "carton_type",
        "design_/_variety_name": "variety_design",
        "design": "variety_design",
        "packing_name": "packaging_size_name",
    },
}

RAW_MATERIAL_SECTIONS = {
    "blank_stock": {"label_row": 1, "header_row": 2, "data_start": 3, "data_end": 15, "title": "SECTION A: CUP BLANK MATERIAL", "marker": "CUP BLANK"},
    "bottom_reel": {"label_row": 17, "header_row": 18, "data_start": 19, "data_end": 35, "title": "SECTION B: BOTTOM REEL MATERIAL", "marker": "BOTTOM REEL"},
    "box_stock": {"label_row": 37, "header_row": 38, "data_start": 39, "data_end": 55, "title": "SECTION C: BOX PACKAGING STOCK", "marker": "BOX PACKAGING"},
    "plastic_stock": {"label_row": 57, "header_row": 58, "data_start": 59, "data_end": 80, "title": "SECTION D: PP PLASTIC PACKAGING STOCK", "marker": "PP PLASTIC"},
}

MASTER_ONBOARDING_FILENAME = "master_onboarding_bulk_upload.xlsx"
OWNER_ONBOARDING_FILENAME = "Munshi_AI_Factory_Owner_Onboarding_Template.xlsx"
OWNER_TEMPLATE_COLUMNS = {
    "Factory_Profile": ["Factory Name", "GST Number", "Factory Address", "Invoice Prefix"],
    "Customers": ["Customer Name", "Phone", "Firm Name", "Place", "Address", "Opening Outstanding"],
    "Workers": ["Worker Name", "Mobile", "Daily Wages", "Duty Hours", "Shift Timing", "Shift Type"],
    "Machines": ["Machine Number", "Machine Name", "Machine Type", "Machine Size ML", "Bottom Size MM"],
    "Cup_Blank": ["Material Name", "Cup Size ML", "Design", "Linked Bottom Size MM", "Weight Per Bora KG", "Total Boras Sacks"],
    "Bottom_Reel": ["Bottom Size MM", "Opening Rolls", "Total Weight KG"],
    "Box_Stock": ["Carton Type", "Carton Quantity", "Price Per Box Rs", "Size For Finished Product"],
    "Plastic_Stock": ["Plastic Type", "Used For Cup Sizes ML", "Total Boras Sacks", "Weight Per Bora KG", "Price Per KG Rs"],
    "Finished_Goods": ["Cup Size ML", "Design", "Packaging Size Name", "Carton Type", "Pieces Per Packet", "Packets Per Carton", "Opening Boxes", "Opening Loose Packets"],
    "Costing_Optional": ["Paper Price Per KG", "Bottom Price Per KG", "Plastic Price Per KG", "Carton Price"],
}
TEXT_BULK_COLUMNS = {
    "row_type",
    "factory_name",
    "gstin",
    "factory_address",
    "invoice_prefix",
    "name",
    "customer_restore_key",
    "worker_restore_key",
    "machine_restore_key",
    "material_restore_key",
    "product_restore_key",
    "mobile_number",
    "firm_name",
    "contact_number",
    "phone_number",
    "place",
    "address",
    "gst_number",
    "machine_name",
    "machine_number",
    "machine_type",
    "material_name",
    "box_type",
    "plastic_size_type",
    "variety_design",
    "packaging_size_name",
    "carton_type",
    "size_for_finished_product",
    "email",
    "opening_outstanding_note",
    "advance_balance_note",
    "shift_timing",
    "shift_type",
}

HEADER_ALIASES = {
    "customer_name": "name",
    "phone": "phone_number",
    "total_weight": "total_weight_automatic_calculation",
    "total_weight_kg_automatic_calculation": "total_weight_kg",
    "total_weight_kg=": "total_weight_kg",
    "total_weight=": "total_weight_automatic_calculation",
    "kg_per_sack": "weight_per_bora_kg",
    "quantity_of_total_bora": "total_boras_sacks",
    "quantity of total bora": "total_boras_sacks",
    "total_plastic_kg": "total_plastic_kg_automatic_calculation",
    "total_plastic_kg=": "total_plastic_kg_automatic_calculation",
    "factory_name": "factory_name",
    "factory_address": "factory_address",
    "gst_number": "gstin",
    "worker_name": "name",
    "worker_mobile": "mobile_number",
    "mobile": "mobile_number",
    "machine_size_ml": "mould_size_ml",
    "cup_size_ml": "product_size_ml",
    "cup_size": "product_size_ml",
    "cup_size_ml_": "product_size_ml",
    "design": "variety_design",
    "variety": "variety_design",
    "packing_name": "packaging_size_name",
    "packaging_name": "packaging_size_name",
    "carton_name": "carton_type",
    "pieces_per_packet": "pcs_per_packet",
    "packets_per_carton": "packets_per_box",
    "carton_quantity": "box_quantity_pieces",
    "bottom_size": "bottom_size_mm",
    "linked_bottom_size": "linked_bottom_size_mm",
    "cup_blank_size_ml": "size_ml",
    "blank_size_ml": "size_ml",
    "stock_bora": "total_boras_sacks",
    "stock_rolls": "total_individual_rolls",
    "opening_boxes": "initial_stock_boxes",
    "opening_loose_packets": "initial_loose_packets",
    "design_/_variety_name": "variety_design",
    "cup_design": "variety_design",
    "opening_boras": "total_boras_sacks",
    "total_bora": "total_boras_sacks",
    "opening_bora_quantity": "total_boras_sacks",
    "total_boras": "total_boras_sacks",
    "blank_description": "material_name",
    "price_per_box": "price_per_box_rs",
    "price_per_kg": "price_per_kg_rs",
    "opening_outstanding_amount": "previous_due",
    "opening_box_quantity": "box_quantity_pieces",
}

OPTIONAL_BULK_HEADERS = {
    "customer": {
        "customer_restore_key",
        "firm_name",
        "contact_number",
        "place",
        "gst_number",
        "previous_due",
        "advance_balance",
        "email",
        "opening_outstanding_date",
        "opening_outstanding_note",
        "advance_balance_date",
        "advance_balance_note",
    },
    "worker": {"worker_restore_key", "shift_timing", "shift_type"},
    "machine": {"machine_restore_key", "machine_number", "machine_type"},
    "blank_stock": {"material_restore_key", "total_boras_sacks"},
    "bottom_reel": {"material_restore_key", "bottom_price_per_kg"},
    "finished_goods": {"product_restore_key", "initial_loose_packets"},
}

BULK_NUMERIC_DEFAULTS = {
    "worker": {
        "daily_wages": Decimal("0"),
        "duty_hours": Decimal("8"),
        "previous_attendance_details": Decimal("0"),
    },
    "customer": {
        "previous_due": Decimal("0"),
        "advance_balance": Decimal("0"),
    },
    "blank_stock": {
        "weight_per_bora_kg": Decimal("0"),
        "total_boras_sacks": Decimal("0"),
    },
    "bottom_reel": {
        "total_individual_rolls": Decimal("0"),
        "total_weight_kg": Decimal("0"),
        "bottom_price_per_kg": Decimal("0"),
    },
    "box_stock": {
        "box_quantity_pieces": Decimal("0"),
        "price_per_box_rs": Decimal("0"),
    },
    "plastic_stock": {
        "total_boras_sacks": Decimal("0"),
        "weight_per_bora_kg": Decimal("0"),
        "price_per_kg_rs": Decimal("0"),
    },
    "finished_goods": {
        "pcs_per_packet": Decimal("1"),
        "packets_per_box": Decimal("1"),
        "initial_stock_boxes": Decimal("0"),
        "initial_loose_packets": Decimal("0"),
    },
}

SAMPLE_BULK_ROWS = {
    "company_profile": ["SAMPLE", "Munshi Demo Factory", "07ABCDE1234F1Z5", "Wazirpur Industrial Area, Delhi", "INV-", 2, 1, 1, 1],
    "worker": ["SAMPLE", "WRK-001", "Akash Kumar", "82858117277", 400, 8, "08:00-16:00", "Day", 0],
    "customer": ["SAMPLE", "CUS-001", "Rajesh Kumar", "Rajesh Traders", "9876543210", "9876543210", "rajesh@example.com", "Delhi", "Wazirpur Industrial Area, Delhi", "07ABCDE1234F1Z5", 1500, "2026-04-01", "Opening balance", 0, None, None],
    "machine": ["SAMPLE", "MAC-001", "M-01", "Hi-Speed Cup Machine X", "Paper Cup", 120, 55000, 210, 68],
    "blank_stock": ["SAMPLE", "MAT-BL-210", "Cup Blank Paper", "Standard/White", 210, 68, 20, 25],
    "bottom_reel": ["SAMPLE", "MAT-BT-68", 68, "Standard/White", 1200, 180, 110],
    "box_stock": ["SAMPLE", "Big Box", 500, 18, "210,250,300"],
    "plastic_stock": ["SAMPLE", "PP 210ml Sleeve", 210, 25, 20, 145],
    "finished_goods": ["SAMPLE", "SKU-210-WHITE", 210, "Standard/White", "210- lovely day - 48*62", "Big Box", 100, 10, 50, 0],
}


class CompanyProfileBulkRow(BaseModel):
    row_type: str = Field(..., max_length=20)
    factory_name: str = Field(..., min_length=1, max_length=255)
    gstin: Optional[str] = Field(default=None, max_length=50)
    factory_address: Optional[str] = Field(default=None, max_length=500)
    invoice_prefix: Optional[str] = Field(default="INV-", max_length=50)
    advance_upi_discount: Decimal = Field(default=Decimal("2.00"), ge=0, le=100)
    bill_of_supply_start_seq: int = Field(default=1, ge=1)
    tax_invoice_start_seq: int = Field(default=1, ge=1)
    bill_of_supply_simple_start_seq: int = Field(default=1, ge=1)


class WorkerBulkRow(BaseModel):
    row_type: str = Field(..., max_length=20)
    worker_restore_key: Optional[str] = Field(default=None, max_length=100)
    name: str = Field(..., min_length=1, max_length=255)
    mobile_number: Optional[str] = Field(default=None, max_length=50)
    daily_wages: Decimal = Field(default=Decimal("0"), ge=0)
    duty_hours: Decimal = Field(default=Decimal("8"), ge=0)
    shift_timing: Optional[str] = Field(default=None, max_length=100)
    shift_type: Optional[str] = Field(default=None, max_length=100)
    previous_attendance_details: Decimal = Field(default=Decimal("0"), ge=0)


class CustomerBulkRow(BaseModel):
    row_type: str = Field(..., max_length=20)
    customer_restore_key: Optional[str] = Field(default=None, max_length=100)
    name: str = Field(..., min_length=1, max_length=255)
    firm_name: Optional[str] = Field(default=None, max_length=255)
    contact_number: Optional[str] = Field(default=None, max_length=50)
    phone_number: Optional[str] = Field(default=None, max_length=50)
    email: Optional[str] = Field(default=None, max_length=255)
    place: Optional[str] = Field(default=None, max_length=255)
    address: Optional[str] = Field(default=None, max_length=500)
    gst_number: Optional[str] = Field(default=None, max_length=50)
    previous_due: Decimal = Field(default=Decimal("0"))
    opening_outstanding_date: Optional[date] = None
    opening_outstanding_note: Optional[str] = None
    advance_balance: Decimal = Field(default=Decimal("0"))
    advance_balance_date: Optional[date] = None
    advance_balance_note: Optional[str] = None

    @model_validator(mode="before")
    @classmethod
    def validate_customer_bulk(cls, data: Any) -> Any:
        if isinstance(data, dict):
            # Check previous_due
            pd = data.get("previous_due")
            if pd == "" or pd is None:
                data["previous_due"] = Decimal("0")
            else:
                try:
                    pd_val = Decimal(str(pd))
                    if pd_val < 0:
                        raise ValueError("Opening outstanding cannot be negative.")
                except (ValueError, TypeError):
                    raise ValueError("Opening outstanding cannot be negative.")
            # Check advance_balance
            ab = data.get("advance_balance")
            if ab == "" or ab is None:
                data["advance_balance"] = Decimal("0")
            else:
                try:
                    ab_val = Decimal(str(ab))
                    if ab_val < 0:
                        raise ValueError("Advance balance cannot be negative.")
                except (ValueError, TypeError):
                    raise ValueError("Advance balance cannot be negative.")
        return data


class MachineBulkRow(BaseModel):
    row_type: str = Field(..., max_length=20)
    machine_restore_key: Optional[str] = Field(default=None, max_length=100)
    machine_number: Optional[str] = Field(default=None, max_length=50)
    machine_name: str = Field(..., min_length=1, max_length=255)
    machine_type: Optional[str] = Field(default=None, max_length=255)
    default_operating_speed: int = Field(default=0, ge=0)
    target_output_per_shift: int = Field(default=0, ge=0)
    mould_size_ml: Optional[int] = Field(default=None, gt=0)
    bottom_size_mm: Optional[int] = Field(default=None, gt=0)


class BlankStockBulkRow(BaseModel):
    row_type: str = Field(..., max_length=20)
    material_restore_key: Optional[str] = Field(default=None, max_length=100)
    material_name: str = Field(..., min_length=1, max_length=255)
    variety_design: str = Field(default="", max_length=100)
    size_ml: int = Field(..., gt=0)
    linked_bottom_size_mm: Optional[int] = Field(default=None, gt=0)
    weight_per_bora_kg: Decimal = Field(default=Decimal("0"), ge=0)
    total_boras_sacks: Decimal = Field(default=Decimal("0"), ge=0)


class BottomReelBulkRow(BaseModel):
    row_type: str = Field(..., max_length=20)
    material_restore_key: Optional[str] = Field(default=None, max_length=100)
    bottom_size_mm: int = Field(..., gt=0)
    variety_design: str = Field(default="", max_length=100)
    total_individual_rolls: int = Field(default=0, ge=0)
    total_weight_kg: Decimal = Field(default=Decimal("0"), ge=0)
    bottom_price_per_kg: Decimal = Field(default=Decimal("0"), ge=0)


class BoxStockBulkRow(BaseModel):
    row_type: str = Field(..., max_length=20)
    box_type: str = Field(..., min_length=1, max_length=100)
    box_quantity_pieces: int = Field(default=0, ge=0)
    price_per_box_rs: float = Field(default=0, ge=0)
    size_for_finished_product: str = Field(..., min_length=1, max_length=500)

    @model_validator(mode="after")
    def validate_finished_product_sizes(self):
        parse_allowed_sizes(self.size_for_finished_product)
        return self


class PlasticStockBulkRow(BaseModel):
    row_type: str = Field(..., max_length=20)
    plastic_size_type: str = Field(..., min_length=1, max_length=100)
    used_for_cup_size_ml: int = Field(..., gt=0)
    total_boras_sacks: int = Field(default=0, ge=0)
    weight_per_bora_kg: float = Field(default=0, ge=0)
    price_per_kg_rs: float = Field(default=0, ge=0)


class FinishedGoodsBulkRow(BaseModel):
    row_type: str = Field(..., max_length=20)
    product_restore_key: Optional[str] = Field(default=None, max_length=100)
    product_size_ml: int = Field(..., gt=0)
    variety_design: str = Field(default="Standard/White", max_length=100)
    packaging_size_name: str = Field(..., min_length=1, max_length=100)
    carton_type: str = Field(default="", max_length=100)
    pcs_per_packet: int = Field(default=1, gt=0)
    packets_per_box: int = Field(default=1, gt=0)
    initial_stock_boxes: int = Field(default=0, ge=0)
    initial_loose_packets: int = Field(default=0, ge=0)


BULK_ROW_MODELS = {
    "company_profile": CompanyProfileBulkRow,
    "worker": WorkerBulkRow,
    "customer": CustomerBulkRow,
    "machine": MachineBulkRow,
    "blank_stock": BlankStockBulkRow,
    "bottom_reel": BottomReelBulkRow,
    "box_stock": BoxStockBulkRow,
    "plastic_stock": PlasticStockBulkRow,
    "finished_goods": FinishedGoodsBulkRow,
}


def get_or_create_factory_settings(db: Session, factory_id: int) -> FactorySettings:
    settings = (
        db.query(FactorySettings)
        .filter(FactorySettings.factory_id == int(factory_id))
        .with_for_update()
        .first()
    )
    if settings is None:
        settings = FactorySettings(factory_id=int(factory_id))
        db.add(settings)
        db.flush()
    return settings


def normalize_bulk_value(value):
    try:
        import pandas as pd
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def bulk_str(value) -> str:
    value = normalize_bulk_value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalized_identity(value) -> str:
    text = bulk_str(value).casefold()
    text = re.sub(r"\s*[-–—]\s*", "-", text)
    return " ".join(text.split())


def normalized_phone(value) -> str:
    digits = re.sub(r"\D", "", bulk_str(value))
    return digits[-10:] if len(digits) >= 10 else digits


def generated_restore_key(prefix: str, *identity_parts) -> str:
    identity = "|".join(normalized_identity(part) for part in identity_parts)
    digest = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:16]
    return f"{prefix}-{digest}"


def generate_owner_restore_keys(valid_by_type: dict[str, list[dict]]) -> None:
    for row in valid_by_type.get("customer", []):
        if not row.get("customer_restore_key"):
            phone = normalized_phone(row.get("phone_number") or row.get("contact_number"))
            row["customer_restore_key"] = generated_restore_key(
                "CUS", phone or row.get("name"), row.get("firm_name"), row.get("place")
            )
    for row in valid_by_type.get("worker", []):
        if not row.get("worker_restore_key"):
            row["worker_restore_key"] = generated_restore_key(
                "WRK", normalized_phone(row.get("mobile_number")) or row.get("name")
            )
    for row in valid_by_type.get("machine", []):
        if not row.get("machine_restore_key"):
            row["machine_restore_key"] = generated_restore_key(
                "MAC", row.get("machine_number") or row.get("machine_name"), row.get("mould_size_ml")
            )
    for sub_tab_type, prefix, size_field in (
        ("blank_stock", "MAT-BL", "size_ml"),
        ("bottom_reel", "MAT-BT", "bottom_size_mm"),
    ):
        for row in valid_by_type.get(sub_tab_type, []):
            if not row.get("material_restore_key"):
                row["material_restore_key"] = generated_restore_key(
                    prefix, sub_tab_type, row.get(size_field), row.get("variety_design")
                )
    for row in valid_by_type.get("finished_goods", []):
        if not row.get("product_restore_key"):
            row["product_restore_key"] = generated_restore_key(
                "SKU",
                row.get("product_size_ml"),
                row.get("variety_design"),
                row.get("packaging_size_name"),
            )


def normalize_bulk_cell(key: str, value):
    value = normalize_bulk_value(value)
    if key in TEXT_BULK_COLUMNS:
        return bulk_str(value)
    return value


def is_blank_bulk_value(value) -> bool:
    value = normalize_bulk_value(value)
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def apply_bulk_numeric_defaults(sub_tab_type: str, row: dict) -> dict:
    for key, default_value in BULK_NUMERIC_DEFAULTS.get(sub_tab_type, {}).items():
        if is_blank_bulk_value(row.get(key)):
            row[key] = default_value
    return row


def coerce_excel_int_token(value) -> int:
    value = normalize_bulk_value(value)
    if value is None:
        raise ValueError("empty integer value")
    if isinstance(value, bool):
        raise ValueError("boolean is not a valid integer value")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = re.sub(r"(?i)\s*ml\s*$", "", bulk_str(value)).strip()
    if not text:
        raise ValueError("empty integer value")
    try:
        decimal_value = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError(
            "Enter a whole number, for example 210 or 210 ml. / "
            "पूरा नंबर लिखें, जैसे 210 या 210 ml।"
        ) from None
    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f"{text} is not a whole number")
    return int(decimal_value)


def split_bulk_int_values(value) -> list[int]:
    value = normalize_bulk_value(value)
    if value is None:
        return []
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return [coerce_excel_int_token(value)]
    tokens = [token.strip() for token in re.split(r"[,;/|]+", bulk_str(value)) if token.strip()]
    return [coerce_excel_int_token(token) for token in tokens]


def expand_bulk_row_variants(sub_tab_type: str, row: dict) -> list[dict]:
    if sub_tab_type != "plastic_stock":
        return [row]
    cup_sizes = split_bulk_int_values(row.get("used_for_cup_size_ml"))
    if not cup_sizes:
        return [row]
    return [{**row, "used_for_cup_size_ml": cup_size} for cup_size in cup_sizes]


def canonical_bulk_header(value) -> str:
    header = bulk_str(value).lower()
    header = header.replace("\n", " ").replace("\r", " ").strip()
    header = " ".join(header.split())
    header = header.replace(" ", "_").replace("-", "_")
    header = header.rstrip("=:")
    return HEADER_ALIASES.get(header, header)


def canonicalize_bulk_frame(frame):
    canonical_columns: list[str] = []
    seen: set[str] = set()
    keep_indices: list[int] = []
    for index, column in enumerate(frame.columns.tolist()):
        canonical = canonical_bulk_header(column)
        if not canonical or canonical in seen:
            continue
        canonical_columns.append(canonical)
        seen.add(canonical)
        keep_indices.append(index)
    canonical_frame = frame.iloc[:, keep_indices].copy()
    canonical_frame.columns = canonical_columns
    return canonical_frame


def actual_rows_only(frame):
    row_type = frame.get("row_type")
    if row_type is None:
        return frame.iloc[0:0]
    mask = row_type.fillna("").astype(str).str.strip().str.upper() == "ACTUAL"
    return frame[mask]


def row_has_owner_data(raw_row, expected: list[str]) -> bool:
    return any(
        key != "row_type" and not is_blank_bulk_value(raw_row.get(key))
        for key in expected
    )


def validate_bulk_frame(
    frame,
    sub_tab_type: str,
    sheet_name: str | None = None,
    row_offset: int = 2,
    *,
    strict_validation: bool = False,
    compatibility_warnings: list[dict] | None = None,
    owner_friendly: bool = False,
) -> tuple[list[dict], list[dict]]:
    expected = BULK_TEMPLATE_COLUMNS[sub_tab_type]
    frame = canonicalize_bulk_frame(frame)
    headers = [str(column).strip() for column in frame.columns.tolist()]
    optional_headers = set(OPTIONAL_BULK_HEADERS.get(sub_tab_type, set()))
    if not strict_validation:
        optional_headers.update({
            "blank_stock": {"variety_design", "linked_bottom_size_mm"},
            "bottom_reel": {"variety_design"},
            "finished_goods": {"initial_loose_packets"},
        }.get(sub_tab_type, set()))

    if sub_tab_type == "blank_stock" and "total_boras_sacks" not in headers:
        unmapped_headers = [h for h in headers if h not in expected]
        for h in unmapped_headers:
            if any(x in h for x in ["bora", "sack", "opening"]):
                return [], [{
                    "sheet": sheet_name or sub_tab_type,
                    "row": row_offset,
                    "error": "Opening Bora Quantity could not be mapped. Expected column: total_boras_sacks / Opening Bora Quantity.",
                    "entity_type": sub_tab_type,
                }]

    missing_headers = [column for column in expected if column not in headers and column not in optional_headers]
    if missing_headers:
        return [], [{
            "sheet": sheet_name or sub_tab_type,
            "row": row_offset,
            "error": "Header mismatch",
            "expected_headers": expected,
            "received_headers": headers,
            "missing_headers": missing_headers,
        }]

    model = BULK_ROW_MODELS[sub_tab_type]
    valid_rows: list[dict] = []
    failed_rows: list[dict] = []
    if sub_tab_type == "worker" and "previous_attendance_details" in frame.columns:
        frame["previous_attendance_details"] = frame["previous_attendance_details"].fillna(0)
        frame.loc[frame["previous_attendance_details"].astype(str).str.strip() == "", "previous_attendance_details"] = 0
    actual_frame = actual_rows_only(frame)
    for index, raw_row in actual_frame.iterrows():
        if not row_has_owner_data(raw_row, expected):
            continue
        row = apply_bulk_numeric_defaults(
            sub_tab_type,
            {key: normalize_bulk_cell(key, raw_row.get(key)) for key in expected},
        )
        if not strict_validation and not owner_friendly and sub_tab_type == "blank_stock":
            if is_blank_bulk_value(row.get("variety_design")):
                row["variety_design"] = bulk_str(row.get("material_name"))
                if compatibility_warnings is not None:
                    compatibility_warnings.append({
                        "sheet": sheet_name or sub_tab_type,
                        "row": int(index) + 1,
                        "error": "Blank variety missing; defaulted to material_name.",
                        "values": {"variety_design": row["variety_design"]},
                    })
        if not strict_validation and not owner_friendly and sub_tab_type == "bottom_reel":
            if is_blank_bulk_value(row.get("variety_design")):
                row["variety_design"] = "Plain White"
                if compatibility_warnings is not None:
                    compatibility_warnings.append({
                        "sheet": sheet_name or sub_tab_type,
                        "row": int(index) + 1,
                        "error": "Bottom variety missing; defaulted to Plain White.",
                        "values": {"variety_design": "Plain White"},
                    })
        try:
            for row_variant in expand_bulk_row_variants(sub_tab_type, row):
                if strict_validation and sub_tab_type == "blank_stock" and is_blank_bulk_value(
                    row_variant.get("linked_bottom_size_mm")
                ):
                    raise ValueError("linked_bottom_size_mm is required in strict validation mode")
                validated_row = model.model_validate(row_variant).model_dump()
                if sub_tab_type == "blank_stock" and "total_boras_sacks" not in headers:
                    validated_row.pop("total_boras_sacks", None)
                validated_row["_row_number"] = int(index) + 1
                valid_rows.append(validated_row)
        except Exception as exc:
            if sub_tab_type == "plastic_stock":
                failed_rows.append({
                    "sheet": sheet_name or sub_tab_type,
                    "row": int(index) + 1,
                    "field": "used_for_cup_size_ml",
                    "error": "Cup size can be written as 210 or 210,250,300.",
                    "suggested_correction": "Enter at least one valid cup size, for example 210 or 55 ml, 65 ml.",
                    "severity": ValidationSeverity.FATAL.value,
                    "action_type": "error",
                    "values": row,
                    "entity_type": sub_tab_type,
                })
                continue
            failed_rows.append({
                "sheet": sheet_name or sub_tab_type,
                "row": int(index) + 1,
                "error": str(exc),
                "values": row,
                "entity_type": sub_tab_type,
            })
    return valid_rows, failed_rows


def bulk_unique_key(sub_tab_type: str, row: dict) -> tuple:
    if sub_tab_type == "company_profile":
        return ("company_profile",)
    if sub_tab_type == "worker":
        restore_key = normalized_identity(row.get("worker_restore_key"))
        return ("restore", restore_key) if restore_key else (
            "fallback",
            normalized_phone(row.get("mobile_number")) or normalized_identity(row.get("name")),
        )
    if sub_tab_type == "customer":
        restore_key = normalized_identity(row.get("customer_restore_key"))
        if restore_key:
            return ("restore", restore_key)
        return (
            "fallback",
            normalized_phone(row.get("phone_number") or row.get("contact_number")),
            normalized_identity(row.get("gst_number")),
            normalized_identity(row.get("name")),
            normalized_identity(row.get("firm_name")),
            normalized_identity(row.get("place")),
        )
    if sub_tab_type == "machine":
        restore_key = normalized_identity(row.get("machine_restore_key"))
        return ("restore", restore_key) if restore_key else (
            "fallback",
            normalized_identity(row.get("machine_number")) or normalized_identity(row.get("machine_name")),
        )
    if sub_tab_type == "blank_stock":
        restore_key = normalized_identity(row.get("material_restore_key"))
        return ("restore", restore_key) if restore_key else (
            "fallback",
            int(row["size_ml"]),
            normalized_identity(row.get("variety_design")),
        )
    if sub_tab_type == "bottom_reel":
        restore_key = normalized_identity(row.get("material_restore_key"))
        return ("restore", restore_key) if restore_key else (
            "fallback",
            int(row["bottom_size_mm"]),
            normalized_identity(row.get("variety_design")),
        )
    if sub_tab_type == "box_stock":
        return (bulk_str(row.get("box_type")).lower(),)
    if sub_tab_type == "plastic_stock":
        return (bulk_str(row.get("plastic_size_type")).lower(), int(row["used_for_cup_size_ml"]))
    if sub_tab_type == "finished_goods":
        restore_key = normalized_identity(row.get("product_restore_key"))
        if restore_key:
            return ("restore", restore_key)
        product_size_ml = int(row["product_size_ml"])
        return (
            "fallback",
            product_size_ml,
            normalized_identity(row.get("variety_design")),
            normalized_identity(row.get("packaging_size_name")),
        )
    return tuple(sorted((key, str(value)) for key, value in row.items() if not key.startswith("_")))


def dedupe_valid_bulk_rows(valid_by_type: dict[str, list[dict]]) -> tuple[dict[str, list[dict]], list[ValidationIssue]]:
    deduped: dict[str, list[dict]] = {key: [] for key in valid_by_type}
    warnings: list[ValidationIssue] = []
    sheet_names = {sub_tab_type: sheet_name for sheet_name, sub_tab_type in BULK_MASTER_SHEETS.items()}
    sheet_names.update({
        "blank_stock": "Raw Materials",
        "bottom_reel": "Raw Materials",
        "box_stock": "Raw Materials",
        "plastic_stock": "Raw Materials",
    })

    for sub_tab_type, rows in valid_by_type.items():
        by_key: dict[tuple, dict] = {}
        duplicate_rows: list[int] = []
        for row in rows:
            key = bulk_unique_key(sub_tab_type, row)
            if key in by_key:
                previous_row = by_key[key]
                duplicate_rows.extend([
                    int(previous_row.get("_row_number") or 0),
                    int(row.get("_row_number") or 0),
                ])
            by_key[key] = row
        if duplicate_rows:
            unique_rows = sorted({row for row in duplicate_rows if row > 0})
            warnings.append(ValidationIssue(
                row=max(unique_rows),
                field="duplicate_rows",
                error=f"{len(unique_rows)} duplicate rows were combined; the last value was used.",
                severity=ValidationSeverity.WARNING,
                suggested_correction=f"Check rows: {', '.join(map(str, unique_rows[:12]))}.",
                sheet=sheet_names.get(sub_tab_type, sub_tab_type),
                raw_value=", ".join(map(str, unique_rows[:12])),
                action_type="updated",
            ))
        deduped[sub_tab_type] = list(by_key.values())
    return deduped, warnings


def auto_normalize_owner_mappings(
    valid_by_type: dict[str, list[dict]],
) -> list[ValidationIssue]:
    fixes: list[ValidationIssue] = []
    blanks_by_size: dict[int, list[dict]] = {}
    bottoms_by_size: dict[int, list[dict]] = {}
    for row in valid_by_type.get("blank_stock", []):
        blanks_by_size.setdefault(int(row["size_ml"]), []).append(row)
    for row in valid_by_type.get("bottom_reel", []):
        if not normalized_identity(row.get("variety_design")):
            row["variety_design"] = "Plain White"
            fixes.append(ValidationIssue(
                row=row.get("_row_number"),
                field="variety_design",
                error=(
                    "Bottom size exists but variety was missing, auto-defaulted to Plain White. / "
                    "Bottom size मिला, variety खाली थी इसलिए Plain White रखा गया।"
                ),
                severity=ValidationSeverity.INFO,
                suggested_correction="No action needed / कोई बदलाव ज़रूरी नहीं।",
                sheet="Bottom_Reel",
                action_type="updated",
            ))
        bottoms_by_size.setdefault(int(row["bottom_size_mm"]), []).append(row)

    for row in valid_by_type.get("finished_goods", []):
        if not normalized_identity(row.get("variety_design")):
            candidates = blanks_by_size.get(int(row["product_size_ml"]), [])
            varieties = {bulk_str(item.get("variety_design")) for item in candidates if bulk_str(item.get("variety_design"))}
            if len(varieties) == 1:
                row["variety_design"] = next(iter(varieties))
                fixes.append(ValidationIssue(
                    row=row.get("_row_number"), field="variety_design",
                    error="Product design was filled from the only matching Cup Blank.",
                    severity=ValidationSeverity.INFO,
                    suggested_correction="No action needed / कोई बदलाव जरूरी नहीं।",
                    sheet="Finished Goods", action_type="updated",
                ))

    for row in valid_by_type.get("blank_stock", []):
        if normalized_identity(row.get("variety_design")):
            continue
        row["variety_design"] = bulk_str(row.get("material_name")) or "Cup Blank"
        fixes.append(ValidationIssue(
            row=row.get("_row_number"), field="variety_design",
            error="Cup Blank design was filled from its material name.",
            severity=ValidationSeverity.INFO,
            suggested_correction="No action needed / कोई बदलाव जरूरी नहीं।",
            sheet="Cup_Blank", action_type="updated",
        ))

    # Re-check products after blank designs have been safely inferred.
    for row in valid_by_type.get("finished_goods", []):
        if normalized_identity(row.get("variety_design")):
            continue
        candidates = blanks_by_size.get(int(row["product_size_ml"]), [])
        varieties = {
            bulk_str(item.get("variety_design"))
            for item in candidates
            if bulk_str(item.get("variety_design"))
        }
        if len(varieties) == 1:
            row["variety_design"] = next(iter(varieties))
            fixes.append(ValidationIssue(
                row=row.get("_row_number"),
                field="variety_design",
                error="Product design was filled from the only matching Cup Blank.",
                severity=ValidationSeverity.INFO,
                suggested_correction="No action needed / कोई बदलाव ज़रूरी नहीं।",
                sheet="Finished Goods",
                action_type="updated",
            ))

    return fixes


def validate_bulk_cross_sheet(
    valid_by_type: dict[str, list[dict]],
    *,
    strict_validation: bool = True,
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    severity = ValidationSeverity.FATAL if strict_validation else ValidationSeverity.WARNING
    action_type = "error" if strict_validation else "unchanged"
    boxes = {
        normalize_carton_type(row.get("box_type")): row
        for row in valid_by_type.get("box_stock", [])
    }
    bottom_sizes = {
        int(row["bottom_size_mm"])
        for row in valid_by_type.get("bottom_reel", [])
    }
    blanks = {
        (int(row["size_ml"]), normalized_identity(row.get("variety_design")))
        for row in valid_by_type.get("blank_stock", [])
    }
    for row in valid_by_type.get("finished_goods", []):
        carton_type = (row.get("carton_type") or "").strip()
        product_size = int(row["product_size_ml"])
        product_variety = normalized_identity(row.get("variety_design"))
        sku = (product_size, product_variety)
        if not carton_type:
            issues.append(ValidationIssue(
                row=row.get("_row_number"), field="carton_type",
                error="Select carton_type: Small Box or Big Box.",
                severity=severity,
                suggested_correction="Select the Box Stock carton type used by this finished product.",
                sheet="Finished Goods", section="Finished Goods",
                raw_value={"carton_type": carton_type, "product_size_ml": product_size}, action_type=action_type,
            ))
        else:
            matched_box = boxes.get(normalize_carton_type(carton_type))
            allowed_sizes = parse_allowed_sizes(
                matched_box.get("size_for_finished_product")
            ) if matched_box else []
            logger.debug(
                "Matched carton_type=%s Allowed sizes=%s Product size=%s",
                matched_box.get("box_type") if matched_box else None,
                allowed_sizes,
                product_size,
            )
            if matched_box is None or product_size not in allowed_sizes:
                issues.append(ValidationIssue(
                    row=row.get("_row_number"), field="product_size_ml",
                    error=(
                        f"This carton type is not configured for product size {product_size}ml. "
                        f"Add {product_size} to Size For Finished Product for {carton_type}."
                    ),
                    severity=severity,
                    suggested_correction=(
                        f"Matched Box Stock carton: {matched_box.get('box_type') if matched_box else 'none'}; "
                        f"Allowed sizes: {', '.join(str(size) for size in allowed_sizes) if allowed_sizes else 'none'}."
                    ),
                    sheet="Finished Goods", section="Finished Goods",
                    raw_value={
                        "carton_type": carton_type,
                        "product_size_ml": product_size,
                        "matched_box_stock_carton_type": matched_box.get("box_type") if matched_box else None,
                        "allowed_sizes": allowed_sizes,
                    },
                    action_type=action_type,
                ))
        if sku not in blanks:
            available_varieties = sorted({
                variety
                for size, variety in blanks
                if size == product_size and variety
            })
            if not product_variety and len(available_varieties) > 1:
                error = "More than one Cup Blank design exists for this size. Please choose the product design."
                correction = (
                    f"Choose one design for {product_size} ml: "
                    f"{', '.join(available_varieties)}."
                )
            else:
                error = "Matching Cup Blank was not found for this product size and design."
                correction = (
                    f"Add Cup Blank row for {product_size} ml "
                    f"{bulk_str(row.get('variety_design')).lower() or 'design'}"
                )
            issues.append(ValidationIssue(
                row=row.get("_row_number"), field="variety_design",
                error=error,
                severity=severity,
                suggested_correction=correction,
                sheet="Finished Goods", section="Finished Goods",
                raw_value=row.get("variety_design"), action_type=action_type,
            ))

    for row in valid_by_type.get("blank_stock", []):
        linked_value = row.get("linked_bottom_size_mm")
        if linked_value is None:
            issues.append(ValidationIssue(
                row=row.get("_row_number"), field="linked_bottom_size_mm",
                error="Cup Blank is not linked to a Bottom Reel size. / Cup Blank का Bottom Reel size नहीं जुड़ा है।",
                severity=severity,
                suggested_correction="Fill Linked Bottom Size MM in Cup_Blank.",
                sheet="Raw Materials", section="Cup Blank",
                raw_value=None, action_type=action_type,
            ))
            continue
        linked_size = int(linked_value)
        if linked_size not in bottom_sizes:
            issues.append(ValidationIssue(
                row=row.get("_row_number"), field="linked_bottom_size_mm",
                error="Matching Bottom Reel was not found for this Cup Blank. / इस Cup Blank का Bottom Reel नहीं मिला।",
                severity=severity,
                suggested_correction=f"Add Bottom_Reel row for {linked_size} mm.",
                sheet="Raw Materials", section="Cup Blank",
                raw_value=row.get("linked_bottom_size_mm"), action_type=action_type,
            ))

    for row in valid_by_type.get("machine", []):
        if row.get("bottom_size_mm") and int(row["bottom_size_mm"]) not in bottom_sizes:
            issues.append(ValidationIssue(
                row=row.get("_row_number"), field="bottom_size_mm",
                error="This machine's bottom size is missing from Bottom Reel stock. / Machine का bottom size stock में नहीं मिला।",
                severity=severity,
                suggested_correction=f"Add Bottom Reel size {row.get('bottom_size_mm')} mm.",
                sheet="Machines", section="Machines",
                raw_value=row.get("bottom_size_mm"), action_type=action_type,
            ))
    return issues


def increment_bulk_stat(stats: dict[str, int] | None, key: str, value: int = 1) -> None:
    if stats is None:
        return
    stats[key] = int(stats.get(key, 0)) + value


def transfer_restore_key(
    db: Session,
    model,
    *,
    factory_id: int,
    column,
    restore_key: str | None,
    target_id: int | None,
) -> None:
    if not restore_key:
        return
    stale_rows = (
        db.query(model)
        .filter(
            model.factory_id == factory_id,
            sql_func.lower(column) == restore_key.lower(),
        )
        .with_for_update()
        .all()
    )
    for stale_row in stale_rows:
        if target_id is None or stale_row.id != target_id:
            setattr(stale_row, column.key, None)
    if any(target_id is None or stale_row.id != target_id for stale_row in stale_rows):
        db.flush()


def read_standard_sheet(
    workbook: dict,
    sheet_name: str,
    sub_tab_type: str,
    *,
    strict_validation: bool,
    compatibility_warnings: list[dict],
) -> tuple[list[dict], list[dict]]:
    if sheet_name not in workbook:
        return [], [{"sheet": sheet_name, "row": None, "error": "Required worksheet is missing"}]
    raw_frame = workbook[sheet_name]
    instruction = bulk_str(raw_frame.iat[0, 0]) if not raw_frame.empty else ""
    if not instruction:
        return [], [{"sheet": sheet_name, "row": 1, "error": "Instruction row is required"}]
    if len(raw_frame.index) < 2:
        return [], [{"sheet": sheet_name, "row": 2, "error": "Header row is missing"}]
    headers = [canonical_bulk_header(value) for value in raw_frame.iloc[1].tolist()]
    frame = raw_frame.iloc[2:].copy()
    frame.columns = headers
    frame = frame.loc[:, [column for column in frame.columns if column]]
    return validate_bulk_frame(
        frame,
        sub_tab_type,
        sheet_name,
        row_offset=2,
        strict_validation=strict_validation,
        compatibility_warnings=compatibility_warnings,
    )


def find_raw_section_label_rows(raw_frame) -> dict[str, int]:
    label_rows: dict[str, int] = {}
    for row_index in range(len(raw_frame.index)):
        row_text = " ".join(bulk_str(value) for value in raw_frame.iloc[row_index].tolist()).upper()
        if not row_text:
            continue
        for sub_tab_type, section in RAW_MATERIAL_SECTIONS.items():
            marker = str(section["marker"]).upper()
            if sub_tab_type not in label_rows and marker in row_text:
                label_rows[sub_tab_type] = row_index
    return label_rows


def read_raw_material_section(
    raw_frame,
    sub_tab_type: str,
    *,
    strict_validation: bool,
    compatibility_warnings: list[dict],
) -> tuple[list[dict], list[dict]]:
    section = RAW_MATERIAL_SECTIONS[sub_tab_type]
    label_rows = find_raw_section_label_rows(raw_frame)
    if sub_tab_type not in label_rows:
        return [], [{"sheet": "Raw Materials", "row": None, "error": f"Missing section marker containing: {section['marker']}"}]

    sorted_label_rows = sorted(label_rows.values())
    label_index = label_rows[sub_tab_type]
    next_label_index = next((row for row in sorted_label_rows if row > label_index), len(raw_frame.index))
    header_index = label_index + 1
    start_index = header_index + 1
    end_index = next_label_index

    if len(raw_frame.index) <= header_index:
        return [], [{"sheet": "Raw Materials", "row": label_index + 2, "error": "Header row is missing"}]

    headers = [canonical_bulk_header(value) for value in raw_frame.iloc[header_index].tolist()]
    frame = raw_frame.iloc[start_index:end_index].copy()
    frame.columns = headers
    frame = frame.loc[:, [column for column in frame.columns if column]]
    return validate_bulk_frame(
        frame,
        sub_tab_type,
        "Raw Materials",
        row_offset=start_index,
        strict_validation=strict_validation,
        compatibility_warnings=compatibility_warnings,
    )


def read_owner_friendly_sheet(
    raw_frame,
    sub_tab_type: str,
    sheet_name: str,
    compatibility_warnings: list[dict],
) -> tuple[list[dict], list[dict]]:
    expected = BULK_TEMPLATE_COLUMNS[sub_tab_type]
    header_index = None
    headers: list[str] = []
    for index, raw_row in raw_frame.head(10).iterrows():
        candidate = []
        for value in raw_row.tolist():
            raw_header = canonical_bulk_header(value)
            candidate.append(OWNER_HEADER_ALIASES.get(sub_tab_type, {}).get(raw_header, raw_header))
        recognized = [header for header in candidate if header in expected]
        if len(recognized) >= 1:
            header_index = int(index)
            headers = candidate
            break
    if header_index is None:
        return [], [{"sheet": sheet_name, "row": 1, "error": "Could not find a recognized header row"}]

    frame = raw_frame.iloc[header_index + 1:].copy()
    frame.columns = headers
    frame = frame.loc[:, [column for column in frame.columns if column]]
    frame = canonicalize_bulk_frame(frame)
    if "row_type" not in frame.columns:
        frame.insert(0, "row_type", "ACTUAL")
    else:
        frame["row_type"] = frame["row_type"].fillna("ACTUAL")
        frame.loc[frame["row_type"].astype(str).str.strip() == "", "row_type"] = "ACTUAL"
    for column in expected:
        if column not in frame.columns:
            frame[column] = None

    owner_defaults = {
        "company_profile": {
            "invoice_prefix": "INV-",
            "advance_upi_discount": Decimal("2"),
            "bill_of_supply_start_seq": 1,
            "tax_invoice_start_seq": 1,
            "bill_of_supply_simple_start_seq": 1,
        },
        "machine": {
            "machine_type": "Paper Cup",
            "default_operating_speed": 0,
            "target_output_per_shift": 0,
        },
        "blank_stock": {"material_name": "Cup Blank"},
    }
    for column, default in owner_defaults.get(sub_tab_type, {}).items():
        frame[column] = frame[column].where(frame[column].notna(), default)
        frame.loc[frame[column].astype(str).str.strip() == "", column] = default

    return validate_bulk_frame(
        frame,
        sub_tab_type,
        sheet_name,
        row_offset=header_index + 2,
        strict_validation=False,
        compatibility_warnings=compatibility_warnings,
        owner_friendly=True,
    )


def read_owner_friendly_excel(workbook: dict) -> tuple[dict[str, list[dict]], list[dict]]:
    valid_by_type: dict[str, list[dict]] = {key: [] for key in BULK_ROW_MODELS}
    failed_rows: list[dict] = []
    compatibility_warnings: list[dict] = []
    normalized_sheets = {canonical_bulk_header(name): name for name in workbook}

    for owner_sheet, sub_tab_type in OWNER_FRIENDLY_SHEETS.items():
        actual_name = normalized_sheets.get(canonical_bulk_header(owner_sheet))
        if sub_tab_type == "costing_optional":
            continue
        if actual_name is None:
            failed_rows.append({"sheet": owner_sheet, "row": None, "error": "Required worksheet is missing"})
            continue
        rows, errors = read_owner_friendly_sheet(
            workbook[actual_name],
            sub_tab_type,
            actual_name,
            compatibility_warnings,
        )
        valid_by_type[sub_tab_type] = rows
        failed_rows.extend(errors)

    auto_issues = auto_normalize_owner_mappings(valid_by_type)
    generate_owner_restore_keys(valid_by_type)
    failed_rows.extend([
        {
            "sheet": issue.sheet,
            "row": issue.row,
            "field": issue.field,
            "error": issue.error,
            "severity": issue.severity.value,
            "suggested_correction": issue.suggested_correction,
            "action_type": issue.action_type,
        }
        for issue in auto_issues
    ])
    failed_rows.extend(compatibility_warnings)
    return valid_by_type, failed_rows


def detect_master_template_version(file_bytes: bytes) -> int:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3, values_only=True):
                for value in row:
                    match = re.search(r"template_version\s*[=:]\s*(\d+)", bulk_str(value), re.IGNORECASE)
                    if match:
                        return int(match.group(1))
    except Exception:
        return 1
    return 1


def is_owner_friendly_workbook(file_bytes: bytes) -> bool:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(file_bytes), read_only=True)
        names = {canonical_bulk_header(name) for name in workbook.sheetnames}
        expected = {canonical_bulk_header(name) for name in OWNER_FRIENDLY_SHEETS}
        return len(names & expected) >= 5
    except Exception:
        return False


def read_master_bulk_excel(
    file_bytes: bytes,
    *,
    strict_validation: bool | None = None,
) -> tuple[dict[str, list[dict]], list[dict]]:
    try:
        import pandas as pd
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Excel parser dependencies are not installed") from exc

    try:
        workbook = pd.read_excel(BytesIO(file_bytes), sheet_name=None, header=None, dtype=object)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to read Excel file: {exc}") from exc

    normalized_sheet_names = {canonical_bulk_header(name) for name in workbook}
    owner_sheet_names = {canonical_bulk_header(name) for name in OWNER_FRIENDLY_SHEETS}
    if len(normalized_sheet_names & owner_sheet_names) >= 5:
        return read_owner_friendly_excel(workbook)

    template_version = detect_master_template_version(file_bytes)
    effective_strict_validation = bool(strict_validation) or template_version >= 2
    valid_by_type: dict[str, list[dict]] = {key: [] for key in BULK_ROW_MODELS}
    failed_rows: list[dict] = []
    compatibility_warnings: list[dict] = []

    for sheet_name, sub_tab_type in BULK_MASTER_SHEETS.items():
        if sub_tab_type == "raw_materials":
            if sheet_name not in workbook:
                failed_rows.append({"sheet": sheet_name, "row": None, "error": "Required worksheet is missing"})
                continue
            for raw_sub_type in RAW_MATERIAL_SECTIONS:
                valid_rows, sheet_errors = read_raw_material_section(
                    workbook[sheet_name],
                    raw_sub_type,
                    strict_validation=effective_strict_validation,
                    compatibility_warnings=compatibility_warnings,
                )
                valid_by_type[raw_sub_type] = valid_rows
                failed_rows.extend(sheet_errors)
            continue

        valid_rows, sheet_errors = read_standard_sheet(
            workbook,
            sheet_name,
            sub_tab_type,
            strict_validation=effective_strict_validation,
            compatibility_warnings=compatibility_warnings,
        )
        valid_by_type[sub_tab_type] = valid_rows
        failed_rows.extend(sheet_errors)
    if not effective_strict_validation:
        bottom_by_cup_size: dict[int, set[int]] = {}
        for machine in valid_by_type.get("machine", []):
            if machine.get("mould_size_ml") and machine.get("bottom_size_mm"):
                bottom_by_cup_size.setdefault(int(machine["mould_size_ml"]), set()).add(int(machine["bottom_size_mm"]))
        for blank in valid_by_type.get("blank_stock", []):
            if blank.get("linked_bottom_size_mm") is not None:
                continue
            matches = bottom_by_cup_size.get(int(blank["size_ml"]), set())
            if len(matches) == 1:
                blank["linked_bottom_size_mm"] = next(iter(matches))
                message = "Blank linked bottom size missing; defaulted from the unique matching machine."
            else:
                message = "Blank linked bottom size missing; mapping left incomplete for production."
            compatibility_warnings.append({
                "sheet": "Raw Materials",
                "row": blank.get("_row_number"),
                "error": message,
                "values": {"linked_bottom_size_mm": blank.get("linked_bottom_size_mm")},
            })
    failed_rows.extend(compatibility_warnings)
    return valid_by_type, failed_rows


def inspect_finished_goods_sheet(file_bytes: bytes) -> dict:
    import pandas as pd

    debug_info = {
        "sheet_name": None,
        "normalized_headers": [],
        "rows_read": 0,
        "rows_inserted": 0,
        "rows_updated": 0,
        "rows_skipped": 0,
        "skip_reasons": [],
        "created_finished_goods_stock_ids": [],
        "created_final_product_stock_ids": [],
        "matched_existing_final_product_stock_ids": [],
    }
    workbook = pd.read_excel(BytesIO(file_bytes), sheet_name=None, header=None, dtype=object)
    sheet_name = next(
        (name for name in workbook if canonical_bulk_header(name) == "finished_goods"),
        None,
    )
    if sheet_name is None:
        available = ", ".join(str(name) for name in workbook) or "none"
        debug_info["skip_reasons"].append(
            f"Finished Goods sheet not imported because the worksheet was not found. Available sheets: {available}"
        )
        return debug_info

    debug_info["sheet_name"] = sheet_name
    frame = workbook[sheet_name]
    if len(frame.index) < 2:
        debug_info["skip_reasons"].append(
            "Finished Goods sheet not imported because the header row is missing"
        )
        return debug_info

    debug_info["normalized_headers"] = [
        header for header in (canonical_bulk_header(value) for value in frame.iloc[1].tolist()) if header
    ]
    return debug_info


def build_master_onboarding_workbook() -> BytesIO:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, sub_tab_type in BULK_MASTER_SHEETS.items():
            if sub_tab_type == "raw_materials":
                pd.DataFrame().to_excel(writer, index=False, header=False, sheet_name=sheet_name)
                continue
            frame = pd.DataFrame(
                [
                    ["Instruction: template_version=2; keep row_type as SAMPLE for examples and ACTUAL for rows to import."],
                    BULK_TEMPLATE_COLUMNS[sub_tab_type],
                    SAMPLE_BULK_ROWS[sub_tab_type],
                    *[[None] * len(BULK_TEMPLATE_COLUMNS[sub_tab_type]) for _ in range(10)],
                ]
            )
            frame.to_excel(writer, index=False, header=False, sheet_name=sheet_name)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="EDE9FE")
        section_fill = PatternFill("solid", fgColor="DCFCE7")
        section_font = Font(bold=True, color="14532D")
        header_font = Font(bold=True, color="111827")

        for sheet_name, sub_tab_type in BULK_MASTER_SHEETS.items():
            worksheet = workbook[sheet_name]
            if sub_tab_type == "raw_materials":
                for raw_sub_type, section in RAW_MATERIAL_SECTIONS.items():
                    label_row = int(section["label_row"])
                    header_row = int(section["header_row"])
                    data_start = int(section["data_start"])
                    worksheet.cell(row=label_row, column=1, value=section["title"])
                    worksheet.cell(row=label_row, column=1).fill = section_fill
                    worksheet.cell(row=label_row, column=1).font = section_font
                    for column_index, header in enumerate(BULK_TEMPLATE_COLUMNS[raw_sub_type], start=1):
                        cell = worksheet.cell(row=header_row, column=column_index, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                    for column_index, value in enumerate(SAMPLE_BULK_ROWS[raw_sub_type], start=1):
                        worksheet.cell(row=data_start, column=column_index, value=value)
                for column_index in range(1, 9):
                    worksheet.column_dimensions[get_column_letter(column_index)].width = 24
                continue

            worksheet.cell(row=1, column=1).fill = section_fill
            worksheet.cell(row=1, column=1).font = section_font
            for column_index in range(1, len(BULK_TEMPLATE_COLUMNS[sub_tab_type]) + 1):
                worksheet.cell(row=2, column=column_index).fill = header_fill
                worksheet.cell(row=2, column=column_index).font = header_font
                worksheet.column_dimensions[get_column_letter(column_index)].width = 24
    output.seek(0)
    return output


def build_owner_onboarding_workbook() -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="EDE9FE")
    for sheet_name, headers in OWNER_TEMPLATE_COLUMNS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(index)].width = max(18, len(header) + 4)
        for _ in range(12):
            sheet.append([None] * len(headers))
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def log_bulk_upload(background_tasks: BackgroundTasks, db: Session, current_user: User, sub_tab_type: str, row_count: int) -> None:
    background_tasks.add_task(
        log_activity,
        db,
        int(current_user.factory_id),
        current_user.id,
        current_user.full_name or current_user.username,
        current_user.role,
        "BULK_UPLOAD_COMPLETED",
        f"System: {current_user.role} completed bulk upload of {row_count} entries in {sub_tab_type}",
        sub_tab_type,
        None,
        {"row_count": row_count, "sub_tab_type": sub_tab_type},
    )


def log_bulk_inventory_uploads(
    background_tasks: BackgroundTasks,
    db: Session,
    current_user: User,
    inserted_counts: dict[str, int],
) -> None:
    log_specs = {
        "blank_stock": (
            "RAW_MATERIAL_ADDED",
            "Raw material bulk upload completed - {count} blank stock entries",
            "raw_material",
        ),
        "bottom_reel": (
            "RAW_MATERIAL_ADDED",
            "Raw material bulk upload completed - {count} bottom reel entries",
            "raw_material",
        ),
        "box_stock": (
            "PACKAGING_MATERIAL_ADDED",
            "Packaging material bulk upload completed - {count} box stock entries",
            "packaging_material",
        ),
        "plastic_stock": (
            "PACKAGING_MATERIAL_ADDED",
            "Packaging material bulk upload completed - {count} plastic stock entries",
            "packaging_material",
        ),
        "finished_goods": (
            "FINISHED_GOODS_STOCK_SAVED",
            "Finished goods bulk upload completed - {count} stock entries",
            "finished_goods_stock",
        ),
    }
    for sub_tab_type, (action_type, summary_template, entity_type) in log_specs.items():
        row_count = int(inserted_counts.get(sub_tab_type) or 0)
        if row_count <= 0:
            continue
        background_tasks.add_task(
            log_activity,
            db,
            int(current_user.factory_id),
            current_user.id,
            current_user.full_name or current_user.username,
            current_user.role,
            action_type,
            summary_template.format(count=row_count),
            entity_type,
            None,
            {"row_count": row_count, "sub_tab_type": sub_tab_type},
        )


def sync_finished_goods_to_final_product_stock(
    db: Session,
    factory_id: int,
    stock: FinishedGoodsStock,
    fg_debug_info: dict | None = None,
) -> FinalProductStock:
    profile = (
        db.query(PackagingProfile)
        .filter(
            PackagingProfile.factory_id == factory_id,
            PackagingProfile.id == stock.packaging_profile_id,
        )
        .first()
    )
    if profile is None:
        raise ValueError(
            f"Finished Goods sheet not imported because packaging profile {stock.packaging_profile_id} "
            f"does not belong to factory {factory_id}"
        )

    variety = (stock.variant_name or "Standard/White").strip() or "Standard/White"
    packaging_size_name = profile.profile_name.strip()
    final_stock = (
        db.query(FinalProductStock)
        .filter(
            FinalProductStock.factory_id == factory_id,
            FinalProductStock.product_size_ml == stock.cup_size_ml,
            sql_func.lower(FinalProductStock.variety) == variety.lower(),
            sql_func.lower(FinalProductStock.packaging_size_name) == packaging_size_name.lower(),
        )
        .with_for_update()
        .first()
    )
    if final_stock is None:
        final_stock = FinalProductStock(
            factory_id=factory_id,
            product_size_ml=stock.cup_size_ml,
            variety=variety,
            packaging_size_name=packaging_size_name,
            carton_type=profile.box_size_name,
            pieces_per_packet=profile.cups_per_poly or 1,
            packets_per_box_limit=profile.polys_per_box or 1,
            current_quantity=stock.boxes_available or 0,
            total_boxes=stock.boxes_available or 0,
            loose_packets=0,
        )
        db.add(final_stock)
        db.flush()
        if fg_debug_info is not None:
            fg_debug_info["created_final_product_stock_ids"].append(final_stock.id)
    else:
        final_stock.pieces_per_packet = profile.cups_per_poly or 1
        final_stock.carton_type = profile.box_size_name
        final_stock.packets_per_box_limit = profile.polys_per_box or 1
        final_stock.current_quantity = stock.boxes_available or 0
        final_stock.total_boxes = stock.boxes_available or 0
        db.flush()
        if fg_debug_info is not None:
            fg_debug_info["matched_existing_final_product_stock_ids"].append(final_stock.id)
    return final_stock


def apply_bulk_rows(db: Session, current_user: User, sub_tab_type: str, valid_rows: list[dict], stats: dict[str, int] | None = None, fg_debug_info: dict | None = None) -> int:
    factory_id = int(current_user.factory_id)
    if not valid_rows:
        return 0

    if sub_tab_type == "company_profile":
        row = valid_rows[0]
        factory = db.query(Factory).filter(Factory.id == factory_id).with_for_update().first()
        if factory is None:
            raise HTTPException(status_code=404, detail="Factory not found")
        settings = get_or_create_factory_settings(db, factory_id)
        factory.factory_name = row["factory_name"].strip()
        factory.name = factory.name or row["factory_name"].strip()
        factory.gst_number = (row.get("gstin") or "").strip() or None
        factory.address = (row.get("factory_address") or "").strip() or None
        factory.invoice_prefix = (row.get("invoice_prefix") or "INV-").strip() or "INV-"
        factory.advance_payment_discount_percentage = row.get("advance_upi_discount") or Decimal("2.00")
        settings.bill_of_supply_start_seq = row["bill_of_supply_start_seq"]
        settings.tax_invoice_start_seq = row["tax_invoice_start_seq"]
        settings.bill_of_supply_simple_start_seq = row["bill_of_supply_simple_start_seq"]
        factory.next_bill_of_supply_number = row["bill_of_supply_start_seq"]
        factory.next_tax_invoice_number = row["tax_invoice_start_seq"]
        factory.next_bill_of_supply_simple_number = row["bill_of_supply_simple_start_seq"]
        increment_bulk_stat(stats, "updated")
        return 1

    if sub_tab_type == "worker":
        worker_rows: list[tuple[Worker, dict]] = []
        pending_workers: dict[tuple[str, str], Worker] = {}
        for row in valid_rows:
            worker_name = row["name"].strip()
            if not worker_name:
                increment_bulk_stat(stats, "skipped")
                continue
            phone, _ = normalize_phone_number(str(row["mobile_number"])) if row.get("mobile_number") else (None, None)
            restore_key = (row.get("worker_restore_key") or "").strip() or None
            pending_key = (
                ("restore", normalized_identity(restore_key))
                if restore_key
                else ("fallback", phone or normalized_identity(worker_name))
            )
            query = db.query(Worker).filter(Worker.factory_id == factory_id)
            worker = pending_workers.get(pending_key)
            if worker is not None:
                increment_bulk_stat(stats, "updated")
            elif phone:
                worker = query.filter(Worker.phone == phone).with_for_update().first()
                if worker is None:
                    worker = query.filter(sql_func.lower(sql_func.trim(Worker.name)) == worker_name.lower()).with_for_update().first()
            else:
                worker = query.filter(sql_func.lower(sql_func.trim(Worker.name)) == worker_name.lower()).with_for_update().first()
            if worker is None and restore_key:
                worker = query.filter(sql_func.lower(Worker.worker_restore_key) == restore_key.lower()).with_for_update().first()
            if worker is None:
                worker = Worker(factory_id=factory_id, name=worker_name)
                db.add(worker)
                increment_bulk_stat(stats, "inserted")
            elif pending_key not in pending_workers:
                increment_bulk_stat(stats, "updated")
            transfer_restore_key(
                db, Worker, factory_id=factory_id, column=Worker.worker_restore_key,
                restore_key=restore_key, target_id=worker.id,
            )
            pending_workers[pending_key] = worker
            worker.worker_restore_key = restore_key
            worker.phone = phone
            worker.daily_wage_rate = row["daily_wages"]
            worker.daily_wages = row["daily_wages"]
            worker.duty_hours = row["duty_hours"]
            worker.salary = worker.salary or 0
            worker.daily_salary = row["daily_wages"]
            worker.shift_hours = row["duty_hours"]
            worker.shift_timing = (row.get("shift_timing") or "").strip() or None
            worker.shift_type = (row.get("shift_type") or "").strip() or worker.shift_type or "worker"
            worker.is_active = True
            worker_rows.append((worker, row))
        db.flush()

        for worker, row in worker_rows:
            if row["previous_attendance_details"] <= 0 or not worker.id:
                continue
            opening_attendance = (
                db.query(WorkerOpeningAttendance)
                .filter(
                    WorkerOpeningAttendance.factory_id == factory_id,
                    WorkerOpeningAttendance.worker_id == worker.id,
                )
                .with_for_update()
                .first()
            )
            if opening_attendance is None:
                opening_attendance = WorkerOpeningAttendance(
                    factory_id=factory_id,
                    worker_id=worker.id,
                    created_by_user_id=current_user.id,
                )
                db.add(opening_attendance)
            opening_attendance.period_start = date.today()
            opening_attendance.period_end = date.today()
            opening_attendance.present_days = row["previous_attendance_details"]
            opening_attendance.half_days = 0
            opening_attendance.absent_days = 0
            opening_attendance.paid_leave_days = 0
            opening_attendance.overtime_hours = 0
            opening_attendance.advance_paid = 0
            opening_attendance.deductions = 0
            opening_attendance.notes = "Opening attendance imported by bulk upload"
        return len(worker_rows)

    if sub_tab_type == "customer":
        saved_count = 0
        for row in valid_rows:
            customer_name = row["name"].strip()
            if not customer_name:
                increment_bulk_stat(stats, "skipped")
                continue
            restore_key = (row.get("customer_restore_key") or "").strip() or None
            phone_number = normalized_phone(row.get("phone_number") or row.get("contact_number")) or None
            gst_number = (row.get("gst_number") or "").strip() or None
            query = db.query(Customer).filter(Customer.factory_id == factory_id)
            if phone_number:
                customer = query.filter(
                    sql_func.replace(sql_func.replace(sql_func.replace(Customer.phone_number, "+91", ""), " ", ""), "-", "")
                    .like(f"%{phone_number}")
                ).with_for_update().first()
                if customer is None and gst_number:
                    customer = query.filter(sql_func.lower(Customer.gst_number) == gst_number.lower()).with_for_update().first()
                if customer is None:
                    customer = query.filter(
                        sql_func.lower(sql_func.trim(Customer.name)) == customer_name.lower(),
                        sql_func.lower(sql_func.trim(sql_func.coalesce(Customer.firm_name, ""))) == normalized_identity(row.get("firm_name")),
                        sql_func.lower(sql_func.trim(sql_func.coalesce(Customer.place, ""))) == normalized_identity(row.get("place")),
                    ).with_for_update().first()
            elif gst_number:
                customer = query.filter(sql_func.lower(Customer.gst_number) == gst_number.lower()).with_for_update().first()
                if customer is None:
                    customer = query.filter(
                        sql_func.lower(sql_func.trim(Customer.name)) == customer_name.lower(),
                        sql_func.lower(sql_func.trim(sql_func.coalesce(Customer.firm_name, ""))) == normalized_identity(row.get("firm_name")),
                        sql_func.lower(sql_func.trim(sql_func.coalesce(Customer.place, ""))) == normalized_identity(row.get("place")),
                    ).with_for_update().first()
            else:
                customer = query.filter(
                    sql_func.lower(sql_func.trim(Customer.name)) == customer_name.lower(),
                    sql_func.lower(sql_func.trim(sql_func.coalesce(Customer.firm_name, ""))) == normalized_identity(row.get("firm_name")),
                    sql_func.lower(sql_func.trim(sql_func.coalesce(Customer.place, ""))) == normalized_identity(row.get("place")),
                ).with_for_update().first()
            if customer is None and restore_key:
                customer = query.filter(sql_func.lower(Customer.customer_restore_key) == restore_key.lower()).with_for_update().first()
            if customer is None:
                customer = Customer(factory_id=factory_id, name=customer_name)
                db.add(customer)
                increment_bulk_stat(stats, "inserted")
            else:
                increment_bulk_stat(stats, "updated")
            transfer_restore_key(
                db, Customer, factory_id=factory_id, column=Customer.customer_restore_key,
                restore_key=restore_key, target_id=customer.id,
            )

            contact_number = (row.get("contact_number") or "").strip() or None
            previous_due = Decimal(str(row.get("previous_due") or "0"))
            advance_balance = Decimal(str(row.get("advance_balance") or "0"))
            customer.customer_restore_key = restore_key
            customer.name = customer_name
            customer.firm_name = (row.get("firm_name") or "").strip() or None
            customer.contact_number = contact_number
            customer.phone_number = phone_number
            customer.phone = phone_number or contact_number
            customer.place = (row.get("place") or "").strip() or None
            customer.address = (row.get("address") or "").strip() or None
            customer.email = (row.get("email") or "").strip() or None
            customer.gst_number = gst_number
            customer.opening_outstanding_date = row.get("opening_outstanding_date")
            customer.opening_outstanding_note = (row.get("opening_outstanding_note") or "").strip() or None
            customer.previous_due = previous_due
            customer.advance_balance = advance_balance
            customer.advance_balance_date = row.get("advance_balance_date")
            customer.advance_balance_note = (row.get("advance_balance_note") or "").strip() or None
            customer.total_due = previous_due
            customer.pending_dues = float(previous_due)
            customer.pending_balance = previous_due
            customer.balance_amount = previous_due
            customer.is_active = True
            
            db.flush()
            if previous_due > 0:
                open_bill = (
                    db.query(OutstandingBill)
                    .filter(OutstandingBill.factory_id == factory_id)
                    .filter(OutstandingBill.customer_id == customer.id)
                    .filter(OutstandingBill.source_type.in_(("opening_balance", "opening_outstanding")))
                    .first()
                )
                if not open_bill:
                    create_outstanding_bill(
                        db,
                        factory_id=factory_id,
                        customer_id=customer.id,
                        source_type="opening_outstanding",
                        tracking_number=f"OPEN-{customer.id}",
                        bill_date=row.get("opening_outstanding_date") or date.today(),
                        bill_amount=previous_due,
                        amount_paid=Decimal("0.00"),
                        note=customer.opening_outstanding_note,
                    )
                else:
                    open_bill.bill_date = row.get("opening_outstanding_date") or open_bill.bill_date
                    open_bill.note = customer.opening_outstanding_note
                    open_bill.bill_amount = previous_due
                    open_bill.balance_amount = max(previous_due - Decimal(open_bill.amount_paid or 0), Decimal("0"))
                    open_bill.status = "paid" if open_bill.balance_amount == 0 else ("partial" if open_bill.amount_paid else "active")
            saved_count += 1
        db.flush()
        return saved_count

    if sub_tab_type == "machine":
        saved_count = 0
        for row in valid_rows:
            machine_name = row["machine_name"].strip()
            if not machine_name:
                increment_bulk_stat(stats, "skipped")
                continue
            restore_key = (row.get("machine_restore_key") or "").strip() or None
            machine_number = (row.get("machine_number") or "").strip() or None
            query = db.query(Machine).filter(Machine.factory_id == factory_id)
            if machine_number:
                machine = query.filter(sql_func.lower(Machine.machine_number) == machine_number.lower()).with_for_update().first()
            else:
                machine = query.filter(sql_func.lower(sql_func.trim(Machine.name)) == machine_name.lower()).with_for_update().first()
            if machine is None:
                machine = query.filter(sql_func.lower(sql_func.trim(Machine.name)) == machine_name.lower()).with_for_update().first()
            if machine is None and restore_key:
                machine = query.filter(sql_func.lower(Machine.machine_restore_key) == restore_key.lower()).with_for_update().first()
            if machine is None:
                machine = Machine(factory_id=factory_id, name=machine_name)
                db.add(machine)
                increment_bulk_stat(stats, "inserted")
            else:
                increment_bulk_stat(stats, "updated")
            transfer_restore_key(
                db, Machine, factory_id=factory_id, column=Machine.machine_restore_key,
                restore_key=restore_key, target_id=machine.id,
            )
            machine.name = machine_name
            machine.machine_restore_key = restore_key
            machine.machine_name = machine_name
            machine.machine_number = machine_number
            machine.machine_type = (row.get("machine_type") or "").strip() or "Custom Machine"
            machine.mould_size_ml = row.get("mould_size_ml")
            machine.bottom_size_mm = row.get("bottom_size_mm")
            machine.speed_per_minute = row["default_operating_speed"]
            machine.speed_bpm = row["default_operating_speed"]
            machine.speed_cups_per_minute = row["default_operating_speed"]
            machine.default_speed = row["default_operating_speed"]
            machine.target_output_per_shift = row["target_output_per_shift"]
            machine.raw_materials_mapped = ["blank_stock", "bottom_reel"]
            machine.is_active = True
            saved_count += 1
        db.flush()
        return saved_count

    if sub_tab_type == "blank_stock":
        # total_boras_sacks is optional so existing customer workbooks remain
        # valid. Missing values preserve the legacy zero-quantity behavior.
        saved_count = 0
        for row in valid_rows:
            blank_size_ml = int(row["size_ml"])
            restore_key = (row.get("material_restore_key") or "").strip() or None
            material_name = row["material_name"].strip()
            variety = row["variety_design"].strip()
            query = db.query(BlankStock).filter(BlankStock.factory_id == factory_id)
            stock = query.filter(
                BlankStock.blank_size_ml == blank_size_ml,
                sql_func.lower(BlankStock.variety) == variety.lower(),
            ).with_for_update().first()
            if stock is None and restore_key:
                stock = query.filter(sql_func.lower(BlankStock.material_restore_key) == restore_key.lower()).with_for_update().first()
            if stock is None:
                stock = BlankStock(factory_id=factory_id, blank_size_ml=blank_size_ml, variety=variety)
                db.add(stock)
                increment_bulk_stat(stats, "inserted")
            else:
                increment_bulk_stat(stats, "updated")
            transfer_restore_key(
                db, BlankStock, factory_id=factory_id, column=BlankStock.material_restore_key,
                restore_key=restore_key, target_id=stock.id,
            )
            stock.material_restore_key = restore_key
            stock.material_name = material_name
            stock.variety = variety
            stock.linked_bottom_size_mm = (
                int(row["linked_bottom_size_mm"])
                if row.get("linked_bottom_size_mm") is not None
                else None
            )
            stock.weight_per_bora_kg = row.get("weight_per_bora_kg") or Decimal("0")
            total_boras = Decimal("0")
            if "total_boras_sacks" in row and row.get("total_boras_sacks") is not None:
                total_boras = Decimal(str(row["total_boras_sacks"]))
            elif "total_boras" in row and row.get("total_boras") is not None:
                total_boras = Decimal(str(row["total_boras"]))
            elif getattr(stock, "total_boras", None) is not None:
                total_boras = Decimal(str(stock.total_boras))

            stock.total_boras = total_boras
            stock.total_qty_kg = total_boras * stock.weight_per_bora_kg
            saved_count += 1
        db.flush()
        return saved_count

    if sub_tab_type == "bottom_reel":
        saved_count = 0
        for row in valid_rows:
            bottom_size_mm = int(row["bottom_size_mm"])
            restore_key = (row.get("material_restore_key") or "").strip() or None
            variety = row["variety_design"].strip()
            query = db.query(BottomStock).filter(BottomStock.factory_id == factory_id)
            stock = query.filter(
                BottomStock.bottom_size_mm == bottom_size_mm,
                sql_func.lower(BottomStock.variety) == variety.lower(),
            ).with_for_update().first()
            if stock is None and restore_key:
                stock = query.filter(sql_func.lower(BottomStock.material_restore_key) == restore_key.lower()).with_for_update().first()
            if stock is None:
                stock = BottomStock(factory_id=factory_id, bottom_size_mm=bottom_size_mm, variety=variety)
                db.add(stock)
                increment_bulk_stat(stats, "inserted")
            else:
                increment_bulk_stat(stats, "updated")
            transfer_restore_key(
                db, BottomStock, factory_id=factory_id, column=BottomStock.material_restore_key,
                restore_key=restore_key, target_id=stock.id,
            )
            stock.total_rolls = row["total_individual_rolls"]
            stock.material_restore_key = restore_key
            stock.variety = variety
            stock.total_weight_kg = row["total_weight_kg"]
            stock.total_qty_kg = row["total_weight_kg"]
            stock.price_per_kg = row["bottom_price_per_kg"]
            saved_count += 1
        db.flush()
        return saved_count

    if sub_tab_type == "box_stock":
        saved_count = 0
        for row in valid_rows:
            packaging_size_name = row["box_type"].strip()
            stock = (
                db.query(BoxStock)
                .filter(
                    BoxStock.factory_id == factory_id,
                    sql_func.lower(BoxStock.packaging_size_name) == packaging_size_name.lower(),
                )
                .with_for_update()
                .first()
            )
            if stock is None:
                stock = BoxStock(factory_id=factory_id, packaging_size_name=packaging_size_name)
                db.add(stock)
                increment_bulk_stat(stats, "inserted")
            else:
                increment_bulk_stat(stats, "updated")
            stock.box_type = packaging_size_name
            stock.quantity = row["box_quantity_pieces"]
            stock.total_boxes = row["box_quantity_pieces"]
            stock.price_per_box = row["price_per_box_rs"]
            stock.size_for_finished_product = serialize_finished_product_sizes(
                row["size_for_finished_product"]
            )
            saved_count += 1
        db.flush()
        return saved_count

    if sub_tab_type == "plastic_stock":
        saved_count = 0
        for row in valid_rows:
            plastic_size_name = row["plastic_size_type"].strip()
            cup_size_ml = int(row["used_for_cup_size_ml"])
            stock = (
                db.query(PlasticStock)
                .filter(
                    PlasticStock.factory_id == factory_id,
                    sql_func.lower(PlasticStock.plastic_size_name) == plastic_size_name.lower(),
                    PlasticStock.cup_size_ml == cup_size_ml,
                )
                .with_for_update()
                .first()
            )
            if stock is None:
                stock = PlasticStock(factory_id=factory_id, plastic_size_name=plastic_size_name, cup_size_ml=cup_size_ml)
                db.add(stock)
                increment_bulk_stat(stats, "inserted")
            else:
                increment_bulk_stat(stats, "updated")
            stock.total_boras = row["total_boras_sacks"]
            stock.weight_per_bora_kg = row["weight_per_bora_kg"]
            stock.price_per_kg = row["price_per_kg_rs"]
            saved_count += 1
        db.flush()
        return saved_count

    if sub_tab_type == "finished_goods":
        saved_count = 0
        if fg_debug_info is not None:
            fg_debug_info["rows_read"] = len(valid_rows)
        for row in valid_rows:
            product_size_ml = int(row["product_size_ml"])
            restore_key = (row.get("product_restore_key") or "").strip() or None
            variety = (row.get("variety_design") or "Standard/White").strip() or "Standard/White"
            packaging_size_name = (row.get("packaging_size_name") or "").strip()
            carton_type = (row.get("carton_type") or "").strip()
            if not packaging_size_name:
                packaging_size_name = f"{product_size_ml}ML - {variety}"
            pieces_per_packet = max(int(row["pcs_per_packet"]), 1)
            packets_per_box = max(int(row["packets_per_box"]), 1)
            initial_stock_boxes = max(int(row["initial_stock_boxes"]), 0)
            initial_loose_packets = max(int(row.get("initial_loose_packets") or 0), 0)

            box_inventory = get_or_create_inventory(db, factory_id, packaging_size_name, "Packaging", "pieces")
            poly_inventory = get_or_create_inventory(db, factory_id, f"{product_size_ml}ml Polybag", "Packaging", "pieces")
            profile = (
                db.query(PackagingProfile)
                .filter(PackagingProfile.factory_id == factory_id)
                .filter(PackagingProfile.cup_size_ml == product_size_ml)
                .filter(sql_func.lower(PackagingProfile.profile_name) == packaging_size_name.lower())
                .filter(sql_func.lower(sql_func.coalesce(PackagingProfile.print_design_name, "")) == variety.lower())
                .with_for_update()
                .first()
            )
            if profile is None:
                profile = PackagingProfile(
                    factory_id=factory_id,
                    profile_name=packaging_size_name,
                    product_name=f"{product_size_ml}ml Paper Cup",
                    product_name_ml=product_size_ml,
                    cup_size_ml=product_size_ml,
                    print_design_name=variety,
                    polybag_capacity=pieces_per_packet,
                    box_capacity=pieces_per_packet * packets_per_box,
                    box_size_name=carton_type,
                    cups_per_poly=pieces_per_packet,
                    cups_per_polybag=pieces_per_packet,
                    polys_per_box=packets_per_box,
                    polybags_per_box=packets_per_box,
                    box_inventory_id=box_inventory.id,
                    poly_inventory_id=poly_inventory.id,
                )
                db.add(profile)
                db.flush()
            else:
                profile.print_design_name = variety
                profile.cup_size_ml = product_size_ml
                profile.product_name_ml = product_size_ml
                profile.polybag_capacity = pieces_per_packet
                profile.box_capacity = pieces_per_packet * packets_per_box
                profile.cups_per_poly = pieces_per_packet
                profile.cups_per_polybag = pieces_per_packet
                profile.polys_per_box = packets_per_box
                profile.polybags_per_box = packets_per_box
                profile.box_size_name = carton_type
                profile.box_inventory_id = box_inventory.id
                profile.poly_inventory_id = poly_inventory.id
                db.flush()

            # Update FinishedGoodsStock
            stock = (
                db.query(FinishedGoodsStock)
                .filter(FinishedGoodsStock.factory_id == factory_id)
                .filter(FinishedGoodsStock.packaging_profile_id == profile.id)
                .with_for_update()
                .first()
            )
            is_new_stock = stock is None
            if is_new_stock:
                stock = FinishedGoodsStock(
                    factory_id=factory_id,
                    cup_size_ml=product_size_ml,
                    packaging_profile_id=profile.id,
                    boxes_available=initial_stock_boxes,
                    category="CUP_FINISHED",
                    variant_name=variety,
                )
                db.add(stock)
                increment_bulk_stat(stats, "inserted")
            else:
                stock.cup_size_ml = product_size_ml
                stock.boxes_available = initial_stock_boxes
                stock.category = "CUP_FINISHED"
                stock.variant_name = variety
                increment_bulk_stat(stats, "updated")

            db.flush()
            if fg_debug_info is not None:
                if is_new_stock:
                    fg_debug_info["rows_inserted"] += 1
                else:
                    fg_debug_info["rows_updated"] += 1
                fg_debug_info["created_finished_goods_stock_ids"].append(stock.id)

            final_stock = sync_finished_goods_to_final_product_stock(db, factory_id, stock, fg_debug_info)
            transfer_restore_key(
                db, FinalProductStock, factory_id=factory_id, column=FinalProductStock.product_restore_key,
                restore_key=restore_key, target_id=final_stock.id,
            )
            final_stock.product_restore_key = restore_key
            final_stock.carton_type = carton_type
            final_stock.loose_packets = initial_loose_packets
            final_stock.current_quantity = initial_stock_boxes

            saved_count += 1
        return saved_count

    return 0


def reset_active_onboarding_master_data(
    db: Session,
    *,
    factory_id: int,
) -> int:
    archived_count = (
        db.query(Worker)
        .filter(Worker.factory_id == factory_id, Worker.is_active.is_(True))
        .update({Worker.is_active: False}, synchronize_session=False)
    )
    archived_count += (
        db.query(Machine)
        .filter(Machine.factory_id == factory_id, Machine.is_active.is_(True))
        .update({Machine.is_active: False}, synchronize_session=False)
    )
    archived_count += (
        db.query(Customer)
        .filter(Customer.factory_id == factory_id, Customer.is_active.is_(True))
        .update({Customer.is_active: False}, synchronize_session=False)
    )

    removed_count = 0
    for model in (BlankStock, BottomStock, BoxStock, PlasticStock, FinalProductStock):
        removed_count += (
            db.query(model)
            .filter(model.factory_id == factory_id)
            .delete(synchronize_session=False)
        )
    db.flush()
    return archived_count + removed_count


def _log_onboarding_change(db: Session, factory_id: int, action: str, subject: str) -> None:
    try:
        log_factory_operation(
            db,
            factory_id=int(factory_id),
            event_type="machine_telemetry",
            description=f"👥 Onboarding Change: {action} {subject} in system configs",
        )
    except Exception as log_error:
        logger.exception("Suppressed activity log failure for onboarding change: %s", log_error)


@v1_router.get("/template/master")
def download_master_onboarding_template(
    current_user: User = Depends(check_permissions(FACTORY_VIEW_ROLES)),
):
    return StreamingResponse(
        build_owner_onboarding_workbook(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{OWNER_ONBOARDING_FILENAME}"'},
    )


@v1_router.post("/bulk-upload/master/validate")
async def validate_master_onboarding(
    file: UploadFile = File(...),
    strict_validation: bool = False,
    current_user: User = Depends(check_permissions(OWNER_ROLES)),
):
    """
    Dry-run validation: parse the Excel workbook and return a detailed row-by-row
    validation report WITHOUT committing anything to the database.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Only .xlsx master onboarding files are supported")

    file_bytes = await file.read()
    effective_strict = (
        strict_validation
        or detect_master_template_version(file_bytes) >= 2
        or is_owner_friendly_workbook(file_bytes)
    )
    valid_by_type, failed_rows = read_master_bulk_excel(file_bytes, strict_validation=effective_strict)
    valid_by_type, duplicate_warnings = dedupe_valid_bulk_rows(valid_by_type)
    cross_sheet_issues = validate_bulk_cross_sheet(valid_by_type, strict_validation=effective_strict)

    # Count total ACTUAL rows across all sheets
    total_attempted = sum(len(rows) for rows in valid_by_type.values())
    successful_rows = total_attempted  # in dry-run all valid rows are "would succeed"
@v1_router.post("/bulk-upload/master")
async def bulk_upload_master_onboarding(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    strict_validation: bool = False,
    current_user: User = Depends(check_permissions(OWNER_ROLES)),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Only .xlsx master onboarding files are supported")

    file_bytes = await file.read()
    effective_strict = (
        strict_validation
        or detect_master_template_version(file_bytes) >= 2
        or is_owner_friendly_workbook(file_bytes)
    )
    fg_debug_info = inspect_finished_goods_sheet(file_bytes)
    valid_by_type, failed_rows = read_master_bulk_excel(file_bytes, strict_validation=effective_strict)

    valid_by_type, duplicate_warnings = dedupe_valid_bulk_rows(valid_by_type)
    cross_sheet_issues = validate_bulk_cross_sheet(valid_by_type, strict_validation=effective_strict)

    # Build enriched validation report
    total_attempted = sum(len(rows) for rows in valid_by_type.values()) + len(failed_rows)
    issues = enrich_failed_rows(failed_rows) + duplicate_warnings + cross_sheet_issues

    # Extract any failures or skip reasons for Finished Goods
    fg_errors = [err for err in failed_rows if err.get("sheet") == "Finished Goods"]
    if fg_errors:
        for err in fg_errors:
            reason = err.get("error") or "unknown validation error"
            missing = err.get("missing_headers")
            if missing:
                reason = f"{reason}; missing headers: {', '.join(missing)}"
            fg_debug_info["skip_reasons"].append(
                f"Finished Goods sheet not imported because {reason}"
            )
            fg_debug_info["rows_skipped"] += 1
    elif fg_debug_info["sheet_name"] and not valid_by_type.get("finished_goods"):
        fg_debug_info["skip_reasons"].append("Finished Goods sheet not imported because no ACTUAL rows were found or sheet is empty")

    # Fatal errors block the entire import
    fatal_issues = [i for i in issues if i.severity == ValidationSeverity.FATAL]
    if fatal_issues:
        report = make_report(issues, successful_rows=0, total_rows_attempted=total_attempted)
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Import rejected due to validation errors. Fix the highlighted rows and upload again.",
                "overall_status": "failed",
                "validation_report": report.to_dict(),
                "failed_rows": failed_rows,
                "fg_debug_info": fg_debug_info,
            },
        )

    if not any(valid_by_type.values()):
        raise HTTPException(
            status_code=422,
            detail={
                "message": "No importable rows found. Mark data rows with row_type = ACTUAL.",
                "overall_status": "failed",
                "validation_report": make_report([], 0, 0).to_dict(),
                "failed_rows": [{"sheet": "Workbook", "row": None, "error": "No valid rows found"}],
                "fg_debug_info": fg_debug_info,
            },
        )

    inserted_counts: dict[str, int] = {}
    operation_counts: dict[str, int] = {
        "inserted": 0,
        "updated": 0,
        "unchanged": 0,
        "skipped": 0,
        "failed": 0,
        "warnings": len([issue for issue in issues if issue.severity == ValidationSeverity.WARNING]),
    }
    sheet_stats = {}
    try:
        operation_counts["skipped"] = reset_active_onboarding_master_data(
            db,
            factory_id=int(current_user.factory_id),
        )
        for sub_tab_type in BULK_TEMPLATE_COLUMNS:
            sub_stats = {
                "inserted": 0,
                "updated": 0,
                "unchanged": 0,
                "skipped": 0,
                "failed": 0,
            }
            inserted_counts[sub_tab_type] = apply_bulk_rows(
                db,
                current_user,
                sub_tab_type,
                valid_by_type.get(sub_tab_type, []),
                sub_stats,
                fg_debug_info=fg_debug_info if sub_tab_type == "finished_goods" else None
            )
            for k in ["inserted", "updated", "unchanged", "skipped", "failed"]:
                operation_counts[k] += sub_stats[k]
            sheet_stats[sub_tab_type] = sub_stats

        operation_counts["warnings"] = len([
            issue for issue in issues if issue.severity == ValidationSeverity.WARNING
        ])
        db.commit()
        total_rows = sum(inserted_counts.values())
        log_bulk_upload(background_tasks, db, current_user, "master_onboarding", total_rows)
        log_bulk_inventory_uploads(background_tasks, db, current_user, inserted_counts)

        report = make_report(issues, successful_rows=total_rows, total_rows_attempted=total_attempted)
        overall_status = "partial" if report.warning_issues else "success"

        # Calculate dynamic summary status counts for response
        summary_payload = {
            "finished_goods": {
                "read": fg_debug_info["rows_read"],
                "inserted": fg_debug_info["rows_inserted"],
                "updated": fg_debug_info["rows_updated"],
                "skipped": fg_debug_info["rows_skipped"],
            },
            "workers": {
                "read": len(valid_by_type.get("worker", [])),
                "inserted": len(valid_by_type.get("worker", [])),
                "updated": 0,
                "skipped": 0,
            },
            "customers": {
                "read": len(valid_by_type.get("customer", [])),
                "inserted": len(valid_by_type.get("customer", [])),
                "updated": 0,
                "skipped": 0,
            },
            "raw_materials": {
                "read": len(valid_by_type.get("blank_stock", [])) + len(valid_by_type.get("bottom_reel", [])),
                "inserted": len(valid_by_type.get("blank_stock", [])) + len(valid_by_type.get("bottom_reel", [])),
                "updated": 0,
                "skipped": 0,
            },
            "machines": {
                "read": len(valid_by_type.get("machine", [])),
                "inserted": len(valid_by_type.get("machine", [])),
                "updated": 0,
                "skipped": 0,
            }
        }

        blank_stock_valid_rows = valid_by_type.get("blank_stock", [])
        total_boras_imported = sum(float(row.get("total_boras_sacks") or 0) for row in blank_stock_valid_rows)
        total_kg_imported = sum(float(row.get("total_boras_sacks") or 0) * float(row.get("weight_per_bora_kg") or 0) for row in blank_stock_valid_rows)

        debug_import_summary = {
            "Cup Blank": {
                "created_count": sheet_stats.get("blank_stock", {}).get("inserted", 0),
                "updated_count": sheet_stats.get("blank_stock", {}).get("updated", 0),
                "unchanged_count": sheet_stats.get("blank_stock", {}).get("unchanged", 0),
                "total_boras_imported": total_boras_imported,
                "total_kg_imported": total_kg_imported,
            }
        }

        return {
            "message": "Master data replaced / updated successfully",
            "overall_status": overall_status,
            "rows_inserted": total_rows,
            "created_count": operation_counts["inserted"],
            "updated_count": operation_counts["updated"],
            "unchanged_count": operation_counts["unchanged"],
            "archived_skipped_count": operation_counts["skipped"],
            "inserted_counts": inserted_counts,
            "operation_counts": operation_counts,
            "validation_report": report.to_dict(),
            "summary": summary_payload,
            "fg_debug_info": fg_debug_info,
            "errors": [err for err in failed_rows],
            "debug_import_summary": debug_import_summary,
            # kept for backward compatibility
            "failed_rows": [],
        }
    except HTTPException:
        db.rollback()
        raise
    except IntegrityError as exc:
        db.rollback()
        db_issues = enrich_failed_rows([{
            "sheet": "Database",
            "row": None,
            "error": "A conflicting master-data identity could not be resolved automatically.",
        }])
        report = make_report(db_issues, successful_rows=0, total_rows_attempted=total_attempted)
        raise HTTPException(
            status_code=409,
            detail={
                "message": "A master-data conflict could not be resolved automatically. Review the workbook for ambiguous identities.",
                "overall_status": "failed",
                "validation_report": report.to_dict(),
                "failed_rows": [{
                    "sheet": "Database",
                    "row": None,
                    "error": "A conflicting master-data identity could not be resolved automatically.",
                }],
            },
        ) from exc
    except Exception as exc:
        db.rollback()
        db_issues = enrich_failed_rows([{"sheet": "Database", "row": None, "error": str(exc)}])
        report = make_report(db_issues, successful_rows=0, total_rows_attempted=total_attempted)
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Unexpected server error during import.",
                "overall_status": "failed",
                "validation_report": report.to_dict(),
                "failed_rows": [{"sheet": "Database", "row": None, "error": str(exc)}],
            },
        ) from exc


class OnboardingWorkerSummary(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    name: str
    daily_wages: Decimal
    duty_hours: float
    previous_attendance: int = 0


class OnboardingMachineSummary(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    machine_type: str
    machine_number: Optional[str] = None
    mould_size_ml: Optional[int] = None
    bottom_size_mm: Optional[int] = None
    speed_per_minute: int
    machine_name: Optional[str] = None
    default_speed: float = 0
    target_output_per_shift: int = 0
    raw_materials_mapped: List[str] = Field(default_factory=list)
    is_active: bool = True


class OnboardingRawMetricSummary(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    material_type: str
    size_ml_or_mm: int
    weight_per_sack_kg: Decimal
    pieces_per_sack: int


class OnboardingPackagingMetricSummary(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    cup_size_ml: int
    kg_per_box: Decimal
    cups_per_box: int
    variety: Optional[str] = "Standard/White"


class OnboardingOverviewResponse(BaseModel):
    workers: List[OnboardingWorkerSummary]
    machines: List[OnboardingMachineSummary]
    raw_material_metrics: List[OnboardingRawMetricSummary]
    packaging_metrics: List[OnboardingPackagingMetricSummary]


class MachineLimitResponse(BaseModel):
    used: int
    limit: int
    plan: str
    nearing_limit: bool
    limit_reached: bool


class OnboardingCustomerSummary(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    name: str
    phone: Optional[str] = None
    total_due: Decimal


def _worker_summary(worker: Worker, previous_attendance: int = 0) -> dict:
    return {
        "id": worker.id,
        "name": worker.name or "Unnamed Worker",
        "daily_wages": Decimal(worker.daily_wages or worker.daily_salary or worker.salary or 0),
        "duty_hours": float(worker.duty_hours or worker.shift_hours or 8.0),
        "previous_attendance": int(previous_attendance or 0),
    }


def _machine_summary(machine: Machine) -> dict:
    machine_number = machine.machine_number or machine.machine_sequence_number or machine.name or f"M-{machine.id}"
    return {
        "id": machine.id,
        "machine_type": machine.machine_type or machine.machine_name or machine.name or "Custom Machine",
        "machine_number": machine_number,
        "mould_size_ml": machine.mould_size_ml or machine.cup_size_ml or machine.current_mould_size,
        "bottom_size_mm": machine.bottom_size_mm or machine.bottom_size or machine.current_bottom_size,
        "speed_per_minute": int(machine.speed_per_minute or machine.speed_cups_per_minute or machine.speed_bpm or 0),
        "machine_name": machine.machine_name or machine.machine_type or machine.name,
        "default_speed": float(machine.default_speed or machine.speed_per_minute or machine.speed_cups_per_minute or machine.speed_bpm or 0),
        "target_output_per_shift": int(machine.target_output_per_shift or 0),
        "raw_materials_mapped": machine.raw_materials_mapped or [],
        "is_active": bool(machine.is_active),
    }


def _raw_metric_summary(metric: RawMaterialMetrics) -> dict:
    return {
        "id": metric.id,
        "material_type": metric.material_type or "Blank",
        "size_ml_or_mm": int(metric.size_ml_or_mm or 1),
        "weight_per_sack_kg": Decimal(metric.weight_per_sack_kg or 0),
        "pieces_per_sack": int(metric.pieces_per_sack or 1),
    }


def _packaging_metric_summary(metric: PackagingMetrics) -> dict:
    return {
        "id": metric.id,
        "cup_size_ml": int(metric.cup_size_ml or 1),
        "kg_per_box": Decimal(metric.kg_per_box or 0),
        "cups_per_box": int(metric.cups_per_box or 1),
        "variety": getattr(metric, "variant_name", None) or "Standard/White",
    }


class FinalProductOpeningStockRequest(BaseModel):
    product_id: Optional[int] = Field(default=None, gt=0)
    product_size_ml: Optional[int] = Field(default=None, gt=0)
    variety: str = Field(default="Standard/White", max_length=100)
    packaging_size: Optional[str] = Field(default=None, max_length=100)
    packaging_size_name: Optional[str] = Field(default=None, max_length=100)
    initial_quantity: int = Field(default=0, ge=0)
    current_quantity: Optional[int] = Field(default=None, ge=0)
    total_boxes: Optional[int] = Field(default=None, ge=0)
    loose_packets: int = Field(default=0, ge=0)
    pieces_per_packet: int = Field(default=1, gt=0)
    packets_per_box: Optional[int] = Field(default=None, gt=0)
    packets_per_box_limit: Optional[int] = Field(default=None, gt=0)
    category: Optional[str] = None


class FinalProductOpeningStockResponse(BaseModel):
    id: int
    factory_id: str
    product_size_ml: Optional[int] = None
    variety: Optional[str] = None
    packaging_size: Optional[str] = None
    packaging_size_name: Optional[str] = None
    pieces_per_packet: Optional[int] = None
    packets_per_box: Optional[int] = None
    current_quantity: Optional[int] = None
    total_boxes: Optional[int] = None
    loose_packets: Optional[int] = None
    packets_per_box_limit: Optional[int] = None


class BottomStockCreate(BaseModel):
    bottom_size_mm: int = Field(..., gt=0)
    bag_weight_kg: Optional[Decimal] = Field(default=None, ge=0)
    rolls_per_bag: Optional[int] = Field(default=None, ge=0)
    total_bags: Optional[int] = Field(default=None, ge=0)
    total_rolls: Optional[int] = Field(default=None, ge=0)
    total_weight_kg: Optional[Decimal] = Field(default=None, ge=0)


class BlankStockCreate(BaseModel):
    material_name: str = Field(default="Blank", max_length=100)
    size_ml: int = Field(..., gt=0)
    kg_per_sack: Optional[Decimal] = Field(default=None, gt=0)
    total_sacks: Decimal = Field(..., ge=0)


@router.get("/factory-profile", response_model=FactoryProfileResponse)
def get_factory_profile(
    current_user: User = Depends(check_permissions(FACTORY_VIEW_ROLES)),
    db: Session = Depends(get_db)
):
    if not current_user.factory_id:
        raise HTTPException(status_code=404, detail="No factory linked to this user")
    factory = db.query(Factory).filter(Factory.id == current_user.factory_id).first()
    if not factory:
        raise HTTPException(status_code=404, detail="Factory not found")
    settings = (
        db.query(FactorySettings)
        .filter(FactorySettings.factory_id == int(factory.id))
        .first()
    )
    return FactoryProfileResponse(
        id=factory.id,
        factory_name=factory.factory_name or factory.name,
        address=factory.address,
        gst_number=getattr(factory, "gst_number", None),
        invoice_prefix=getattr(factory, "invoice_prefix", "INV-") or "INV-",
        advance_payment_discount_percentage=getattr(factory, "advance_payment_discount_percentage", Decimal("2.00")) or Decimal("2.00"),
        digital_signature_url=getattr(factory, "digital_signature_url", None),
        bill_of_supply_start_seq=(settings.bill_of_supply_start_seq if settings else getattr(factory, "next_bill_of_supply_number", 1)) or 1,
        tax_invoice_start_seq=(settings.tax_invoice_start_seq if settings else getattr(factory, "next_tax_invoice_number", 1)) or 1,
        bill_of_supply_simple_start_seq=(settings.bill_of_supply_simple_start_seq if settings else getattr(factory, "next_bill_of_supply_simple_number", 1)) or 1,
        next_tax_invoice_number=getattr(factory, "next_tax_invoice_number", 1) or 1,
        next_bill_of_supply_number=getattr(factory, "next_bill_of_supply_number", 1) or 1,
        next_bill_of_supply_simple_number=getattr(factory, "next_bill_of_supply_simple_number", 1) or 1,
    )

@router.post("/factory-profile", response_model=FactoryProfileResponse)
def update_factory_profile(
    payload: FactoryProfileUpdate,
    current_user: User = Depends(check_permissions(FACTORY_VIEW_ROLES)),
    db: Session = Depends(get_db)
):
    if not current_user.factory_id:
        raise HTTPException(status_code=404, detail="No factory linked to this user")
    factory = db.query(Factory).filter(Factory.id == current_user.factory_id).with_for_update().first()
    if not factory:
        raise HTTPException(status_code=404, detail="Factory not found")
    settings = get_or_create_factory_settings(db, factory.id)
    
    factory.factory_name = payload.factory_name.strip()
    factory.address = payload.address.strip() if payload.address else None
    factory.gst_number = payload.gst_number.strip() if payload.gst_number else None

    if payload.bill_of_supply_start_seq is not None:
        settings.bill_of_supply_start_seq = int(payload.bill_of_supply_start_seq)
        factory.next_bill_of_supply_number = int(payload.bill_of_supply_start_seq)
    if payload.tax_invoice_start_seq is not None:
        settings.tax_invoice_start_seq = int(payload.tax_invoice_start_seq)
        factory.next_tax_invoice_number = int(payload.tax_invoice_start_seq)
    if payload.bill_of_supply_simple_start_seq is not None:
        settings.bill_of_supply_simple_start_seq = int(payload.bill_of_supply_simple_start_seq)
        factory.next_bill_of_supply_simple_number = int(payload.bill_of_supply_simple_start_seq)

    if payload.invoice_prefix is not None:
        factory.invoice_prefix = payload.invoice_prefix.strip()

    if payload.advance_payment_discount_percentage is not None:
        factory.advance_payment_discount_percentage = payload.advance_payment_discount_percentage

    if payload.digital_signature_url is not None:
        if current_user.role != "Owner":
            raise HTTPException(status_code=403, detail="Only Owner can update invoice signature")
        factory.digital_signature_url = payload.digital_signature_url.strip() if payload.digital_signature_url else None

    db.commit()
    db.refresh(factory)
    return FactoryProfileResponse(
        id=factory.id,
        factory_name=factory.factory_name or factory.name,
        address=factory.address,
        gst_number=factory.gst_number,
        invoice_prefix=factory.invoice_prefix or "INV-",
        advance_payment_discount_percentage=factory.advance_payment_discount_percentage,
        digital_signature_url=factory.digital_signature_url,
        bill_of_supply_start_seq=settings.bill_of_supply_start_seq or 1,
        tax_invoice_start_seq=settings.tax_invoice_start_seq or 1,
        bill_of_supply_simple_start_seq=settings.bill_of_supply_simple_start_seq or 1,
        next_tax_invoice_number=getattr(factory, "next_tax_invoice_number", 1) or 1,
        next_bill_of_supply_number=getattr(factory, "next_bill_of_supply_number", 1) or 1,
        next_bill_of_supply_simple_number=getattr(factory, "next_bill_of_supply_simple_number", 1) or 1,
    )


@router.post("/factory-profile/signature")
async def upload_invoice_signature(
    file: UploadFile = File(...),
    current_user: User = Depends(check_permissions(OWNER_ROLES)),
    db: Session = Depends(get_db),
):
    extension = Path(file.filename or "").suffix.lower()
    if extension not in {".png", ".jpg", ".jpeg"}:
        raise HTTPException(status_code=422, detail="Signature must be PNG, JPG, or JPEG")
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=422, detail="Signature image must be 2 MB or smaller")
    signature_dir = Path("volumes/media/signatures") / str(current_user.factory_id)
    signature_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid4().hex}{extension}"
    target = signature_dir / filename
    target.write_bytes(content)
    factory = db.query(Factory).filter(Factory.id == current_user.factory_id).with_for_update().first()
    if factory is None:
        target.unlink(missing_ok=True)
        raise HTTPException(status_code=404, detail="Factory not found")
    factory.digital_signature_url = f"/media/signatures/{current_user.factory_id}/{filename}"
    db.commit()
    return {"digital_signature_url": factory.digital_signature_url}


@router.delete("/factory-profile/signature")
def remove_invoice_signature(
    current_user: User = Depends(check_permissions(OWNER_ROLES)),
    db: Session = Depends(get_db),
):
    factory = db.query(Factory).filter(Factory.id == current_user.factory_id).with_for_update().first()
    if factory is None:
        raise HTTPException(status_code=404, detail="Factory not found")
    old_url = factory.digital_signature_url
    factory.digital_signature_url = None
    db.commit()
    if old_url and old_url.startswith(f"/media/signatures/{current_user.factory_id}/"):
        (Path("volumes/media") / old_url.removeprefix("/media/")).unlink(missing_ok=True)
    return {"digital_signature_url": None}


@router.get("/overview", response_model=OnboardingOverviewResponse)
def onboarding_overview(
    current_user: User = Depends(check_permissions(FACTORY_VIEW_ROLES)),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)
    workers = (
        db.query(Worker)
        .filter(Worker.factory_id == factory_id)
        .filter(Worker.is_active.is_(True))
        .order_by(Worker.name.asc().nullslast(), Worker.id.asc())
        .all()
    )
    opening_by_worker_id = {
        row.worker_id: int(row.present_days or 0)
        for row in db.query(WorkerOpeningAttendance)
        .filter(WorkerOpeningAttendance.factory_id == factory_id)
        .filter(WorkerOpeningAttendance.worker_id.in_([worker.id for worker in workers] or [0]))
        .all()
    }
    machines = (
        db.query(Machine)
        .filter(Machine.factory_id == factory_id)
        .filter(Machine.is_active.is_(True))
        .order_by(Machine.machine_number.asc().nullslast(), Machine.name.asc().nullslast(), Machine.id.asc())
        .all()
    )
    raw_metrics = (
        db.query(RawMaterialMetrics)
        .filter(RawMaterialMetrics.factory_id == factory_id)
        .order_by(RawMaterialMetrics.material_type.asc().nullslast(), RawMaterialMetrics.size_ml_or_mm.asc().nullslast(), RawMaterialMetrics.id.asc())
        .all()
    )
    packaging_metrics = (
        db.query(PackagingMetrics)
        .filter(PackagingMetrics.factory_id == factory_id)
        .order_by(PackagingMetrics.cup_size_ml.asc().nullslast(), PackagingMetrics.id.asc())
        .all()
    )
    return OnboardingOverviewResponse(
        workers=[_worker_summary(worker, opening_by_worker_id.get(worker.id, 0)) for worker in workers],
        machines=[_machine_summary(machine) for machine in machines],
        raw_material_metrics=[_raw_metric_summary(metric) for metric in raw_metrics],
        packaging_metrics=[_packaging_metric_summary(metric) for metric in packaging_metrics],
    )


@router.post("/final-stock", response_model=FinalProductOpeningStockResponse)
def save_final_product_opening_stock(
    payload: FinalProductOpeningStockRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)
    stock = None
    provided_fields = getattr(payload, "model_fields_set", set())

    try:
        product_size_ml = payload.product_size_ml
        packaging_size_name = (payload.packaging_size_name or payload.packaging_size or "").strip()
        variety = (payload.variety or "Standard/White").strip() or "Standard/White"
        pieces_per_packet = int(payload.pieces_per_packet or 1)
        packets_per_box_limit = int(payload.packets_per_box_limit or payload.packets_per_box or 1)
        quantity = payload.current_quantity
        if quantity is None:
            quantity = payload.total_boxes
        if quantity is None:
            quantity = payload.initial_quantity
        quantity = int(quantity or 0)
        loose_packets = int(payload.loose_packets or 0)

        if payload.product_id:
            stock = (
                db.query(FinalProductStock)
                .filter(FinalProductStock.factory_id == factory_id)
                .filter(FinalProductStock.id == payload.product_id)
                .with_for_update()
                .first()
            )
            if stock is None:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Final product stock item not found")
            product_size_ml = product_size_ml if "product_size_ml" in provided_fields else stock.product_size_ml
            packaging_size_name = packaging_size_name if ("packaging_size_name" in provided_fields or "packaging_size" in provided_fields) else (stock.packaging_size_name or f"{product_size_ml or 210}ml Standard Box")
            variety = variety.strip() if "variety" in provided_fields else (stock.variety or "Standard/White")
            pieces_per_packet = int(pieces_per_packet if "pieces_per_packet" in provided_fields else (stock.pieces_per_packet or 1))
            packets_per_box_limit = int(
                packets_per_box_limit
                if ("packets_per_box_limit" in provided_fields or "packets_per_box" in provided_fields)
                else (stock.packets_per_box_limit or 1)
            )
        else:
            product_size_ml = int(product_size_ml or 210)
            packaging_size_name = packaging_size_name or f"{product_size_ml}ml Standard Box"
            stock = (
                db.query(FinalProductStock)
                .filter(FinalProductStock.factory_id == factory_id)
                .filter(FinalProductStock.product_size_ml == product_size_ml)
                .filter(sql_func.lower(sql_func.trim(FinalProductStock.variety)) == variety.strip().lower())
                .filter(sql_func.lower(sql_func.trim(FinalProductStock.packaging_size_name)) == packaging_size_name.strip().lower())
                .with_for_update()
                .first()
            )
            if stock is None:
                stock = FinalProductStock(
                    factory_id=factory_id,
                    product_size_ml=product_size_ml,
                    variety=variety,
                    packaging_size_name=packaging_size_name,
                    pieces_per_packet=max(pieces_per_packet, 1),
                    packets_per_box_limit=max(packets_per_box_limit, 1),
                    current_quantity=0,
                    total_boxes=0,
                    loose_packets=0,
                )
                db.add(stock)

        if not product_size_ml:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="product_size_ml is required")
        if not packaging_size_name:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="packaging_size_name is required")

        stock.product_size_ml = int(product_size_ml)
        stock.variety = variety
        stock.packaging_size_name = packaging_size_name
        stock.pieces_per_packet = max(int(pieces_per_packet or 1), 1)
        stock.packets_per_box_limit = max(int(packets_per_box_limit or 1), 1)
        stock.current_quantity = max(quantity, 0)
        stock.total_boxes = max(quantity, 0)
        stock.loose_packets = max(loose_packets, 0)
        db.flush()
        try:
            db.commit()
        except Exception as exc:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Variant combination already initialized: {exc}"
            )
        db.refresh(stock)
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception("Final product opening stock core save failed")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Final product stock save failed: {exc}") from exc

    # Optional dependency/bootstrap work must never break the saved opening stock.
    try:
        poly_inventory = get_or_create_inventory(db, factory_id, f"{stock.product_size_ml}ml Polybag", "Packaging", "pieces")
        box_inventory = get_or_create_inventory(db, factory_id, stock.packaging_size_name, "Packaging", "pieces")
        profile = (
            db.query(PackagingProfile)
            .filter(PackagingProfile.factory_id == factory_id)
            .filter(sql_func.lower(PackagingProfile.profile_name) == stock.packaging_size_name.lower())
            .first()
        )
        if profile is None:
            profile = PackagingProfile(
                factory_id=factory_id,
                profile_name=stock.packaging_size_name,
                product_name=f"{stock.product_size_ml}ml Paper Cup",
                product_name_ml=stock.product_size_ml,
                cup_size_ml=stock.product_size_ml,
                polybag_capacity=stock.pieces_per_packet,
                box_capacity=stock.pieces_per_packet * stock.packets_per_box_limit,
                box_size_name=stock.packaging_size_name,
                cups_per_poly=stock.pieces_per_packet,
                cups_per_polybag=stock.pieces_per_packet,
                polys_per_box=stock.packets_per_box_limit,
                polybags_per_box=stock.packets_per_box_limit,
                box_inventory_id=box_inventory.id,
                poly_inventory_id=poly_inventory.id,
            )
            db.add(profile)
            db.flush()
        finished = (
            db.query(FinishedGoodsStock)
            .filter(FinishedGoodsStock.factory_id == factory_id)
            .filter(FinishedGoodsStock.packaging_profile_id == profile.id)
            .first()
        )
        if finished is None:
            finished = FinishedGoodsStock(
                factory_id=factory_id,
                cup_size_ml=stock.product_size_ml,
                packaging_profile_id=profile.id,
                boxes_available=0,
            )
            db.add(finished)
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.exception("Optional final-stock packaging bootstrap failed and was suppressed: %s", exc)

    try:
        metric = (
            db.query(PackagingMetrics)
            .filter(PackagingMetrics.factory_id == factory_id)
            .filter(PackagingMetrics.cup_size_ml == stock.product_size_ml)
            .filter(sql_func.lower(PackagingMetrics.variant_name) == stock.variety.lower())
            .first()
        )
        if metric is None:
            metric = PackagingMetrics(
                factory_id=factory_id,
                cup_size_ml=stock.product_size_ml,
                variant_name=stock.variety,
                kg_per_box=Decimal("10.000"),
                cups_per_box=stock.packets_per_box_limit,
            )
            db.add(metric)
        else:
            metric.cups_per_box = stock.packets_per_box_limit
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.exception("Optional packaging metrics sync failed and was suppressed: %s", exc)

    try:
        from routers.inventory import recalculate_and_sync_sku_stock
        recalculate_and_sync_sku_stock(
            db=db,
            factory_id=factory_id,
            product_size_ml=stock.product_size_ml,
            variety=stock.variety,
            packaging_size_name=stock.packaging_size_name,
        )
        db.commit()
        db.refresh(stock)
    except Exception as exc:
        db.rollback()
        logger.exception("Optional final-stock live cache sync failed and was suppressed: %s", exc)
        stock = db.query(FinalProductStock).filter(FinalProductStock.id == stock.id).first()

    try:
        background_tasks.add_task(
            sync_data_to_n8n_bg,
            factory_id=factory_id,
            sync_type="onboarding",
            action="insert",
            data=payload,
        )
    except Exception as exc:
        logger.exception("Optional final-stock n8n enqueue failed and was suppressed: %s", exc)

    return FinalProductOpeningStockResponse(
        id=stock.id,
        factory_id=str(stock.factory_id),
        product_size_ml=stock.product_size_ml,
        variety=stock.variety or "Standard/White",
        packaging_size=stock.packaging_size_name,
        packaging_size_name=stock.packaging_size_name,
        pieces_per_packet=stock.pieces_per_packet,
        packets_per_box=stock.packets_per_box_limit,
        current_quantity=stock.current_quantity if stock.current_quantity is not None else 0,
        total_boxes=stock.total_boxes or 0,
        loose_packets=stock.loose_packets or 0,
        packets_per_box_limit=stock.packets_per_box_limit,
    )

@router.get("/workers", response_model=List[OnboardingWorkerSummary])
def list_onboarding_workers(
    current_user: User = Depends(check_permissions(FACTORY_VIEW_ROLES)),
    db: Session = Depends(get_db),
):
    workers = (
        db.query(Worker)
        .filter(Worker.factory_id == str(current_user.factory_id))
        .filter(Worker.is_active.is_(True))
        .order_by(Worker.name.asc())
        .all()
    )
    opening_by_worker_id = {
        row.worker_id: int(row.present_days or 0)
        for row in db.query(WorkerOpeningAttendance)
        .filter(WorkerOpeningAttendance.factory_id == str(current_user.factory_id))
        .filter(WorkerOpeningAttendance.worker_id.in_([worker.id for worker in workers] or [0]))
        .all()
    }
    return [_worker_summary(worker, opening_by_worker_id.get(worker.id, 0)) for worker in workers]


@router.get("/machines", response_model=List[OnboardingMachineSummary])
def list_onboarding_machines(
    current_user: User = Depends(check_permissions(FACTORY_VIEW_ROLES)),
    db: Session = Depends(get_db),
):
    machines = (
        db.query(Machine)
        .filter(Machine.factory_id == str(current_user.factory_id))
        .filter(Machine.is_active.is_(True))
        .order_by(Machine.machine_number.asc().nullslast(), Machine.name.asc().nullslast(), Machine.id.asc())
        .all()
    )
    return [_machine_summary(machine) for machine in {machine.id: machine for machine in machines}.values()]


@router.get("/machines/limits", response_model=MachineLimitResponse)
def get_machine_limits(
    current_user: User = Depends(check_permissions(FACTORY_VIEW_ROLES)),
    db: Session = Depends(get_db),
):
    usage = get_machine_limit_usage(db, str(current_user.factory_id))
    return MachineLimitResponse(
        used=usage.used,
        limit=usage.limit,
        plan=usage.plan,
        nearing_limit=usage.used >= max(usage.limit - 1, 0),
        limit_reached=usage.used >= usage.limit,
    )


@router.get("/materials")
def list_onboarding_materials(
    current_user: User = Depends(check_permissions(FACTORY_VIEW_ROLES)),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)
    return {
        "raw_material_metrics": (
            db.query(RawMaterialMetrics)
            .filter(RawMaterialMetrics.factory_id == factory_id)
            .order_by(RawMaterialMetrics.material_type.asc(), RawMaterialMetrics.size_ml_or_mm.asc())
            .all()
        ),
        "packaging_metrics": (
            db.query(PackagingMetrics)
            .filter(PackagingMetrics.factory_id == factory_id)
            .order_by(PackagingMetrics.cup_size_ml.asc())
            .all()
        ),
    }


@router.get("/customers", response_model=List[OnboardingCustomerSummary])
def list_onboarding_customers(
    current_user: User = Depends(check_permissions(FACTORY_VIEW_ROLES)),
    db: Session = Depends(get_db),
):
    return (
        db.query(Customer)
        .filter(Customer.factory_id == str(current_user.factory_id))
        .filter(Customer.is_active.is_(True))
        .order_by(Customer.name.asc())
        .all()
    )


@router.post("/raw-material/blank", status_code=status.HTTP_201_CREATED)
def create_blank_stock(
    payload: BlankStockCreate,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)
    total_weight_kg = (payload.kg_per_sack or Decimal("0")) * payload.total_sacks
    stock = (
        db.query(BlankStock)
        .filter(BlankStock.factory_id == factory_id)
        .filter(BlankStock.blank_size_ml == payload.size_ml)
        .first()
    )
    if stock is None:
        stock = BlankStock(
            factory_id=factory_id,
            blank_size_ml=payload.size_ml,
            linked_bottom_size_mm=payload.size_ml,
        )
        db.add(stock)

    stock.weight_per_bora_kg = payload.kg_per_sack
    stock.total_boras = payload.total_sacks
    stock.total_qty_kg = total_weight_kg
    db.commit()

    # Synchronize RawMaterialMetrics for Dashboard mapping
    metric = (
        db.query(RawMaterialMetrics)
        .filter(RawMaterialMetrics.factory_id == factory_id)
        .filter(RawMaterialMetrics.material_type == "Blank")
        .filter(RawMaterialMetrics.size_ml_or_mm == payload.size_ml)
        .first()
    )
    if metric is None:
        metric = RawMaterialMetrics(
            factory_id=factory_id,
            material_type="Blank",
            size_ml_or_mm=payload.size_ml,
            weight_per_sack_kg=payload.kg_per_sack or Decimal("20.000"),
            pieces_per_sack=1000,
        )
        db.add(metric)
    else:
        if payload.kg_per_sack is not None:
            metric.weight_per_sack_kg = payload.kg_per_sack
    db.commit()

    background_tasks.add_task(
        log_activity,
        db=db,
        factory_id=int(current_user.factory_id),
        user_id=current_user.id,
        user_name=current_user.username,
        user_role=current_user.role,
        action_type="RAW_MATERIAL_ADDED",
        action_summary=f"Raw material '{payload.material_name}' added - {total_weight_kg} kg",
        entity_type="raw_material",
        entity_id=stock.id,
        metadata=None,
    )
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=factory_id,
        sync_type="onboarding",
        action="insert",
        data=payload,
    )

    db.refresh(stock)
    return {
        "id": stock.id,
        "material_name": payload.material_name,
        "size_ml": stock.blank_size_ml,
        "kg_per_sack": payload.kg_per_sack,
        "total_sacks": payload.total_sacks,
        "total_weight_kg": stock.total_qty_kg,
    }


@router.post("/raw-material/bottom", status_code=status.HTTP_201_CREATED)
def create_bottom_stock(
    payload: BottomStockCreate,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)
    calculated_rolls = (payload.rolls_per_bag or 0) * (payload.total_bags or 0)
    calculated_weight_kg = (payload.bag_weight_kg or Decimal("0.000")) * Decimal(payload.total_bags or 0)
    total_rolls = payload.total_rolls if payload.total_rolls is not None else calculated_rolls
    total_weight_kg = payload.total_weight_kg if payload.total_weight_kg is not None else calculated_weight_kg
    stock = (
        db.query(BottomStock)
        .filter(BottomStock.factory_id == factory_id)
        .filter(BottomStock.bottom_size_mm == payload.bottom_size_mm)
        .first()
    )
    if stock is None:
        stock = BottomStock(factory_id=factory_id, bottom_size_mm=payload.bottom_size_mm)
        db.add(stock)

    stock.bag_weight_kg = payload.bag_weight_kg
    stock.rolls_per_bag = payload.rolls_per_bag
    stock.total_bags = payload.total_bags
    stock.total_rolls = total_rolls
    stock.total_weight_kg = total_weight_kg
    stock.total_qty_kg = total_weight_kg

    db.commit()

    # Synchronize RawMaterialMetrics for Dashboard mapping
    metric = (
        db.query(RawMaterialMetrics)
        .filter(RawMaterialMetrics.factory_id == factory_id)
        .filter(RawMaterialMetrics.material_type == "Bottom")
        .filter(RawMaterialMetrics.size_ml_or_mm == payload.bottom_size_mm)
        .first()
    )
    if metric is None:
        metric = RawMaterialMetrics(
            factory_id=factory_id,
            material_type="Bottom",
            size_ml_or_mm=payload.bottom_size_mm,
            weight_per_sack_kg=payload.bag_weight_kg if (payload.bag_weight_kg and payload.bag_weight_kg > 0) else Decimal("10.000"),
            pieces_per_sack=1,
        )
        db.add(metric)
    else:
        if payload.bag_weight_kg and payload.bag_weight_kg > 0:
            metric.weight_per_sack_kg = payload.bag_weight_kg
    db.commit()

    background_tasks.add_task(
        log_activity,
        db=db,
        factory_id=int(current_user.factory_id),
        user_id=current_user.id,
        user_name=current_user.username,
        user_role=current_user.role,
        action_type="RAW_MATERIAL_ADDED",
        action_summary=f"Raw material '{payload.bottom_size_mm}mm Bottom' added - {total_weight_kg} kg",
        entity_type="raw_material",
        entity_id=stock.id,
        metadata=None,
    )
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=factory_id,
        sync_type="onboarding",
        action="insert",
        data=payload,
    )

    db.refresh(stock)
    return {
        "id": stock.id,
        "bottom_size_mm": stock.bottom_size_mm,
        "bag_weight_kg": stock.bag_weight_kg,
        "rolls_per_bag": stock.rolls_per_bag,
        "total_bags": stock.total_bags,
        "total_rolls": stock.total_rolls,
        "total_weight_kg": stock.total_weight_kg,
    }


@router.post("/raw-material/box", status_code=status.HTTP_201_CREATED)
def create_box_stock(
    payload: BoxStockCreate,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)
    stock = (
        db.query(BoxStock)
        .filter(BoxStock.factory_id == factory_id)
        .filter(sql_func.lower(BoxStock.packaging_size_name) == payload.box_type.lower())
        .first()
    )
    if stock is None:
        stock = BoxStock(factory_id=factory_id, packaging_size_name=payload.box_type)
        db.add(stock)

    stock.box_type = payload.box_type
    stock.quantity = payload.quantity
    stock.total_boxes = payload.quantity
    stock.price_per_box = payload.price_per_box
    db.commit()
    background_tasks.add_task(
        log_activity,
        db=db,
        factory_id=int(current_user.factory_id),
        user_id=current_user.id,
        user_name=current_user.username,
        user_role=current_user.role,
        action_type="PACKAGING_MATERIAL_ADDED",
        action_summary=f"Packaging material '{payload.box_type}' added",
        entity_type="packaging_material",
        entity_id=stock.id,
        metadata=None,
    )
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=factory_id,
        sync_type="onboarding",
        action="insert",
        data=payload,
    )
    db.refresh(stock)
    return {
        "id": stock.id,
        "box_type": stock.box_type or stock.packaging_size_name,
        "quantity": stock.quantity,
        "price_per_box": stock.price_per_box,
    }


@router.post("/raw-material/plastic", status_code=status.HTTP_201_CREATED)
def create_plastic_stock(
    payload: PlasticStockCreate,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)
    stock = (
        db.query(PlasticStock)
        .filter(PlasticStock.factory_id == factory_id)
        .filter(sql_func.lower(PlasticStock.plastic_size_name) == payload.plastic_size_name.lower())
        .filter(PlasticStock.cup_size_ml == payload.cup_size_ml)
        .first()
    )
    if stock is None:
        stock = PlasticStock(
            factory_id=factory_id,
            plastic_size_name=payload.plastic_size_name.strip(),
            cup_size_ml=payload.cup_size_ml,
        )
        db.add(stock)

    stock.total_boras = payload.total_boras
    stock.weight_per_bora_kg = payload.weight_per_bora_kg
    stock.price_per_kg = payload.price_per_kg
    db.commit()

    # Synchronize PackagingMetrics for Dashboard mapping
    metric = (
        db.query(PackagingMetrics)
        .filter(PackagingMetrics.factory_id == factory_id)
        .filter(PackagingMetrics.cup_size_ml == payload.cup_size_ml)
        .filter(sql_func.lower(PackagingMetrics.variant_name) == "standard/white")
        .first()
    )
    if metric is None:
        metric = PackagingMetrics(
            factory_id=factory_id,
            cup_size_ml=payload.cup_size_ml,
            variant_name="Standard/White",
            kg_per_box=Decimal(str(payload.weight_per_bora_kg or 0)) if payload.weight_per_bora_kg > 0 else Decimal("10.000"),
            cups_per_box=1000,
        )
        db.add(metric)
    else:
        if payload.weight_per_bora_kg > 0:
            metric.kg_per_box = payload.weight_per_bora_kg
    db.commit()

    background_tasks.add_task(
        log_activity,
        db=db,
        factory_id=int(current_user.factory_id),
        user_id=current_user.id,
        user_name=current_user.username,
        user_role=current_user.role,
        action_type="PACKAGING_MATERIAL_ADDED",
        action_summary=f"Packaging material '{payload.plastic_size_name}' added",
        entity_type="packaging_material",
        entity_id=stock.id,
        metadata=None,
    )
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=factory_id,
        sync_type="onboarding",
        action="insert",
        data=payload,
    )

    db.refresh(stock)
    return {
        "id": stock.id,
        "plastic_size_name": stock.plastic_size_name,
        "cup_size_ml": stock.cup_size_ml,
        "total_boras": stock.total_boras,
        "weight_per_bora_kg": stock.weight_per_bora_kg,
        "total_plastic_kg": stock.total_boras * stock.weight_per_bora_kg,
        "price_per_kg": stock.price_per_kg,
    }


@router.delete("/worker/{worker_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_onboarding_worker(
    worker_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(check_permissions(OWNER_ROLES)),
    db: Session = Depends(get_db),
):
    assert_owner_delete_permission(current_user)
    can_cleanup_null_factory = (current_user.role or "").lower() in {"admin", "owner"}
    worker = (
        db.query(Worker)
        .filter(Worker.id == worker_id)
        .first()
    )
    if worker is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Worker not found")

    if str(worker.factory_id) != str(current_user.factory_id):
        if not (worker.factory_id is None and can_cleanup_null_factory):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Worker not found")

    try:
        worker.is_active = False
        _log_onboarding_change(db, int(current_user.factory_id), "Deleted", worker.name)
        db.commit()
        background_tasks.add_task(
            sync_data_to_n8n_bg,
            factory_id=str(current_user.factory_id),
            sync_type="worker",
            action="delete",
            data={"worker_id": worker_id},
        )
    except Exception as e:
        db.rollback()
        logger.warning("Worker delete failed for worker_id=%s factory_id=%s", worker_id, current_user.factory_id, exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Worker delete karte waqt error aaya.",
        ) from e
    return None


@router.delete("/machine/{machine_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_onboarding_machine(
    machine_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner_delete),
    db: Session = Depends(get_db),
):
    machine = (
        db.query(Machine)
        .filter(Machine.factory_id == str(current_user.factory_id))
        .filter(Machine.id == machine_id)
        .first()
    )
    if machine is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Machine not found")
    machine_spec = machine.machine_name or machine.name or machine.machine_number or f"Machine {machine.id}"
    db.delete(machine)
    _log_onboarding_change(db, int(current_user.factory_id), "Deleted", machine_spec)
    db.commit()
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=str(current_user.factory_id),
        sync_type="onboarding",
        action="delete",
        data={"machine_id": machine_id},
    )
    return None


@router.delete("/raw-material/{raw_material_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_onboarding_raw_material(
    raw_material_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner_delete),
    db: Session = Depends(get_db),
):
    metric = (
        db.query(RawMaterialMetrics)
        .filter(RawMaterialMetrics.factory_id == str(current_user.factory_id))
        .filter(RawMaterialMetrics.id == raw_material_id)
        .first()
    )
    if metric is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Raw material metric not found")
    db.delete(metric)
    db.commit()
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=str(current_user.factory_id),
        sync_type="onboarding",
        action="delete",
        data={"raw_material_id": raw_material_id},
    )
    return None


@router.delete("/customer/{customer_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_onboarding_customer(
    customer_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner_delete),
    db: Session = Depends(get_db),
):
    customer = (
        db.query(Customer)
        .filter(Customer.factory_id == str(current_user.factory_id))
        .filter(Customer.id == customer_id)
        .first()
    )
    if customer is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    db.delete(customer)
    db.commit()
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=str(current_user.factory_id),
        sync_type="onboarding",
        action="delete",
        data={"customer_id": customer_id},
    )
    return None


# =============================================================================
# LEVEL 1 ONBOARDING — Factory & Owner
# =============================================================================

@router.post("/step1", response_model=Step1Response)
def onboarding_step1(
    payload: Step1Request,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(check_permissions(OWNER_ROLES)),
    db: Session = Depends(get_db),
):
    if current_user.factory_id and current_user.factory_id > 0:
        existing = db.query(Factory).filter(Factory.id == current_user.factory_id).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="User is already linked to a factory",
            )

    factory = Factory(
        name=payload.factory_name.strip(),
        owner_phone_number=current_user.phone_number,
    )
    db.add(factory)
    db.commit()
    db.refresh(factory)

    current_user.factory_id = factory.id
    db.commit()
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=str(factory.id),
        sync_type="onboarding",
        action="insert",
        data=payload,
    )

    return Step1Response(
        message="Factory created successfully",
        factory_id=factory.id,
        factory_name=factory.name,
    )


@router.post("/step1/workers", response_model=WorkerResponse, status_code=status.HTTP_201_CREATED)
def onboarding_step1_create_worker(
    payload: WorkerCreate,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner),
    db: Session = Depends(get_db),
):
    worker = (
        db.query(Worker)
        .filter(Worker.factory_id == str(current_user.factory_id))
        .filter(sql_func.lower(Worker.name) == payload.name.strip().lower())
        .first()
    )
    created = worker is None
    if worker is None:
        worker = Worker(factory_id=str(current_user.factory_id), name=payload.name.strip())
        db.add(worker)
        db.flush()

    if payload.phone:
        worker.phone, _ = normalize_phone_number(payload.phone, payload.country_code)
    worker.daily_wages = payload.daily_wages
    worker.duty_hours = payload.duty_hours
    worker.daily_salary = payload.daily_wages
    worker.salary = payload.daily_wages
    worker.shift_hours = payload.duty_hours
    worker.is_active = True
    db.flush()

    if payload.opening_attendance is not None:
        existing_oa = (
            db.query(WorkerOpeningAttendance)
            .filter(WorkerOpeningAttendance.factory_id == str(current_user.factory_id))
            .filter(WorkerOpeningAttendance.worker_id == worker.id)
            .first()
        )
        if existing_oa is not None:
            existing_oa.period_start = payload.opening_attendance.period_start
            existing_oa.period_end = payload.opening_attendance.period_end
            existing_oa.present_days = payload.opening_attendance.present_days
            existing_oa.half_days = payload.opening_attendance.half_days
            existing_oa.absent_days = payload.opening_attendance.absent_days
            existing_oa.paid_leave_days = payload.opening_attendance.paid_leave_days
            existing_oa.overtime_hours = payload.opening_attendance.overtime_hours
            existing_oa.advance_paid = payload.opening_attendance.advance_paid
            existing_oa.deductions = payload.opening_attendance.deductions
            existing_oa.notes = payload.opening_attendance.notes
        else:
            opening_att = WorkerOpeningAttendance(
                factory_id=str(current_user.factory_id),
                worker_id=worker.id,
                period_start=payload.opening_attendance.period_start,
                period_end=payload.opening_attendance.period_end,
                present_days=payload.opening_attendance.present_days,
                half_days=payload.opening_attendance.half_days,
                absent_days=payload.opening_attendance.absent_days,
                paid_leave_days=payload.opening_attendance.paid_leave_days,
                overtime_hours=payload.opening_attendance.overtime_hours,
                advance_paid=payload.opening_attendance.advance_paid,
                deductions=payload.opening_attendance.deductions,
                notes=payload.opening_attendance.notes,
                created_by_user_id=current_user.id,
            )
            db.add(opening_att)

    _log_onboarding_change(db, int(current_user.factory_id), "Added" if created else "Updated", worker.name)
    db.commit()
    db.refresh(worker)
    background_tasks.add_task(
        log_activity,
        db,
        int(current_user.factory_id),
        current_user.id,
        getattr(current_user, "full_name", None) or current_user.username,
        current_user.role,
        "ONBOARDING_UPDATED",
        f"{'Added' if created else 'Updated'} worker {worker.name}",
        "onboarding",
        worker.id,
        {"entity": "worker"},
    )
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=str(current_user.factory_id),
        sync_type="worker",
        action="insert",
        data=payload,
    )
    return worker


# =============================================================================
# LEVEL 2 ONBOARDING — Machine Configuration
# =============================================================================

@router.post("/step2/machines", response_model=Step2Response)
def onboarding_step2_machines(
    payload: Step2Request,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)
    check_machine_limit(factory_id, db, requested_count=len(payload.machines))

    machine_specs: list[str] = []
    for item in payload.machines:
        seq = item.machine_sequence_number.strip().upper()
        existing = (
            db.query(Machine)
            .filter(Machine.factory_id == factory_id)
            .filter(sql_func.lower(Machine.machine_sequence_number) == seq.lower())
            .first()
        )
        if existing is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Machine sequence {seq} already exists for this factory",
            )

        machine = Machine(
            factory_id=factory_id,
            name=item.name or seq,
            machine_name=item.name or seq,
            machine_type=item.name or "Custom Machine",
            machine_sequence_number=seq,
            cup_size_ml=item.cup_size_ml,
            bottom_size_mm=item.bottom_size_mm,
            speed_cups_per_minute=item.speed_cups_per_minute,
            speed_per_minute=item.speed_cups_per_minute,
            speed_bpm=item.speed_cups_per_minute,
            default_speed=float(item.speed_cups_per_minute or 0),
            can_swap_moulds=item.can_swap_moulds,
        )
        db.add(machine)
        machine_specs.append(machine.name or machine.machine_sequence_number or f"{item.cup_size_ml}ml machine")

    if machine_specs:
        _log_onboarding_change(db, int(factory_id), "Added", ", ".join(machine_specs))
    db.commit()
    background_tasks.add_task(
        log_activity,
        db,
        int(current_user.factory_id),
        current_user.id,
        current_user.full_name or current_user.username,
        current_user.role,
        "ONBOARDING_UPDATED",
        f"Added {len(payload.machines)} machine setup entries",
        "onboarding",
        None,
        {"entity": "machine", "machines": machine_specs},
    )
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=factory_id,
        sync_type="onboarding",
        action="insert",
        data=payload,
    )
    return Step2Response(
        message="Machines configured successfully",
        factory_id=factory_id,
        machines_saved=len(payload.machines),
    )


# =============================================================================
# LEVEL 3 ONBOARDING — Raw Material & Packaging Metrics
# =============================================================================

@router.post("/step3/materials", response_model=Step3Response)
def onboarding_step3_materials(
    payload: Step3Request,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)

    raw_saved = 0
    for item in payload.raw_material_metrics:
        metric = (
            db.query(RawMaterialMetrics)
            .filter(RawMaterialMetrics.factory_id == factory_id)
            .filter(RawMaterialMetrics.material_type == item.material_type)
            .filter(RawMaterialMetrics.size_ml_or_mm == item.size_ml_or_mm)
            .first()
        )
        if metric is None:
            metric = RawMaterialMetrics(
                factory_id=factory_id,
                material_type=item.material_type,
                size_ml_or_mm=item.size_ml_or_mm,
            )
            db.add(metric)

        metric.weight_per_sack_kg = item.weight_per_sack_kg
        metric.pieces_per_sack = item.pieces_per_sack
        raw_saved += 1

    pack_saved = 0
    for item in payload.packaging_metrics:
        variety = getattr(item, "variety", "Standard/White") or "Standard/White"
        metric = (
            db.query(PackagingMetrics)
            .filter(PackagingMetrics.factory_id == factory_id)
            .filter(PackagingMetrics.cup_size_ml == item.cup_size_ml)
            .filter(sql_func.lower(PackagingMetrics.variant_name) == variety.lower())
            .first()
        )
        if metric is None:
            metric = PackagingMetrics(
                factory_id=factory_id,
                cup_size_ml=item.cup_size_ml,
                variant_name=variety,
            )
            db.add(metric)

        metric.kg_per_box = item.kg_per_box
        metric.cups_per_box = item.cups_per_box
        packaging_size_name = f"{item.cup_size_ml}ml Standard Box"
        box_stock = (
            db.query(BoxStock)
            .filter(BoxStock.factory_id == factory_id)
            .filter(sql_func.lower(BoxStock.packaging_size_name) == packaging_size_name.lower())
            .first()
        )
        if box_stock is None:
            db.add(
                BoxStock(
                    factory_id=factory_id,
                    packaging_size_name=packaging_size_name,
                    total_boxes=0,
                )
            )

        final_stock = (
            db.query(FinalProductStock)
            .filter(FinalProductStock.factory_id == factory_id)
            .filter(FinalProductStock.product_size_ml == item.cup_size_ml)
            .filter(sql_func.lower(FinalProductStock.packaging_size_name) == packaging_size_name.lower())
            .first()
        )
        if final_stock is None:
            db.add(
                FinalProductStock(
                    factory_id=factory_id,
                    product_size_ml=item.cup_size_ml,
                    packaging_size_name=packaging_size_name,
                    current_quantity=0,
                    total_boxes=0,
                    loose_packets=0,
                    packets_per_box_limit=item.cups_per_box,
                )
            )
        pack_saved += 1

    db.commit()
    if raw_saved:
        background_tasks.add_task(
            log_activity,
            db=db,
            factory_id=int(current_user.factory_id),
            user_id=current_user.id,
            user_name=current_user.username,
            user_role=current_user.role,
            action_type="RAW_MATERIAL_ADDED",
            action_summary=f"Raw material metrics updated - {raw_saved} entries",
            entity_type="raw_material",
            entity_id=None,
            metadata={"raw_material_metrics_saved": raw_saved},
        )
    if pack_saved:
        background_tasks.add_task(
            log_activity,
            db=db,
            factory_id=int(current_user.factory_id),
            user_id=current_user.id,
            user_name=current_user.username,
            user_role=current_user.role,
            action_type="PACKAGING_MATERIAL_ADDED",
            action_summary=f"Packaging material metrics updated - {pack_saved} entries",
            entity_type="packaging_material",
            entity_id=None,
            metadata={"packaging_metrics_saved": pack_saved},
        )
    background_tasks.add_task(
        sync_data_to_n8n_bg,
        factory_id=factory_id,
        sync_type="onboarding",
        action="insert",
        data=payload,
    )
    return Step3Response(
        message="Material and packaging metrics saved successfully",
        factory_id=factory_id,
        raw_material_metrics_saved=raw_saved,
        packaging_metrics_saved=pack_saved,
    )


# =============================================================================
# LEGACY BULK ONBOARDING (Backward Compatible)
# =============================================================================

@router.post("/complete", response_model=OnboardingCompleteResponse)
def complete_onboarding(
    payload: OnboardingCompleteRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)

    try:
        machine_names = []
        for machine_payload in payload.machines:
            upsert_machine(db, factory_id, machine_payload)
            machine_names.append(machine_payload.name.strip())

        for material_payload in payload.raw_materials:
            material = upsert_raw_material(db, factory_id, material_payload)
            upsert_inventory_from_raw_material(db, factory_id, material)

        for profile_payload in payload.packaging_profiles:
            upsert_packaging_profile(db, factory_id, profile_payload)

        for yield_payload in payload.material_yields:
            upsert_material_yield(db, factory_id, yield_payload)

        upsert_costing_master(db, factory_id, payload.costing_master)

        worker_names = []
        for worker_payload in payload.workers:
            upsert_worker(db, factory_id, worker_payload)
            worker_names.append(worker_payload.name.strip())

        for customer_payload in payload.customers:
            upsert_customer(db, factory_id, customer_payload)

        if machine_names:
            _log_onboarding_change(db, int(factory_id), "Updated", ", ".join(machine_names))
        if worker_names:
            _log_onboarding_change(db, int(factory_id), "Updated", ", ".join(worker_names))
        db.commit()
        if payload.raw_materials:
            background_tasks.add_task(
                log_activity,
                db=db,
                factory_id=int(current_user.factory_id),
                user_id=current_user.id,
                user_name=current_user.username,
                user_role=current_user.role,
                action_type="RAW_MATERIAL_ADDED",
                action_summary=f"Raw materials updated during onboarding - {len(payload.raw_materials)} entries",
                entity_type="raw_material",
                entity_id=None,
                metadata={"raw_materials_saved": len(payload.raw_materials)},
            )
        if payload.packaging_profiles:
            background_tasks.add_task(
                log_activity,
                db=db,
                factory_id=int(current_user.factory_id),
                user_id=current_user.id,
                user_name=current_user.username,
                user_role=current_user.role,
                action_type="PACKAGING_MATERIAL_ADDED",
                action_summary=f"Packaging profiles updated during onboarding - {len(payload.packaging_profiles)} entries",
                entity_type="packaging_material",
                entity_id=None,
                metadata={"packaging_profiles_saved": len(payload.packaging_profiles)},
            )
        background_tasks.add_task(
            sync_data_to_n8n_bg,
            factory_id=factory_id,
            sync_type="onboarding",
            action="insert",
            data=payload,
        )
        return OnboardingCompleteResponse(
            message="Factory onboarding completed successfully",
            factory_id=factory_id,
            machines_saved=len(payload.machines),
            raw_materials_saved=len(payload.raw_materials),
            packaging_profiles_saved=len(payload.packaging_profiles),
            material_yields_saved=len(payload.material_yields),
            workers_saved=len(payload.workers),
            customers_saved=len(payload.customers),
        )
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Onboarding failed and was rolled back: {exc}",
        ) from exc


# ---------------------------------------------------------------------------
# Legacy Upsert Helpers
# ---------------------------------------------------------------------------

def upsert_machine(db: Session, factory_id: int, payload: MachinePayload) -> Machine:
    name = payload.name.strip()
    machine = (
        db.query(Machine)
        .filter(Machine.factory_id == factory_id)
        .filter(sql_func.lower(Machine.name) == name.lower())
        .first()
    )
    if machine is None:
        machine = Machine(factory_id=factory_id, name=name)
        db.add(machine)

    machine.speed_bpm = payload.speed_bpm
    machine.speed_per_minute = payload.speed_bpm
    machine.current_mould_size = payload.current_mould_size
    machine.default_mould_size = payload.current_mould_size
    machine.current_bottom_size = payload.current_bottom_size
    machine.bottom_size = payload.current_bottom_size
    machine.can_swap_moulds = payload.can_swap_moulds
    return machine


def upsert_raw_material(db: Session, factory_id: int, payload: RawMaterialPayload) -> RawMaterial:
    material_name = payload.name or build_raw_material_name(payload)
    material_type = payload.type
    unit = payload.unit or ("pieces" if material_type in {"Polybag", "Carton Box"} else "kg")
    size_name = f"{payload.size_ml}ml" if payload.size_ml else None

    raw_material = (
        db.query(RawMaterial)
        .filter(RawMaterial.factory_id == factory_id)
        .filter(sql_func.lower(RawMaterial.name) == material_name.lower())
        .filter(RawMaterial.material_type == material_type)
        .filter(RawMaterial.size_name == size_name)
        .first()
    )
    if raw_material is None:
        raw_material = RawMaterial(
            factory_id=factory_id,
            name=material_name,
            material_type=material_type,
            size_name=size_name,
        )
        db.add(raw_material)

    raw_material.type = material_type
    raw_material.size_ml = payload.size_ml
    raw_material.gsm = payload.gsm
    raw_material.unit = unit
    raw_material.opening_stock = payload.stock_quantity
    raw_material.current_stock = payload.stock_quantity
    raw_material.stock_quantity = payload.stock_quantity
    raw_material.price_per_unit = payload.price_per_unit
    return raw_material


def build_raw_material_name(payload: RawMaterialPayload) -> str:
    parts = [payload.type]
    if payload.size_ml:
        parts.append(f"{payload.size_ml}ml")
    if payload.gsm:
        parts.append(f"{payload.gsm}gsm")
    return " ".join(parts)


def upsert_inventory_from_raw_material(db: Session, factory_id: int, raw_material: RawMaterial) -> Inventory:
    category = "Packaging" if raw_material.material_type in {"Polybag", "Carton Box"} else "Raw"
    item = get_or_create_inventory(db, factory_id, raw_material.name, category, raw_material.unit)
    item.quantity = raw_material.stock_quantity
    item.price_per_unit = raw_material.price_per_unit
    return item


def upsert_packaging_profile(
    db: Session,
    factory_id: int,
    payload: PackagingProfilePayload,
) -> PackagingProfile:
    product_name = f"{payload.product_name_ml}ml Paper Cup"
    profile_name = f"{payload.product_name_ml}ml Standard Box"
    poly_inventory = get_or_create_inventory(db, factory_id, f"{payload.product_name_ml}ml Polybag", "Packaging", "pieces")
    box_inventory = get_or_create_inventory(db, factory_id, payload.box_size_name or f"{payload.product_name_ml}ml Carton Box", "Packaging", "pieces")

    profile = (
        db.query(PackagingProfile)
        .filter(PackagingProfile.factory_id == factory_id)
        .filter(sql_func.lower(PackagingProfile.profile_name) == profile_name.lower())
        .first()
    )
    if profile is None:
        profile = PackagingProfile(factory_id=factory_id, profile_name=profile_name)
        db.add(profile)

    cups_per_box = payload.cups_per_polybag * payload.polybags_per_box
    profile.product_name = product_name
    profile.product_name_ml = payload.product_name_ml
    profile.cup_size_ml = payload.product_name_ml
    profile.polybag_capacity = payload.cups_per_polybag
    profile.box_capacity = cups_per_box
    profile.box_size_name = payload.box_size_name
    profile.cups_per_poly = payload.cups_per_polybag
    profile.cups_per_polybag = payload.cups_per_polybag
    profile.polys_per_box = payload.polybags_per_box
    profile.polybags_per_box = payload.polybags_per_box
    profile.box_inventory_id = box_inventory.id
    profile.poly_inventory_id = poly_inventory.id
    return profile


def upsert_material_yield(db: Session, factory_id: int, payload: MaterialYieldPayload) -> MaterialYield:
    material_yield = (
        db.query(MaterialYield)
        .filter(MaterialYield.factory_id == factory_id)
        .filter(MaterialYield.material_type == payload.material_type)
        .filter(MaterialYield.size_ml == payload.size_ml)
        .filter(MaterialYield.gsm == payload.gsm)
        .first()
    )
    if material_yield is None:
        material_yield = MaterialYield(
            factory_id=factory_id,
            material_type=payload.material_type,
            size_ml=payload.size_ml,
            gsm=payload.gsm,
        )
        db.add(material_yield)

    material_yield.pieces_per_kg = payload.pieces_per_kg
    return material_yield


def upsert_costing_master(db: Session, factory_id: int, payload) -> CostingMaster:
    costing = db.query(CostingMaster).filter(CostingMaster.factory_id == factory_id).first()
    if costing is None:
        costing = CostingMaster(factory_id=factory_id)
        db.add(costing)

    costing.paper_price_per_kg = payload.paper_price_per_kg
    costing.bottom_roll_price_per_kg = payload.bottom_roll_price_per_kg
    costing.polybag_price = payload.polybag_price
    costing.carton_price = payload.carton_price
    costing.labour_cost_per_box = payload.labour_cost_per_box
    costing.electricity_cost_per_box = payload.electricity_cost_per_box
    return costing


def upsert_worker(db: Session, factory_id: int, payload: WorkerPayload) -> Worker:
    worker = (
        db.query(Worker)
        .filter(Worker.factory_id == factory_id)
        .filter(sql_func.lower(Worker.name) == payload.name.lower())
        .first()
    )
    if worker is None:
        worker = Worker(factory_id=factory_id, name=payload.name.strip())
        db.add(worker)

    worker.is_active = True
    worker.daily_salary = payload.daily_salary
    worker.salary = payload.daily_salary
    worker.shift_type = payload.shift_type
    worker.shift_timing = payload.shift_type
    return worker


def upsert_customer(db: Session, factory_id: int, payload: CustomerPayload) -> Customer:
    customer_name = (payload.name or payload.firm_name).strip()
    customer = (
        db.query(Customer)
        .filter(Customer.factory_id == factory_id)
        .filter(sql_func.lower(Customer.name) == customer_name.lower())
        .first()
    )
    if customer is None:
        customer = Customer(factory_id=factory_id, name=customer_name)
        db.add(customer)

    customer.firm_name = payload.firm_name.strip()
    customer.contact_number = payload.contact_number
    customer.pending_balance = payload.pending_balance
    customer.pending_dues = float(payload.pending_balance)
    customer.balance_amount = payload.pending_balance
    return customer


def get_or_create_inventory(db: Session, factory_id: int, item_name: str, category: str, unit: str) -> Inventory:
    db.flush()
    item = (
        db.query(Inventory)
        .filter(Inventory.factory_id == factory_id)
        .filter(sql_func.lower(Inventory.item_name) == item_name.lower())
        .first()
    )
    if item is None:
        item = Inventory(
            factory_id=factory_id,
            item_name=item_name.strip(),
            category=category,
            unit=unit,
            quantity=Decimal("0.000"),
            price_per_unit=Decimal("0.00"),
        )
        db.add(item)
        db.flush()
    return item


@router.delete("/entry/{entry_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_onboarding_entry(
    entry_id: str,
    background_tasks: BackgroundTasks,
    type: Optional[str] = None,
    current_user: User = Depends(require_owner_delete),
    db: Session = Depends(get_db),
):
    factory_id = str(current_user.factory_id)
    
    # Try to parse string ID containing prefix
    try:
        actual_id = int(entry_id)
    except ValueError:
        parts = entry_id.split("-")
        if len(parts) == 2:
            type = parts[0]
            try:
                actual_id = int(parts[1])
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid entry_id format")
        else:
            raise HTTPException(status_code=400, detail="Invalid entry_id format")

    if not type:
        type = "final"
        
    type_lower = type.lower()
    
    def trigger_delete_sync():
        background_tasks.add_task(
            sync_data_to_n8n_bg,
            factory_id=str(factory_id),
            sync_type="onboarding",
            action="delete",
            data={"entry_id": actual_id, "type": type_lower}
        )
    
    try:
        if type_lower in ("final", "final product", "cups", "final-product"):
            entry = db.query(FinalProductStock).filter(FinalProductStock.id == actual_id, FinalProductStock.factory_id == factory_id).first()
            if entry:
                product_size_ml = entry.product_size_ml
                variety = entry.variety
                packaging_size_name = entry.packaging_size_name
                
                db.delete(entry)
                db.flush()
                
                # Recalculate dynamic live stock balance and sync caches
                from routers.inventory import recalculate_and_sync_sku_stock
                recalculate_and_sync_sku_stock(
                    db=db,
                    factory_id=factory_id,
                    product_size_ml=product_size_ml,
                    variety=variety,
                    packaging_size_name=packaging_size_name,
                )
                db.commit()
                trigger_delete_sync()
                return None
        elif type_lower in ("blank", "blankstock"):
            entry = db.query(BlankStock).filter(BlankStock.id == actual_id, BlankStock.factory_id == factory_id).first()
            if entry:
                db.delete(entry)
                db.commit()
                trigger_delete_sync()
                return None
        elif type_lower in ("bottom", "bottomstock"):
            entry = db.query(BottomStock).filter(BottomStock.id == actual_id, BottomStock.factory_id == factory_id).first()
            if entry:
                db.delete(entry)
                db.commit()
                trigger_delete_sync()
                return None
        elif type_lower in ("box", "boxstock", "carton box", "carton", "packaging", "inventory"):
            entry = db.query(BoxStock).filter(BoxStock.id == actual_id, BoxStock.factory_id == factory_id).first()
            if entry:
                db.delete(entry)
                db.commit()
                trigger_delete_sync()
                return None
            inv_entry = db.query(Inventory).filter(Inventory.id == actual_id, Inventory.factory_id == factory_id).first()
            if inv_entry:
                profiles = db.query(PackagingProfile).filter(
                    (PackagingProfile.box_inventory_id == actual_id) |
                    (PackagingProfile.poly_inventory_id == actual_id)
                ).all()

                from models import DailyProduction
                has_production_links = False
                for profile in profiles:
                    linked_log = db.query(DailyProduction).filter(
                        DailyProduction.factory_id == factory_id,
                        sql_func.lower(DailyProduction.packaging_size_name) == profile.profile_name.lower(),
                    ).first()
                    if linked_log:
                        has_production_links = True
                        break

                if has_production_links:
                    inv_entry.item_name = f"[DELETED] {inv_entry.item_name}"
                    db.add(inv_entry)
                else:
                    for profile in profiles:
                        db.query(FinishedGoodsStock).filter(FinishedGoodsStock.packaging_profile_id == profile.id).delete()
                        db.delete(profile)
                    db.delete(inv_entry)
                db.commit()
                trigger_delete_sync()
                return None
        elif type_lower in ("plastic", "plasticstock", "polybag"):
            entry = db.query(PlasticStock).filter(PlasticStock.id == actual_id, PlasticStock.factory_id == factory_id).first()
            if entry:
                db.delete(entry)
                db.commit()
                trigger_delete_sync()
                return None
        elif type_lower in ("polybag", "polybagstock"):
            entry = db.query(PolybagStock).filter(PolybagStock.id == actual_id, PolybagStock.factory_id == factory_id).first()
            if entry:
                db.delete(entry)
                db.commit()
                trigger_delete_sync()
                return None
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete entry: {exc}"
        )
        
    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Onboarding entry of type '{type}' with ID {actual_id} not found")


@router.delete("/v1/onboarding/items/{item_id}")
@v1_router.delete("/items/{item_id}")
def delete_onboarding_item(
    item_id: int,
    type: str,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_owner_delete),
    db: Session = Depends(get_db)
):
    factory_id = str(current_user.factory_id)
    type_lower = type.lower()
    
    try:
        if type_lower in ("blank", "blankstock"):
            entry = db.query(BlankStock).filter(BlankStock.id == item_id, BlankStock.factory_id == factory_id).first()
            if not entry:
                raise HTTPException(status_code=404, detail="Blank stock not found")
            db.delete(entry)
            db.commit()
            
        elif type_lower in ("bottom", "bottomstock"):
            entry = db.query(BottomStock).filter(BottomStock.id == item_id, BottomStock.factory_id == factory_id).first()
            if not entry:
                raise HTTPException(status_code=404, detail="Bottom stock not found")
            db.delete(entry)
            db.commit()
            
        elif type_lower in ("box", "boxstock", "carton box", "carton", "packaging", "inventory"):
            # Check BoxStock first
            entry = db.query(BoxStock).filter(BoxStock.id == item_id, BoxStock.factory_id == factory_id).first()
            if entry:
                db.delete(entry)
                db.commit()
            else:
                # Check standard Inventory table (Corrugated Box variants)
                inv_entry = db.query(Inventory).filter(Inventory.id == item_id, Inventory.factory_id == factory_id).first()
                if not inv_entry:
                    raise HTTPException(status_code=404, detail="Item not found")
                
                # Deletion constraints: override any hardcoded constraints blocking "Standard" boxes 
                # Check for dependencies in PackagingProfile
                profiles = db.query(PackagingProfile).filter(
                    (PackagingProfile.box_inventory_id == item_id) |
                    (PackagingProfile.poly_inventory_id == item_id)
                ).all()
                
                can_hard_delete = True
                from models import DailyProduction
                for p in profiles:
                    # check if active production logs exist
                    has_logs = db.query(DailyProduction).filter(
                        (DailyProduction.factory_id == factory_id) &
                        (sql_func.lower(DailyProduction.packaging_size_name) == p.profile_name.lower())
                    ).first()
                    if has_logs:
                        can_hard_delete = False
                        break
                
                if can_hard_delete:
                    for p in profiles:
                        db.query(FinishedGoodsStock).filter(FinishedGoodsStock.packaging_profile_id == p.id).delete()
                        db.delete(p)
                    db.flush()
                    
                    # Direct SQL Deletion against the primary table identifier string
                    from sqlalchemy import text
                    db.execute(
                        text("DELETE FROM inventory WHERE factory_id = :factory_id AND id = :item_id"),
                        {"factory_id": factory_id, "item_id": item_id}
                    )
                    db.commit()
                else:
                    # Apply soft-delete status by renaming mapping
                    inv_entry.item_name = f"[DELETED] {inv_entry.item_name}"
                    db.commit()
            
        elif type_lower in ("plastic", "plasticstock", "polybag"):
            entry = db.query(PlasticStock).filter(PlasticStock.id == item_id, PlasticStock.factory_id == factory_id).first()
            if not entry:
                raise HTTPException(status_code=404, detail="Plastic stock not found")
            db.delete(entry)
            db.commit()
            
        elif type_lower in ("polybag", "polybagstock"):
            entry = db.query(PolybagStock).filter(PolybagStock.id == item_id, PolybagStock.factory_id == factory_id).first()
            if not entry:
                raise HTTPException(status_code=404, detail="Polybag stock not found")
            db.delete(entry)
            db.commit()
            
        else:
            raise HTTPException(status_code=400, detail="Invalid stock type")
            
        background_tasks.add_task(
            sync_data_to_n8n_bg,
            factory_id=str(factory_id),
            sync_type="onboarding",
            action="delete",
            data={"entry_id": item_id, "type": type_lower}
        )
        return {"message": "Item deleted successfully"}
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot delete item because it contains active data associations: {exc}"
        )
