from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import hashlib
import json
import logging
from pathlib import Path
import subprocess
import os
import traceback
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session
from sqlalchemy.engine import make_url

from db import DATABASE_URL
from models import (
    BillPayment, BlankStock, BottomStock, BoxStock, Customer, CustomerLedgerAdjustment,
    DailyProduction, Employee, ExpenseLog, Factory, FactoryExpense, FinalProductStock,
    InvoiceDocument, Machine, OutstandingBill, PaymentCollection, PlasticStock, Supplier,
    WastageLog, Worker,
)


API_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BACKUP_ROOT = (
    API_ROOT.parent.parent / "storage" / "backups"
    if API_ROOT.name == "api" and API_ROOT.parent.name == "apps"
    else API_ROOT / "storage" / "backups"
)
BACKUP_ROOT = Path(os.getenv("BACKUP_ROOT", str(DEFAULT_BACKUP_ROOT))).expanduser()
STAGING_ROOT = BACKUP_ROOT / "restore-staging"
META_SHEET = "Backup Metadata"
logger = logging.getLogger(__name__)


class PreRestoreBackupError(RuntimeError):
    """Raised when the mandatory database safety backup cannot be created."""


def _configured_database_url():
    database_url = (
        os.getenv("DATABASE_URL")
        or os.getenv("SQLALCHEMY_DATABASE_URL")
        or DATABASE_URL
    )
    return make_url(database_url)


SHEETS = {
    "Factory Profile": Factory,
    "Customers": Customer,
    "Customer Opening Outstanding": OutstandingBill,
    "Customer Ledger Adjustments": CustomerLedgerAdjustment,
    "Invoices": InvoiceDocument,
    "Invoice Payments": PaymentCollection,
    "Payment Allocations": BillPayment,
    "Outstanding Bills": OutstandingBill,
    "Raw Materials": BlankStock,
    "Paper Stock": BlankStock,
    "Plastic Stock": PlasticStock,
    "Bottom Stock": BottomStock,
    "Box Stock": BoxStock,
    "Finished Goods Stock": FinalProductStock,
    "Daily Production": DailyProduction,
    "Wastage": WastageLog,
    "Expenses": ExpenseLog,
    "Factory Expenses": FactoryExpense,
    "Workers": Worker,
    "Employees": Employee,
    "Machines": Machine,
    "Suppliers": Supplier,
}

RESTORE_ORDER = [
    "Customers", "Workers", "Employees", "Machines", "Suppliers", "Raw Materials",
    "Plastic Stock", "Bottom Stock", "Box Stock", "Finished Goods Stock", "Invoices",
    "Outstanding Bills", "Customer Opening Outstanding", "Customer Ledger Adjustments",
    "Invoice Payments", "Payment Allocations", "Daily Production", "Wastage",
    "Expenses", "Factory Expenses",
]

NATURAL_KEYS = {
    "Customers": ("phone_number", "name"),
    "Workers": ("phone", "name"),
    "Employees": ("name", "role"),
    "Machines": ("machine_number", "name"),
    "Suppliers": ("name",),
    "Raw Materials": ("blank_size_ml", "variety"),
    "Paper Stock": ("blank_size_ml", "variety"),
    "Plastic Stock": ("plastic_size_name", "cup_size_ml"),
    "Bottom Stock": ("bottom_size_mm", "variety"),
    "Box Stock": ("packaging_size_name",),
    "Finished Goods Stock": ("product_size_ml", "variety", "packaging_size_name"),
    "Invoices": ("invoice_number",),
    "Outstanding Bills": ("tracking_number",),
    "Customer Opening Outstanding": ("tracking_number",),
    "Customer Ledger Adjustments": ("customer_id", "adjustment_type", "amount", "reason", "created_at"),
    "Invoice Payments": ("customer_id", "outstanding_bill_id", "amount_collected", "collection_date", "reference_number"),
    "Payment Allocations": ("bill_id", "amount_allocated", "payment_date", "created_at"),
    "Daily Production": ("date", "worker_id", "machine_id", "product_size_ml", "variety", "packaging_size_name", "created_at"),
    "Wastage": ("date", "wastage_weight", "created_at"),
    "Expenses": ("date", "category", "description", "amount"),
    "Factory Expenses": ("expense_name", "amount", "category", "timestamp"),
}


def _jsonable(value: Any):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True, default=str)
    return value


def _restore_key(sheet: str, row: Any) -> str:
    keys = NATURAL_KEYS.get(sheet, ("id",))
    raw = "|".join(str(getattr(row, key, "") or "") for key in keys)
    return hashlib.sha256(f"{sheet}|{raw}".encode()).hexdigest()[:32]


def _rows_for_sheet(db: Session, factory_id: int, sheet: str, model):
    if sheet == "Factory Profile":
        return db.query(model).filter(model.id == factory_id).all()
    query = db.query(model).filter(model.factory_id == factory_id)
    if sheet == "Customer Opening Outstanding":
        query = query.filter(OutstandingBill.source_type.in_(("opening_outstanding", "opening_balance")))
    elif sheet == "Outstanding Bills":
        query = query.filter(~OutstandingBill.source_type.in_(("opening_outstanding", "opening_balance")))
    return query.all()


def build_master_backup(db: Session, factory_id: int) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    factory = db.query(Factory).filter(Factory.id == factory_id).one()
    meta = workbook.create_sheet(META_SHEET)
    meta.append(["backup_version", "factory_id", "factory_name", "generated_at"])
    meta.append(["1", factory_id, factory.factory_name or factory.name, datetime.utcnow().isoformat()])

    invoice_items = workbook.create_sheet("Invoice Items")
    invoice_items.append(["invoice_restore_key", "item_index", "item_json"])

    for sheet, model in SHEETS.items():
        ws = workbook.create_sheet(sheet)
        columns = [column.name for column in model.__table__.columns]
        ws.append(["restore_key", *columns])
        for row in _rows_for_sheet(db, factory_id, sheet, model):
            key = _restore_key(sheet, row)
            ws.append([key, *[_jsonable(getattr(row, column)) for column in columns]])
            if sheet == "Invoices":
                items = (row.payload_json or {}).get("items") or []
                for index, item in enumerate(items):
                    invoice_items.append([key, index, json.dumps(item, ensure_ascii=True, default=str)])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _sheet_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]


def validate_backup(file_bytes: bytes, expected_factory_id: int) -> dict:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return {"fatal": True, "errors": [{"sheet": "Workbook", "error": f"Corrupt Excel: {exc}"}]}
    errors = []
    if META_SHEET not in workbook.sheetnames:
        errors.append({"sheet": META_SHEET, "error": "Backup metadata sheet is missing"})
        return {"fatal": True, "errors": errors}
    metadata = _sheet_records(workbook[META_SHEET])
    source_factory_id = int(metadata[0].get("factory_id") or 0) if metadata else 0
    if source_factory_id != expected_factory_id:
        errors.append({"sheet": META_SHEET, "error": "Cross-factory restore is not allowed"})
    missing = [sheet for sheet in SHEETS if sheet not in workbook.sheetnames]
    errors.extend({"sheet": sheet, "error": "Required sheet is missing"} for sheet in missing)
    counts = {sheet: len(_sheet_records(workbook[sheet])) for sheet in workbook.sheetnames if sheet != META_SHEET}
    return {"fatal": bool(errors), "errors": errors, "source_factory_id": source_factory_id, "sheet_counts": counts}


class RestoreFailure(RuntimeError):
    def __init__(self, message: str, *, sheet: str | None = None, table: str | None = None):
        super().__init__(message)
        self.sheet = sheet
        self.table = table

    @property
    def detail(self) -> str:
        context = self.sheet or self.table
        return f"Restore failed in {context}: {self}" if context else str(self)


def stage_backup(file_bytes: bytes, factory_id: int, filename: str | None = None) -> tuple[str, dict]:
    report = validate_backup(file_bytes, factory_id)
    restore_id = uuid4().hex
    STAGING_ROOT.mkdir(parents=True, exist_ok=True)
    (STAGING_ROOT / f"{factory_id}_{restore_id}.xlsx").write_bytes(file_bytes)
    staged_backup_metadata_path(factory_id, restore_id).write_text(
        json.dumps({
            "restore_session_id": restore_id,
            "factory_id": factory_id,
            "filename": filename or "master_backup.xlsx",
            "validation_status": "VALID" if not report["fatal"] else "FAILED",
            "fatal_count": len(report["errors"]),
            "sheet_counts": report.get("sheet_counts", {}),
        }, ensure_ascii=True),
        encoding="utf-8",
    )
    return restore_id, report


def staged_backup_path(factory_id: int, restore_id: str) -> Path:
    return STAGING_ROOT / f"{factory_id}_{restore_id}.xlsx"


def staged_backup_metadata_path(factory_id: int, restore_id: str) -> Path:
    return STAGING_ROOT / f"{factory_id}_{restore_id}.json"


def build_validation_report(file_bytes: bytes, factory_id: int) -> BytesIO:
    report = validate_backup(file_bytes, factory_id)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Validation Summary"
    summary.append(["status", "fatal_count", "source_factory_id"])
    summary.append([
        "FAILED" if report["fatal"] else "VALID",
        len(report["errors"]),
        report.get("source_factory_id"),
    ])
    errors = workbook.create_sheet("Validation Errors")
    errors.append(["sheet", "row", "column", "bad_value", "error", "correction"])
    for issue in report["errors"]:
        errors.append([
            issue.get("sheet"),
            issue.get("row"),
            issue.get("column"),
            issue.get("bad_value"),
            issue.get("error"),
            issue.get("correction") or "Correct the source sheet and validate again.",
        ])
    counts = workbook.create_sheet("Sheet Counts")
    counts.append(["sheet", "rows"])
    for sheet, count in report.get("sheet_counts", {}).items():
        counts.append([sheet, count])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def preview_backup(db: Session, file_bytes: bytes, factory_id: int) -> dict:
    report = validate_backup(file_bytes, factory_id)
    preview = {"new": {}, "existing": {}, "updated": {}}
    if report["fatal"]:
        return preview
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    for sheet in RESTORE_ORDER:
        if sheet not in workbook.sheetnames:
            continue
        model = SHEETS[sheet]
        records = _sheet_records(workbook[sheet])
        existing_count = sum(1 for record in records if _find_existing(db, model, factory_id, sheet, record))
        preview["existing"][sheet] = existing_count
        preview["updated"][sheet] = existing_count
        preview["new"][sheet] = len(records) - existing_count
    return preview


def _coerce(column, value):
    if value in (None, ""):
        return None
    try:
        python_type = column.type.python_type
    except NotImplementedError:
        return value
    if python_type is dict:
        return json.loads(value) if isinstance(value, str) else value
    if python_type is datetime and isinstance(value, str):
        return datetime.fromisoformat(value)
    if python_type is date and isinstance(value, str):
        return date.fromisoformat(value)
    if python_type is bool and isinstance(value, str):
        return value.lower() in {"true", "1", "yes"}
    return python_type(value)


def _find_existing(db: Session, model, factory_id: int, sheet: str, record: dict):
    query = db.query(model).filter(model.factory_id == factory_id)
    for key in NATURAL_KEYS.get(sheet, ("id",)):
        value = record.get(key)
        if value not in (None, "") and hasattr(model, key):
            query = query.filter(getattr(model, key) == value)
    return query.first()


def create_pre_restore_backup(db: Session, factory_id: int) -> Path:
    try:
        BACKUP_ROOT.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        message = f"Backup directory is not writable: {exc}"
        logger.exception("pre_restore_backup_directory_failed backup_root=%s", BACKUP_ROOT)
        raise PreRestoreBackupError(message) from exc
    bind = db.get_bind()
    if bind.dialect.name == "postgresql":
        path = BACKUP_ROOT / f"pre_restore_factory_{factory_id}_{datetime.utcnow():%Y%m%d_%H%M%S}.dump"
        url = _configured_database_url()
        if not all((url.host, url.username, url.password, url.database)):
            raise PreRestoreBackupError(
                "Pre-restore PostgreSQL backup failed: configured database URL must include host, username, password, and database"
            )
        env = os.environ.copy()
        env["PGPASSWORD"] = url.password
        command = ["pg_dump", "-Fc", "--file", str(path)]
        command.extend(["--host", url.host])
        command.extend(["--port", str(url.port or 5432)])
        command.extend(["--username", url.username])
        command.append(url.database)
        try:
            subprocess.run(
                command,
                check=True,
                capture_output=True,
                text=True,
                timeout=300,
                env=env,
            )
        except FileNotFoundError as exc:
            path.unlink(missing_ok=True)
            message = "Pre-restore PostgreSQL backup failed: pg_dump executable was not found"
            logger.exception("pg_dump executable not found command=%s", command[0])
            raise PreRestoreBackupError(message) from exc
        except subprocess.CalledProcessError as exc:
            path.unlink(missing_ok=True)
            stderr = (exc.stderr or "").strip()
            message = stderr or f"pg_dump exited with status {exc.returncode}"
            logger.exception("pg_dump failed stderr=%s", message)
            raise PreRestoreBackupError(
                f"Pre-restore PostgreSQL backup failed: {message}"
            ) from exc
        except (OSError, subprocess.TimeoutExpired) as exc:
            path.unlink(missing_ok=True)
            message = str(exc).strip() or exc.__class__.__name__
            logger.exception("pg_dump could not create backup path=%s", path)
            raise PreRestoreBackupError(
                f"Pre-restore PostgreSQL backup failed: {message}"
            ) from exc
        if not path.is_file() or path.stat().st_size == 0:
            path.unlink(missing_ok=True)
            message = "pg_dump completed without creating a backup file"
            logger.error("pg_dump output missing or empty path=%s", path)
            raise PreRestoreBackupError(
                f"Pre-restore PostgreSQL backup failed: {message}"
            )
        return path
    path = BACKUP_ROOT / f"pre_restore_factory_{factory_id}_{datetime.utcnow():%Y%m%d_%H%M%S}.xlsx"
    path.write_bytes(build_master_backup(db, factory_id).getvalue())
    return path


def restore_staged_backup(db: Session, factory_id: int, restore_id: str) -> dict:
    path = staged_backup_path(factory_id, restore_id)
    metadata_path = staged_backup_metadata_path(factory_id, restore_id)
    if not path.exists():
        raise ValueError("Restore upload has expired or does not exist")
    if not metadata_path.exists():
        raise ValueError("Validated restore session metadata is missing; validate the file again")
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    if metadata.get("validation_status") != "VALID" or int(metadata.get("fatal_count") or 0) != 0:
        raise ValueError("Restore session did not pass validation")
    file_bytes = path.read_bytes()
    report = validate_backup(file_bytes, factory_id)
    if report["fatal"]:
        raise ValueError("Backup contains fatal validation errors")
    filename = str(metadata.get("filename") or path.name)
    logger.info(
        "master_restore_start restore_session_id=%s filename=%s factory_id=%s sheet_counts=%s",
        restore_id, filename, factory_id, report.get("sheet_counts", {}),
    )
    try:
        backup_path = create_pre_restore_backup(db, factory_id)
    except Exception as exc:
        logger.exception(
            "master_restore_pre_backup_failed restore_session_id=%s filename=%s factory_id=%s",
            restore_id, filename, factory_id,
        )
        raise RestoreFailure("Database safety backup could not be created; no data was changed") from exc
    logger.info(
        "master_restore_pre_backup_created restore_session_id=%s backup_path=%s",
        restore_id, backup_path,
    )
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    result = {"inserted": 0, "updated": 0, "deleted": 0, "skipped": 0}
    id_maps: dict[str, dict[int, int]] = {}
    restored_ids: dict[str, set[int]] = {}
    current_sheet: str | None = None
    current_table: str | None = None

    try:
        for sheet in RESTORE_ORDER:
            if sheet not in workbook.sheetnames:
                continue
            model = SHEETS[sheet]
            table_name = model.__tablename__
            current_sheet = sheet
            current_table = table_name
            id_maps.setdefault(table_name, {})
            restored_ids.setdefault(table_name, set())
            records = _sheet_records(workbook[sheet])
            sheet_created = 0
            sheet_updated = 0
            logger.info(
                "master_restore_sheet_start restore_session_id=%s sheet=%s table=%s rows_parsed=%s",
                restore_id, sheet, table_name, len(records),
            )
            for record in records:
                source_id = int(record.get("id") or 0)
                normalized_record = dict(record)
                for column in model.__table__.columns:
                    value = normalized_record.get(column.name)
                    if not column.name.endswith("_id") or value in (None, ""):
                        continue
                    reference_table = next(iter(column.foreign_keys)).column.table.name if column.foreign_keys else None
                    if reference_table and int(value) in id_maps.get(reference_table, {}):
                        normalized_record[column.name] = id_maps[reference_table][int(value)]
                    elif reference_table in {"users", "orders", "payments"}:
                        normalized_record[column.name] = None
                existing = _find_existing(db, model, factory_id, sheet, normalized_record)
                target = existing or model()
                for column in model.__table__.columns:
                    if column.name in {"id", "factory_id"} or column.name not in normalized_record:
                        continue
                    setattr(target, column.name, _coerce(column, normalized_record[column.name]))
                target.factory_id = factory_id
                if not existing:
                    db.add(target)
                    db.flush()
                    result["inserted"] += 1
                    sheet_created += 1
                else:
                    result["updated"] += 1
                    sheet_updated += 1
                restored_ids[table_name].add(int(target.id))
                if source_id:
                    id_maps[table_name][source_id] = int(target.id)
            logger.info(
                "master_restore_sheet_complete restore_session_id=%s sheet=%s table=%s "
                "created=%s updated=%s",
                restore_id, sheet, table_name, sheet_created, sheet_updated,
            )

        cleaned_tables: set[str] = set()
        for sheet in reversed(RESTORE_ORDER):
            model = SHEETS[sheet]
            table_name = model.__tablename__
            if table_name in cleaned_tables:
                continue
            cleaned_tables.add(table_name)
            current_sheet = sheet
            current_table = table_name
            query = db.query(model).filter(model.factory_id == factory_id)
            keep_ids = restored_ids.get(table_name, set())
            if keep_ids:
                query = query.filter(~model.id.in_(keep_ids))
            deleted = query.delete(synchronize_session=False)
            result["deleted"] += deleted
            logger.info(
                "master_restore_snapshot_cleanup restore_session_id=%s sheet=%s table=%s deleted=%s",
                restore_id, sheet, table_name, deleted,
            )

        db.flush()
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error(
            "master_restore_failed restore_session_id=%s filename=%s sheet=%s table=%s traceback=%s",
            restore_id, filename, current_sheet, current_table, traceback.format_exc(),
        )
        raise RestoreFailure(
            "A workbook row or snapshot cleanup operation could not be applied. "
            "Review this sheet's values and retry validation.",
            sheet=current_sheet,
            table=current_table,
        ) from exc
    path.unlink(missing_ok=True)
    metadata_path.unlink(missing_ok=True)
    logger.info(
        "master_restore_complete restore_session_id=%s filename=%s factory_id=%s "
        "created=%s updated=%s deleted=%s",
        restore_id, filename, factory_id,
        result["inserted"], result["updated"], result["deleted"],
    )
    return result
